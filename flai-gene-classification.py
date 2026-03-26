#!/usr/bin/env python3
"""
Fly Gene Classification + Reagent Finder Pipeline (Standalone)

Usage:
    python flai-gene-classification.py <input_directory> [--keywords "kw1,kw2,..."] [--reference-limit N]

This script processes each CSV file in the specified directory, reads gene symbols from the
configured input column, converts them to FlyBase IDs, gathers literature evidence, classifies
genes against the supplied keywords, and extracts gene-linked reagent records from supporting papers.

Output:
    For each input file <name>.csv, creates <name>_classification.xlsx in the same directory.
"""

import os
import sys
import argparse
import requests
import time
import random
import re
import io
import subprocess
import json
import hashlib
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import pandas as pd
from collections import defaultdict
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from functools import lru_cache
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
# Prefer repo-local .env for batch jobs, then fall back to default dotenv search.
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")
load_dotenv()

# Detect turbo mount: macOS uses /Volumes/umms-rallada, ARC uses /nfs/turbo/umms-rallada
_TURBO_CANDIDATES = [
    Path("/Volumes/umms-rallada"),       # macOS
    Path("/nfs/turbo/umms-rallada"),      # ARC-HPC (Linux)
]
TURBO_ROOT = next((p for p in _TURBO_CANDIDATES if p.exists()), _TURBO_CANDIDATES[0])

# API Keys
NCBI_API_KEY = os.getenv("NCBI_API_KEY")
UNPAYWALL_TOKEN = os.getenv("UNPAYWALL_TOKEN") or "aadish98@gmail.com"
# Shared turbo-server PubMed title/abstract cache used by fly_stocker_v2
PUBMED_CACHE_PATH = TURBO_ROOT / "UM Lab Users" / "Aadish" / "Data" / "PubMed Cache" / "pmid_to_title_abstract.csv"
# Shared turbo-server cache used by fly_stocker_v2
FULLTEXT_METHOD_CACHE_PATH = TURBO_ROOT / "UM Lab Users" / "Aadish" / "Data" / "PubMed Cache" / "pmid_to_fulltext_method.csv"
PUBMED_CACHE_COLUMNS = [
    "pmid", "pmcid", "title", "abstract", "year", "journal",
    "authors", "doi", "source", "updated_at"
]

from metapub import PubMedFetcher
from openai import OpenAI
from glob import glob
from urllib.parse import quote
from HelperScripts.metadata_resolver import normalize_authors, resolve_reference_metadata

DEFAULT_OPENAI_MODEL = "gpt-5.4"
_openai_client: Optional[OpenAI] = None


def _get_openai_model(model_name: Optional[str] = None) -> str:
    """Resolve model from explicit arg, env var, or safe default."""
    model = str(model_name or os.getenv("OPENAI_MODEL") or DEFAULT_OPENAI_MODEL).strip()
    return model or DEFAULT_OPENAI_MODEL


def _get_openai_client() -> OpenAI:
    """Create OpenAI client lazily with a clear missing-key error."""
    global _openai_client
    if _openai_client is not None:
        return _openai_client
    api_key = str(os.getenv("OPENAI_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY is not set. Export it in the environment "
            "or add it to a .env file in the project root."
        )
    _openai_client = OpenAI(api_key=api_key)
    return _openai_client

# Shared FlyBase data location for this standalone script.
FLYBASE_DATA = TURBO_ROOT / "UM Lab Users" / "Aadish" / "Data" / "FlyBase"
GET_FBGN_IDS_SCRIPT = Path(__file__).parent / "HelperScripts" / "GetFBgnIDs.py"
BATCH_SIZE = 50
BATCH_STATE_DIRNAME = ".batch_state"
RUN_STORE_FILENAME = "run_store.json"
MAX_PROMPT_WORDS = 5000
SHARED_PROMPT_OVERHEAD_WORDS = 800
MIN_CHUNK_WORDS = 750


###############################################################################
# Full-Text Method Cache (PMID -> successful retrieval method)
###############################################################################

_fulltext_method_cache: Dict[str, str] = {}
_fulltext_method_pending: Dict[str, str] = {}
_fulltext_cache_loaded = False

_pmid_title_abstract_cache: Dict[str, dict] = {}
_pmid_title_abstract_pending: Dict[str, dict] = {}
_pmid_title_abstract_loaded = False


def _load_fulltext_method_cache() -> Dict[str, str]:
    """Load persistent PMID->method mapping from turbo-server cache."""
    global _fulltext_cache_loaded, _fulltext_method_cache
    if _fulltext_cache_loaded:
        return _fulltext_method_cache
    try:
        if FULLTEXT_METHOD_CACHE_PATH.exists():
            cache_df = pd.read_csv(
                FULLTEXT_METHOD_CACHE_PATH,
                dtype={"pmid": str, "method": str}
            )
            cache_df["pmid"] = cache_df["pmid"].apply(
                lambda x: str(x).strip() if pd.notna(x) else ""
            )
            cache_df = cache_df[cache_df["pmid"] != ""].copy()
            if len(cache_df) > 0:
                _fulltext_method_cache = {
                    row["pmid"]: str(row["method"]) if pd.notna(row["method"]) else ""
                    for _, row in cache_df.iterrows()
                }
            print(f"  Loaded {len(_fulltext_method_cache)} full-text cache entries")
    except Exception as e:
        print(f"  [Warning] Could not load full-text method cache: {e}")
    _fulltext_cache_loaded = True
    return _fulltext_method_cache


def _get_cached_fulltext_method(pmid: str) -> Optional[str]:
    """Get cached successful retrieval method for a PMID."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return None
    cache = _load_fulltext_method_cache()
    method = cache.get(pmid_clean)
    return method if method else None


def _set_cached_fulltext_method(pmid: str, method: str):
    """Store PMID->method in memory and pending-write buffer."""
    pmid_clean = str(pmid or "").strip()
    method_clean = str(method or "").strip()
    if not pmid_clean or not method_clean:
        return
    _load_fulltext_method_cache()
    _fulltext_method_cache[pmid_clean] = method_clean
    _fulltext_method_pending[pmid_clean] = method_clean


def _save_fulltext_method_cache_pending():
    """Append pending PMID->method mappings to shared cache CSV."""
    if not _fulltext_method_pending:
        return
    try:
        rows = [{"pmid": p, "method": m} for p, m in _fulltext_method_pending.items()]
        new_df = pd.DataFrame(rows)
        FULLTEXT_METHOD_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        if FULLTEXT_METHOD_CACHE_PATH.exists():
            new_df.to_csv(FULLTEXT_METHOD_CACHE_PATH, mode="a", header=False, index=False)
        else:
            new_df.to_csv(FULLTEXT_METHOD_CACHE_PATH, index=False)
        print(f"  Saved {len(_fulltext_method_pending)} new full-text cache entries")
        _fulltext_method_pending.clear()
    except Exception as e:
        print(f"  [Warning] Could not save full-text method cache: {e}")


def _load_pmid_title_abstract_cache() -> Dict[str, dict]:
    """Load persistent PMID metadata cache from turbo-server CSV."""
    global _pmid_title_abstract_loaded, _pmid_title_abstract_cache
    if _pmid_title_abstract_loaded:
        return _pmid_title_abstract_cache
    try:
        if PUBMED_CACHE_PATH.exists():
            cache_df = pd.read_csv(PUBMED_CACHE_PATH, dtype=str, keep_default_na=False)
            drop_cols = [c for c in cache_df.columns if str(c).startswith("Unnamed")]
            if drop_cols:
                cache_df = cache_df.drop(columns=drop_cols)
            if "pmid" not in cache_df.columns:
                cache_df = pd.DataFrame(columns=PUBMED_CACHE_COLUMNS)
            for col in PUBMED_CACHE_COLUMNS:
                if col not in cache_df.columns:
                    cache_df[col] = ""
            cache_df["pmid"] = cache_df["pmid"].apply(
                lambda x: str(x).strip() if pd.notna(x) else ""
            )
            cache_df = cache_df[cache_df["pmid"].str.isdigit()].copy()
            cache_df = cache_df[PUBMED_CACHE_COLUMNS]
            if len(cache_df) > 0:
                for _, row in cache_df.iterrows():
                    pmid = str(row.get("pmid", "") or "").strip()
                    if not pmid:
                        continue
                    authors_list = normalize_authors(row.get("authors", ""))
                    _pmid_title_abstract_cache[pmid] = {
                        "title": str(row.get("title", "") or ""),
                        "abstract": str(row.get("abstract", "") or ""),
                        "year": str(row.get("year", "") or ""),
                        "journal": str(row.get("journal", "") or ""),
                        "authors": authors_list,
                        "doi": str(row.get("doi", "") or ""),
                        "pmcid": str(row.get("pmcid", "") or ""),
                        "source": str(row.get("source", "") or ""),
                        "updated_at": str(row.get("updated_at", "") or ""),
                    }
            print(f"  Loaded {len(_pmid_title_abstract_cache)} PubMed metadata cache entries")
    except Exception as e:
        print(f"  [Warning] Could not load PubMed metadata cache: {e}")
    _pmid_title_abstract_loaded = True
    return _pmid_title_abstract_cache


def _get_cached_pmid_title_abstract(pmid: str) -> Optional[dict]:
    """Get cached metadata for PMID."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return None
    cache = _load_pmid_title_abstract_cache()
    out = cache.get(pmid_clean)
    if not out:
        return None
    return {
        "title": str(out.get("title", "") or ""),
        "abstract": str(out.get("abstract", "") or ""),
        "year": str(out.get("year", "") or ""),
        "journal": str(out.get("journal", "") or ""),
        "authors": normalize_authors(out.get("authors", [])),
        "doi": str(out.get("doi", "") or ""),
        "pmcid": str(out.get("pmcid", "") or ""),
        "source": str(out.get("source", "") or ""),
        "updated_at": str(out.get("updated_at", "") or ""),
    }


def _set_cached_pmid_title_abstract(
    pmid: str,
    title: str,
    abstract: str,
    *,
    year: str = "",
    journal: str = "",
    authors: Optional[list[str]] = None,
    doi: str = "",
    pmcid: str = "",
    source: str = "",
    updated_at: str = "",
):
    """Store PMID metadata in memory and pending-write buffer."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return
    entry = {
        "title": str(title or ""),
        "abstract": str(abstract or ""),
        "year": str(year or ""),
        "journal": str(journal or ""),
        "authors": normalize_authors(authors or []),
        "doi": str(doi or ""),
        "pmcid": str(pmcid or ""),
        "source": str(source or ""),
        "updated_at": str(updated_at or ""),
    }
    _load_pmid_title_abstract_cache()
    prev = _pmid_title_abstract_cache.get(pmid_clean, {})
    merged = {
        "title": entry["title"] or str(prev.get("title", "") or ""),
        "abstract": entry["abstract"] or str(prev.get("abstract", "") or ""),
        "year": entry["year"] or str(prev.get("year", "") or ""),
        "journal": entry["journal"] or str(prev.get("journal", "") or ""),
        "authors": entry["authors"] or normalize_authors(prev.get("authors", [])),
        "doi": entry["doi"] or str(prev.get("doi", "") or ""),
        "pmcid": entry["pmcid"] or str(prev.get("pmcid", "") or ""),
        "source": entry["source"] or str(prev.get("source", "") or ""),
        "updated_at": entry["updated_at"] or str(prev.get("updated_at", "") or ""),
    }
    _pmid_title_abstract_cache[pmid_clean] = merged
    _pmid_title_abstract_pending[pmid_clean] = entry


def _save_pmid_title_abstract_cache_pending():
    """Persist pending PMID metadata updates to shared cache CSV."""
    if not _pmid_title_abstract_pending:
        return
    try:
        _load_pmid_title_abstract_cache()
        merged_cache: Dict[str, dict] = {}

        if PUBMED_CACHE_PATH.exists():
            try:
                disk_df = pd.read_csv(PUBMED_CACHE_PATH, dtype=str, keep_default_na=False)
                drop_cols = [c for c in disk_df.columns if str(c).startswith("Unnamed")]
                if drop_cols:
                    disk_df = disk_df.drop(columns=drop_cols)
                if "pmid" in disk_df.columns:
                    for col in PUBMED_CACHE_COLUMNS:
                        if col not in disk_df.columns:
                            disk_df[col] = ""
                    disk_df = disk_df[PUBMED_CACHE_COLUMNS]
                    for _, row in disk_df.iterrows():
                        pmid = str(row.get("pmid", "") or "").strip()
                        if not pmid.isdigit():
                            continue
                        merged_cache[pmid] = {
                            "pmcid": str(row.get("pmcid", "") or ""),
                            "title": str(row.get("title", "") or ""),
                            "abstract": str(row.get("abstract", "") or ""),
                            "year": str(row.get("year", "") or ""),
                            "journal": str(row.get("journal", "") or ""),
                            "authors": normalize_authors(row.get("authors", "")),
                            "doi": str(row.get("doi", "") or ""),
                            "source": str(row.get("source", "") or ""),
                            "updated_at": str(row.get("updated_at", "") or ""),
                        }
            except Exception:
                # Keep best-effort behavior even if disk cache is malformed.
                pass

        for pmid, meta in _pmid_title_abstract_cache.items():
            prior = merged_cache.get(pmid, {})
            merged_cache[pmid] = {
                "pmcid": str(meta.get("pmcid", "") or prior.get("pmcid", "")),
                "title": str(meta.get("title", "") or prior.get("title", "")),
                "abstract": str(meta.get("abstract", "") or prior.get("abstract", "")),
                "year": str(meta.get("year", "") or prior.get("year", "")),
                "journal": str(meta.get("journal", "") or prior.get("journal", "")),
                "authors": normalize_authors(meta.get("authors", []) or prior.get("authors", [])),
                "doi": str(meta.get("doi", "") or prior.get("doi", "")),
                "source": str(meta.get("source", "") or prior.get("source", "")),
                "updated_at": str(meta.get("updated_at", "") or prior.get("updated_at", "")),
            }

        rows = []
        for pmid, meta in merged_cache.items():
            rows.append({
                "pmid": str(pmid),
                "pmcid": str(meta.get("pmcid", "") or ""),
                "title": str(meta.get("title", "") or ""),
                "abstract": str(meta.get("abstract", "") or ""),
                "year": str(meta.get("year", "") or ""),
                "journal": str(meta.get("journal", "") or ""),
                "authors": "; ".join(normalize_authors(meta.get("authors", []))),
                "doi": str(meta.get("doi", "") or ""),
                "source": str(meta.get("source", "") or ""),
                "updated_at": str(meta.get("updated_at", "") or ""),
            })
        out_df = pd.DataFrame(rows, columns=PUBMED_CACHE_COLUMNS)
        out_df = out_df.sort_values(by=["pmid"], ascending=True).reset_index(drop=True)
        PUBMED_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        out_df.to_csv(PUBMED_CACHE_PATH, index=False)
        print(
            f"  Saved {len(_pmid_title_abstract_pending)} pending updates "
            f"({len(out_df)} total PubMed metadata rows)"
        )
        _pmid_title_abstract_pending.clear()
    except Exception as e:
        print(f"  [Warning] Could not save PubMed metadata cache: {e}")


def find_latest_tsv(directory: Path, pattern: str) -> Path:
    """
    Find the latest TSV file matching a pattern in a directory.
    Prefers .gz files, falls back to .tsv if no .gz exists.
    """
    # Try gzipped first
    gz_files = sorted(glob(str(directory / f"{pattern}*.tsv.gz")), reverse=True)
    if gz_files:
        return Path(gz_files[0])
    
    # Fall back to uncompressed
    tsv_files = sorted(glob(str(directory / f"{pattern}*.tsv")), reverse=True)
    if tsv_files:
        return Path(tsv_files[0])
    
    raise FileNotFoundError(f"No TSV file matching '{pattern}' found in {directory}")


def load_flybase_tsv(filepath, **kwargs) -> pd.DataFrame:
    """
    Load a FlyBase TSV file, handling various comment/header formats.
    
    FlyBase TSV files have varying formats:
      - Some have ## metadata comments, then #header, then data
      - Some have multiple # comment lines before #header with tabs
      - Some have NO # prefix on headers (plain column names on line 1)
      - Some have #---- separator lines after the header
    
    This function detects and handles all these cases.
    Supports both .tsv and .tsv.gz files.
    """
    import gzip
    
    filepath = Path(filepath)
    
    # Determine opener based on file extension
    if filepath.suffix == '.gz' or str(filepath).endswith('.tsv.gz'):
        opener = lambda f: gzip.open(f, 'rt', encoding='utf-8')
    else:
        opener = lambda f: open(f, 'r', encoding='utf-8')
    
    # First pass: find the header line and count rows to skip
    skip_rows = 0
    header_line = None
    use_custom_header = False
    header_found = False
    
    with opener(filepath) as f:
        for i, line in enumerate(f):
            stripped = line.strip()
            
            # Skip empty lines
            if not stripped:
                skip_rows = i + 1
                continue
            
            # If we already found the header, we're now at data - break
            if header_found:
                if stripped.startswith('#'):
                    skip_rows = i + 1
                    continue
                skip_rows = i
                break
            
            # This line might be the header
            # Check if it starts with # (then strip it) or is a plain header
            if stripped.startswith('#'):
                # Lines with no tabs are comments/metadata.
                if '\t' not in stripped:
                    skip_rows = i + 1
                    continue

                # Header prefixed with # or ## (e.g. "#FBrf\tPMID..." or "##primary_FBid\t...")
                header_content = stripped.lstrip('#')
                cols = header_content.split('\t')
                # Clean column names
                col_names = [c.strip() for c in cols if c.strip()]
                
                if len(col_names) >= 2:
                    # This looks like a valid header
                    header_line = col_names
                    use_custom_header = True
                    header_found = True
                    skip_rows = i + 1
                    # Don't break - continue to skip any separator lines after header
                    continue
                else:
                    # Single column or empty - treat as comment
                    skip_rows = i + 1
                    continue
            else:
                # No # prefix - could be header or data
                cols = stripped.split('\t')
                
                # Heuristic: if first field looks like a FlyBase ID (starts with FB), 
                # this is likely data, not header - let pandas auto-detect
                first_field = cols[0].strip() if cols else ""
                if first_field.startswith('FB') and len(first_field) > 4:
                    # Looks like data (FBst, FBrf, FBal, FBgn, etc.)
                    # This file has no # on header, just let pandas handle it
                    break
                else:
                    # This is likely a plain header (no # prefix)
                    # Let pandas auto-detect it
                    break
    
    # Read the file
    if use_custom_header and header_line:
        df = pd.read_csv(
            filepath,
            sep='\t',
            skiprows=skip_rows,
            names=header_line,
            low_memory=False,
            on_bad_lines='warn',
            **kwargs
        )
    else:
        df = pd.read_csv(
            filepath,
            sep='\t',
            skiprows=skip_rows,
            low_memory=False,
            on_bad_lines='warn',
            **kwargs
        )
    
    return df


###############################################################################
# FlyBase Data Loading (Cached)
###############################################################################

@lru_cache(maxsize=1)
def load_flybase_synonym_table():
    """Load synonyms table from local TSV (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "Genes", "fb_synonym")
    print(f"  Loading FlyBase synonym table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    df = df[df.organism_abbreviation == "Dmel"]
    return df


@lru_cache(maxsize=1)
def build_fbgn_to_symbol_map():
    """Build mapping from FBgn -> current_symbol."""
    df = load_flybase_synonym_table()
    df['current_symbol'] = df['current_symbol'].astype(str).str.strip()
    df['primary_FBid'] = df['primary_FBid'].astype(str).str.strip()
    return dict(zip(df['primary_FBid'], df['current_symbol']))


@lru_cache(maxsize=1)
def build_fbgn_to_all_names_map():
    """Build mapping from FBgn -> set of names including current_symbol and all synonyms."""
    df = load_flybase_synonym_table()
    df = df.rename(columns={"symbol_synonym(s)": "symbol_synonyms"})
    
    fbgn_to_names = defaultdict(set)
    for row in df.itertuples():
        fbgn = str(getattr(row, 'primary_FBid', '')).strip()
        curr = str(getattr(row, 'current_symbol', '')).strip()
        if fbgn:
            if curr:
                fbgn_to_names[fbgn].add(curr)
            syn_field = getattr(row, "symbol_synonyms", "")
            if isinstance(syn_field, str) and syn_field.strip():
                for s in syn_field.split("|"):
                    s = s.strip()
                    if s:
                        fbgn_to_names[fbgn].add(s)
    return {k: set(v) for k, v in fbgn_to_names.items()}


@lru_cache(maxsize=1)
def build_fbgn_to_primary_names_map():
    """Build mapping from FBgn -> set of primary names (current_symbol and current_fullname only)."""
    df = load_flybase_synonym_table()
    
    fbgn_to_names = defaultdict(set)
    for row in df.itertuples():
        fbgn = str(getattr(row, 'primary_FBid', '')).strip()
        if fbgn:
            curr_symbol = str(getattr(row, 'current_symbol', '')).strip()
            if curr_symbol:
                fbgn_to_names[fbgn].add(curr_symbol)
            curr_fullname = str(getattr(row, 'current_fullname', '')).strip()
            if curr_fullname:
                fbgn_to_names[fbgn].add(curr_fullname)
    return {k: set(v) for k, v in fbgn_to_names.items()}


###############################################################################
# FBgn -> PMCID Retrieval from FlyBase Files
###############################################################################

@lru_cache(maxsize=1)
def load_entity_publication():
    """Load entity_publication TSV with FBgn -> PMID mapping (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "FlyBase_References", "entity_publication")
    print(f"  Loading entity publication table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    df = df[df["entity_id"].str.startswith("FBgn", na=False)]
    df["PubMed_id"] = pd.to_numeric(df["PubMed_id"], errors="coerce")
    df.dropna(subset=["PubMed_id"], inplace=True)
    df["PubMed_id"] = df["PubMed_id"].astype(int)
    return df


@lru_cache(maxsize=1)
def load_fbrf_pmid_pmcid():
    """Load fbrf_pmid_pmcid TSV with PMID -> PMCID mapping (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "FlyBase_References", "fbrf_pmid_pmcid_doi")
    print(f"  Loading PMID/PMCID mapping table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    if "pub_type" in df.columns:
        df = df[df["pub_type"].str.strip().str.lower() == "paper"].copy()
    df["PMID"] = pd.to_numeric(df["PMID"], errors="coerce")
    df.dropna(subset=["PMID"], inplace=True)
    df["PMID"] = df["PMID"].astype(int)
    
    # Extract year from miniref
    if "miniref" in df.columns:
        try:
            def extract_year_after_comma(miniref_str):
                miniref_str = str(miniref_str) if miniref_str else ""
                comma_idx = miniref_str.find(",")
                if comma_idx == -1:
                    return None
                after_comma = miniref_str[comma_idx + 1:]
                match = re.search(r"\b(\d{4})\b", after_comma)
                if match:
                    year = int(match.group(1))
                    if 1900 <= year <= 2100:
                        return year
                return None
            _yr = df["miniref"].apply(extract_year_after_comma)
            df["miniref_year"] = pd.to_numeric(_yr, errors="coerce").fillna(0).astype(int)
        except Exception:
            df["miniref_year"] = 0
    else:
        df["miniref_year"] = 0
    return df


def merge_on_pmid():
    """Merge entity_publication & fbrf_pmid_pmcid on PMID => PMCID."""
    entity_df = load_entity_publication().rename(columns={"PubMed_id": "PMID"})
    fbrf_df = load_fbrf_pmid_pmcid()
    merged = pd.merge(entity_df, fbrf_df, on="PMID", how="inner")
    return merged


@lru_cache(maxsize=1)
def build_pmcid_to_year():
    """Build mapping PMCID -> miniref_year."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_year = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip()
        if not pmc or not pmc.lower().startswith("pmc"):
            continue
        try:
            yr = int(getattr(row, "miniref_year", 0) or 0)
        except Exception:
            yr = 0
        pmc_norm = pmc.upper()
        if pmc_norm not in pmcid_to_year or yr > pmcid_to_year[pmc_norm]:
            pmcid_to_year[pmc_norm] = yr
    return pmcid_to_year


@lru_cache(maxsize=1)
def build_pmcid_to_pmid():
    """Build mapping PMCID -> PMID from FlyBase reference metadata."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_pmid = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip().upper()
        pmid = str(getattr(row, "PMID", "")).strip()
        if pmc.startswith("PMC") and pmid.isdigit():
            pmcid_to_pmid[pmc] = pmid
    return pmcid_to_pmid


@lru_cache(maxsize=1)
def build_pmcid_to_doi():
    """Build mapping PMCID -> DOI from FlyBase reference metadata."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_doi = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip().upper()
        doi = str(getattr(row, "DOI", "")).strip()
        if pmc.startswith("PMC") and doi:
            pmcid_to_doi[pmc] = doi
    return pmcid_to_doi


def get_pmcids_for_fbgn_list(fbgn_ids):
    """Get PMCIDs for a list of FBgn IDs from FlyBase precomputed files."""
    merged = merge_on_pmid()
    sub = merged[merged["entity_id"].isin(fbgn_ids)]
    
    fbgn_to_pmcids = defaultdict(set)
    for row in sub.itertuples():
        pmc = str(row.PMCID).strip() if pd.notna(row.PMCID) else None
        if pmc and pmc.lower().startswith("pmc"):
            fbgn_to_pmcids[row.entity_id].add(pmc)
    return fbgn_to_pmcids


###############################################################################
# HTTP Helper
###############################################################################

def _http_get(url, headers=None, params=None, timeout=15, max_retries=3, backoff=0.75):
    """HTTP GET with retries for transient failures."""
    merged_headers = {
        "User-Agent": "fly-gene-classifier/1.0 (+https://github.com/Allada-Lab)",
        "Accept": "*/*",
    }
    if headers:
        merged_headers.update(headers)
    for attempt in range(max_retries):
        try:
            r = requests.get(
                url,
                headers=merged_headers,
                params=params or {},
                timeout=timeout,
                allow_redirects=True
            )
            if r.status_code == 200:
                return r
            if r.status_code in (429, 500, 502, 503, 504):
                wait_s = backoff * (2 ** attempt) + random.uniform(0, backoff)
                time.sleep(wait_s)
                continue
            return None
        except Exception:
            wait_s = backoff * (2 ** attempt) + random.uniform(0, backoff)
            time.sleep(wait_s)
    return None


###############################################################################
# PubMed / Europe PMC Search
###############################################################################

def search_pubmed_pmids_for_fly_gene(symbol: str, primary_names: list[str], keywords: list[str]) -> list[int]:
    """Search PubMed for fly gene references."""
    safe_gene_terms = []
    if symbol and symbol.strip():
        safe_gene_terms.append(symbol.strip())
    
    if primary_names:
        for name in primary_names:
            name = (name or "").strip()
            if not name or name == symbol:
                continue
            if len(name) >= 4 or " " in name or "-" in name:
                safe_gene_terms.append(name)
    
    if not safe_gene_terms:
        return []
    
    gene_or = " OR ".join([f'"{t}"' for t in sorted(safe_gene_terms)])
    kw_terms = [k.strip() for k in (keywords or []) if k and k.strip()]
    kw_or = " OR ".join([f'"{k}"' for k in kw_terms]) if kw_terms else ""
    drosophila_filter = '("drosophila" OR "fruit fly")'
    
    if kw_or:
        term = f"({gene_or}) AND ({kw_or}) AND {drosophila_filter}"
    else:
        term = f"({gene_or}) AND {drosophila_filter}"
    
    params = {
        "db": "pubmed",
        "term": term,
        "retmax": 200,
        "retmode": "json",
        "sort": "pub+date",
    }
    if NCBI_API_KEY:
        params["api_key"] = NCBI_API_KEY
    
    print(f"  [PubMed] Query: {term[:100]}...")
    r = _http_get("https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi", params=params, timeout=20)
    if r is None:
        return []
    try:
        data = r.json()
        ids = data.get("esearchresult", {}).get("idlist", [])
        return [int(x) for x in ids if str(x).isdigit()]
    except Exception:
        return []


def search_europe_pmc_pmids_for_fly_gene(symbol: str, primary_names: list[str], keywords: list[str]) -> list[int]:
    """Search Europe PMC for fly gene references."""
    safe_gene_terms = []
    if symbol and symbol.strip():
        safe_gene_terms.append(symbol.strip())
    
    if primary_names:
        for name in primary_names:
            name = (name or "").strip()
            if not name or name == symbol:
                continue
            if len(name) >= 4 or " " in name or "-" in name:
                safe_gene_terms.append(name)
    
    if not safe_gene_terms:
        return []
    
    gene_or = " OR ".join([f'"{t}"' for t in sorted(safe_gene_terms)])
    kw_terms = [k.strip() for k in (keywords or []) if k and k.strip()]
    kw_or = " OR ".join([f'"{k}"' for k in kw_terms]) if kw_terms else ""
    drosophila_filter = '("drosophila" OR "fruit fly") AND "gene"'
    
    if kw_or:
        query = f"({gene_or}) AND ({kw_or}) AND {drosophila_filter} AND OPEN_ACCESS:Y"
    else:
        query = f"({gene_or}) AND {drosophila_filter} AND OPEN_ACCESS:Y"
    
    params = {
        "query": query,
        "resultType": "core",
        "pageSize": 200,
        "format": "json",
    }
    
    print(f"  [Europe PMC] Query: {query[:100]}...")
    r = _http_get("https://www.ebi.ac.uk/europepmc/webservices/rest/search", params=params, timeout=20)
    if r is None:
        return []
    try:
        data = r.json() or {}
        results = (data.get("resultList") or {}).get("result", []) or []
        pmids = []
        for rec in results:
            if rec.get("source") == "MED":
                pid = rec.get("pmid") or rec.get("id")
                if pid and str(pid).isdigit():
                    pmids.append(int(pid))
        return pmids
    except Exception:
        return []


def batch_pmids_to_pmcids(pmids: list[int]) -> dict[int, str]:
    """Convert batch of PMIDs to PMCIDs."""
    if not pmids:
        return {}
    
    result = {}
    batch_size = 200
    for i in range(0, len(pmids), batch_size):
        batch = pmids[i:i + batch_size]
        ids_str = ",".join(str(p) for p in batch)
        url = "https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/"
        params = {
            "tool": "fly_gene_classifier",
            "email": "aadish98@gmail.com",
            "ids": ids_str,
            "format": "json"
        }
        if NCBI_API_KEY:
            params["api_key"] = NCBI_API_KEY
        
        try:
            r = _http_get(url, params=params, timeout=30)
            if r and r.status_code == 200:
                data = r.json()
                recs = data.get("records", [])
                for rec in recs:
                    pmid_val = rec.get("pmid")
                    pmcid_val = rec.get("pmcid")
                    if pmid_val and pmcid_val:
                        try:
                            result[int(pmid_val)] = pmcid_val
                        except ValueError:
                            pass
        except Exception as e:
            print(f"  [Warning] Batch PMID->PMCID error: {e}")
        time.sleep(0.1)
    
    return result


def get_pubmed_pmcids_for_fly_gene(fbgn_id: str, symbol: str, primary_names: set[str], keywords: list[str]) -> set[str]:
    """Search PubMed and return PMCIDs."""
    pmids = search_pubmed_pmids_for_fly_gene(symbol, list(primary_names), keywords)
    if not pmids:
        return set()
    pmid_to_pmc = batch_pmids_to_pmcids(pmids)
    return set(pmid_to_pmc.values())


def get_europe_pmc_pmcids_for_fly_gene(fbgn_id: str, symbol: str, primary_names: set[str], keywords: list[str]) -> set[str]:
    """Search Europe PMC and return PMCIDs."""
    pmids = search_europe_pmc_pmids_for_fly_gene(symbol, list(primary_names), keywords)
    if not pmids:
        return set()
    pmid_to_pmc = batch_pmids_to_pmcids(pmids)
    return set(pmid_to_pmc.values())


###############################################################################
# Metadata Fetching
###############################################################################

_meta_cache = {}


@lru_cache(maxsize=1)
def _get_pubmed_fetcher():
    return PubMedFetcher(email="aadish98@gmail.com", api_key=NCBI_API_KEY)


def fetch_title_abstract_pmid(pmid: str) -> tuple[str, str]:
    """Return (title, abstract) for PMID, using persistent cache first."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean or not pmid_clean.isdigit():
        return "", ""

    cached = _get_cached_pmid_title_abstract(pmid_clean)
    if cached:
        return str(cached.get("title", "") or ""), str(cached.get("abstract", "") or "")

    try:
        fetcher = _get_pubmed_fetcher()
        art = fetcher.article_by_pmid(pmid_clean)
        title = str(getattr(art, "title", "") or "")
        abstract = str(getattr(art, "abstract", "") or "")
        _set_cached_pmid_title_abstract(pmid_clean, title, abstract)
        return title, abstract
    except Exception:
        return "", ""


def fetch_title_abstract_pmcid(pmcid, pmid: str = ""):
    """Return (title, abstract, year, journal, authors, doi) for PMCID."""
    pmcid_norm = str(pmcid or "").strip().upper()
    if pmcid_norm in _meta_cache:
        return _meta_cache[pmcid_norm]
    
    title_from_pmid, abstract_from_pmid = fetch_title_abstract_pmid(pmid)

    fetcher = _get_pubmed_fetcher()
    try:
        art = fetcher.article_by_pmcid(pmcid_norm)
        yr = art.year or "n.d."
        journal = getattr(art, "journal", "") or ""
        doi = getattr(art, "doi", "") or ""
        authors = []
        try:
            authors = list(getattr(art, "authors", []) or [])
        except Exception:
            authors = []

        title = str(getattr(art, "title", "") or "") or title_from_pmid
        abstract = str(getattr(art, "abstract", "") or "") or abstract_from_pmid
        if pmid and (title or abstract or journal or authors or doi):
            _set_cached_pmid_title_abstract(
                pmid,
                title,
                abstract,
                year=str(yr or ""),
                journal=journal,
                authors=authors,
                doi=doi,
                pmcid=pmcid_norm,
                source="metapub",
            )

        result = (title, abstract, str(yr), journal, authors, doi)
        _meta_cache[pmcid_norm] = result
        return result
    except Exception:
        result = ("", "", "n.d.", "", [], "")
        _meta_cache[pmcid_norm] = result
        return result


def matches_keywords_list(title, abstract, keywords_list):
    """True if ANY keyword appears in title+abstract."""
    if not keywords_list:
        return True
    text = (title + " " + abstract).lower()
    for kw in keywords_list:
        if not kw:
            continue
        kw_lower = kw.strip().lower()
        if not kw_lower:
            continue
        words = re.findall(r'\b\w+\b', text)
        for word in words:
            if kw_lower in word:
                return True
        if kw_lower in text:
            return True
    return False


###############################################################################
# Full Text Retrieval
###############################################################################

def fetch_from_pmcoa_pdf(pmcid):
    """Fetch PDF from PMC OA."""
    pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf"
    r = _http_get(pdf_url, timeout=25)
    if r is not None and r.status_code == 200:
        text = _extract_pdf_text(r.content)
        if len(text) > 300:
            return text
    return ""


def fetch_from_pmcoa_xml(pmcid):
    """Fetch XML from PMC OA."""
    xml_url = f"https://www.ncbi.nlm.nih.gov/pmc/oai/oai.cgi?verb=GetRecord&identifier=oai:pubmedcentral.nih.gov:{pmcid.replace('PMC', '')}&metadataPrefix=pmc"
    r = _http_get(xml_url, timeout=25)
    if r is not None and r.status_code == 200 and len(r.content) > 500:
        text = re.sub(r"<[^>]+>", " ", r.text)
        return re.sub(r"\s+", " ", text).strip()
    return ""


def fetch_from_europepmc_api(identifier):
    """Fetch from Europe PMC API."""
    if identifier.lower().startswith("pmc"):
        id_param = "PMC" + identifier.replace("PMC", "")
    else:
        id_param = identifier
    url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/{id_param}/fullTextXML"
    r = _http_get(url, timeout=25)
    if r is not None and r.status_code == 200 and len(r.content) > 500:
        text = re.sub(r"<[^>]+>", " ", r.text)
        return re.sub(r"\s+", " ", text).strip()
    return ""


def fetch_from_pmc_html(pmcid):
    """Scrape HTML from PMC page."""
    page_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/"
    r = _http_get(page_url, timeout=25)
    if r is not None and r.status_code == 200:
        text = _extract_html_text(r.text, min_chars=1000)
        if text:
            return text
    return ""


def _normalize_doi(doi: str) -> str:
    """Normalize DOI string to bare DOI value."""
    d = str(doi or "").strip()
    if not d:
        return ""
    d_low = d.lower()
    if d_low.startswith("https://doi.org/"):
        return d[len("https://doi.org/"):]
    if d_low.startswith("http://doi.org/"):
        return d[len("http://doi.org/"):]
    if d_low.startswith("doi:"):
        return d[4:].strip()
    return d


def _is_pdf_response(response) -> bool:
    """Heuristic check for PDF content."""
    if response is None:
        return False
    content_type = str(response.headers.get("content-type", "")).lower()
    if "application/pdf" in content_type or "pdf" in content_type:
        return True
    return bool((response.content or b"")[:5] == b"%PDF-")


def _extract_pdf_text(content: bytes) -> str:
    """Extract text from raw PDF bytes."""
    if not content or len(content) <= 500:
        return ""
    try:
        reader = PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            text += "\n"
        return text
    except Exception:
        return ""


def _extract_html_text(html_text: str, min_chars: int = 300) -> str:
    """Extract cleaned text from HTML."""
    if not html_text or len(html_text) <= 300:
        return ""
    txt = re.sub(r"<script[\s\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
    txt = re.sub(r"<style[\s\S]*?</style>", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"<[^>]+>", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    if len(txt) >= min_chars:
        return txt
    return ""


def fetch_full_text_via_unpaywall(doi: str) -> tuple[str, str]:
    """Try to get full text via Unpaywall."""
    if not doi:
        return "", ""
    doi_norm = _normalize_doi(doi)
    url = f"https://api.unpaywall.org/v2/{quote(doi_norm, safe='')}"
    r = _http_get(url, params={"email": UNPAYWALL_TOKEN}, timeout=20)
    if r is None:
        return "", ""
    try:
        data = r.json()
        oa_locs = data.get("oa_locations", []) or []
        best = data.get("best_oa_location")
        if best:
            oa_locs = [best] + [loc for loc in oa_locs if loc != best]

        for loc in oa_locs:
            pdf_url = (loc or {}).get("url_for_pdf") or ""
            page_url = (loc or {}).get("url") or ""

            if pdf_url:
                r_pdf = _http_get(pdf_url, timeout=30)
                if r_pdf is not None and r_pdf.status_code == 200 and _is_pdf_response(r_pdf):
                    text = _extract_pdf_text(r_pdf.content)
                    if len(text) > 300:
                        return text, "Unpaywall PDF"

            if page_url:
                r_page = _http_get(page_url, timeout=25)
                if r_page is not None and r_page.status_code == 200:
                    if _is_pdf_response(r_page):
                        text = _extract_pdf_text(r_page.content)
                        if len(text) > 300:
                            return text, "Unpaywall PDF"
                    text = _extract_html_text(r_page.text, min_chars=300)
                    if text:
                        return text, "Unpaywall HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_openalex(doi: str) -> tuple[str, str]:
    """Try OpenAlex OA links for DOI retrieval."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""

    r = _http_get(
        "https://api.openalex.org/works",
        params={"filter": f"doi:{doi_norm.lower()}", "per-page": 1},
        timeout=20
    )
    if r is None:
        return "", ""
    try:
        data = r.json() or {}
        results = data.get("results", []) or []
        if not results:
            return "", ""
        work = results[0] or {}
        oa_url = (
            ((work.get("open_access") or {}).get("oa_url"))
            or ((work.get("primary_location") or {}).get("pdf_url"))
            or ((work.get("primary_location") or {}).get("landing_page_url"))
        )
        if not oa_url:
            return "", ""

        rr = _http_get(oa_url, timeout=25)
        if rr is None:
            return "", ""
        if _is_pdf_response(rr):
            text = _extract_pdf_text(rr.content)
            if len(text) > 300:
                return text, "OpenAlex PDF"
        text = _extract_html_text(rr.text, min_chars=300)
        if text:
            return text, "OpenAlex HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_crossref(doi: str) -> tuple[str, str]:
    """Try DOI links exposed by Crossref metadata."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""
    r = _http_get(f"https://api.crossref.org/works/{quote(doi_norm, safe='')}", timeout=20)
    if r is None:
        return "", ""
    try:
        message = (r.json() or {}).get("message", {}) or {}
        links = message.get("link", []) or []
        for link in links:
            url = str((link or {}).get("URL", "")).strip()
            if not url:
                continue
            rr = _http_get(url, timeout=25)
            if rr is None:
                continue
            if _is_pdf_response(rr):
                text = _extract_pdf_text(rr.content)
                if len(text) > 300:
                    return text, "Crossref PDF"
            text = _extract_html_text(rr.text, min_chars=300)
            if text:
                return text, "Crossref HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_doi_resolver(doi: str) -> tuple[str, str]:
    """Last-resort retrieval from DOI landing page."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""
    rr = _http_get(f"https://doi.org/{doi_norm}", timeout=25, max_retries=2)
    if rr is None:
        return "", ""
    if _is_pdf_response(rr):
        text = _extract_pdf_text(rr.content)
        if len(text) > 300:
            return text, "DOI Resolver PDF"
    text = _extract_html_text(rr.text, min_chars=300)
    if text:
        return text, "DOI Resolver HTML"
    return "", ""


def enrich_missing_ids_from_europepmc(pmcid: str = "", doi: str = "") -> tuple[str, str]:
    """Resolve missing PMCID/DOI via Europe PMC search."""
    pmcid_clean = str(pmcid or "").strip().upper()
    doi_clean = _normalize_doi(doi)
    query = ""
    if pmcid_clean:
        query = f"PMCID:{pmcid_clean}"
    elif doi_clean:
        query = f'DOI:"{doi_clean}"'
    else:
        return "", ""

    r = _http_get(
        "https://www.ebi.ac.uk/europepmc/webservices/rest/search",
        params={"query": query, "format": "json", "pageSize": 1},
        timeout=20
    )
    if r is None:
        return "", ""
    try:
        data = r.json() or {}
        results = (data.get("resultList") or {}).get("result", []) or []
        if not results:
            return "", ""
        rec = results[0] or {}
        out_pmcid = str(rec.get("pmcid", "") or "").strip().upper()
        out_doi = _normalize_doi(rec.get("doi", ""))
        return out_pmcid, out_doi
    except Exception:
        return "", ""


def fetch_full_text_by_doi(doi: str) -> tuple[str, str]:
    """DOI-based full-text retrieval cascade."""
    for method in [
        fetch_full_text_via_unpaywall,
        fetch_full_text_via_openalex,
        fetch_full_text_via_crossref,
        fetch_full_text_via_doi_resolver,
    ]:
        try:
            text, label = method(doi)
            if text and len(text) > 300:
                return text, label
        except Exception:
            continue
    return "", ""


def _coerce_fulltext_from_cached_method(cached_method: str, pmcid: str, doi_norm: str) -> tuple[str, str]:
    """Replay one known-good method label and return a valid full-text match."""
    for label, method in [
        ("PMC OA PDF", lambda: (fetch_from_pmcoa_pdf(pmcid), "PMC OA PDF") if pmcid else ("", "")),
        ("PMC OA XML", lambda: (fetch_from_pmcoa_xml(pmcid), "PMC OA XML") if pmcid else ("", "")),
        ("Europe PMC", lambda: (fetch_from_europepmc_api(pmcid), "Europe PMC") if pmcid else ("", "")),
        ("PMC HTML", lambda: (fetch_from_pmc_html(pmcid), "PMC HTML") if pmcid else ("", "")),
    ]:
        if cached_method != label:
            continue
        text, out_label = method()
        if text and len(text) > 300:
            return text, out_label

    for marker, method in [
        ("Unpaywall", fetch_full_text_via_unpaywall),
        ("OpenAlex", fetch_full_text_via_openalex),
        ("Crossref", fetch_full_text_via_crossref),
        ("DOI Resolver", fetch_full_text_via_doi_resolver),
    ]:
        if marker not in cached_method or not doi_norm:
            continue
        text, out_label = method(doi_norm)
        if text and len(text) > 300:
            return text, out_label
    return "", ""


def fetch_full_text_by_id(identifier, id_type="pmcid", doi=None, pmid=None):
    """Cascade through methods to get full text."""
    pmcid = str(identifier or "").strip().upper()
    doi_norm = _normalize_doi(doi)
    pmid_norm = str(pmid or "").strip()

    # Try cached known-good method for this PMID first.
    cached_method = _get_cached_fulltext_method(pmid_norm) if pmid_norm else None
    if cached_method:
        try:
            text, label = _coerce_fulltext_from_cached_method(cached_method, pmcid, doi_norm)
            if text and len(text) > 300:
                return text, label
        except Exception:
            pass

    if not doi_norm and pmcid:
        _epmc_pmcid, epmc_doi = enrich_missing_ids_from_europepmc(pmcid=pmcid, doi=doi_norm)
        if epmc_doi and not doi_norm:
            doi_norm = epmc_doi

    for method, label in [
        (fetch_from_pmcoa_pdf, "PMC OA PDF"),
        (fetch_from_pmcoa_xml, "PMC OA XML"),
        (fetch_from_europepmc_api, "Europe PMC"),
        (fetch_from_pmc_html, "PMC HTML"),
    ]:
        try:
            text = method(pmcid)
            if len(text) > 300:
                if pmid_norm:
                    _set_cached_fulltext_method(pmid_norm, label)
                return text, label
        except Exception:
            continue
    
    # DOI-based fallback cascade
    if doi_norm:
        try:
            text, label = fetch_full_text_by_doi(doi_norm)
            if text and len(text) > 300:
                if pmid_norm:
                    _set_cached_fulltext_method(pmid_norm, label)
                return text, label
        except Exception:
            pass
    
    # Fallback: title+abstract (prefer persistent PMID cache when available).
    title, abstract = fetch_title_abstract_pmid(pmid_norm)
    if not title and not abstract:
        title, abstract, *_ = fetch_title_abstract_pmcid(pmcid, pmid=pmid_norm)
    fallback_text = f"TITLE: {title}\n\nABSTRACT: {abstract}"
    return fallback_text, "Title+Abstract only"


###############################################################################
# Summarization
###############################################################################

class FunctionPhenotypeSummary(BaseModel):
    function: str = ""
    phenotypes: str = ""
    skip_reference: bool = False
    skip_reason: str = ""


class FinalFunctionPhenotypeSummary(BaseModel):
    function: str = ""
    phenotypes: str = ""


class ReferenceReagent(BaseModel):
    stock_id: str = ""
    collection: str = ""
    reagent_type: str = ""
    evidence_snippet: str = ""
    functional_validity: str = ""
    reagent_name: str = ""


class ReagentExtraction(BaseModel):
    reagents: List[ReferenceReagent] = Field(default_factory=list)


def _count_words(text: str) -> int:
    """Approximate word count using whitespace-delimited tokens."""
    return len(re.findall(r"\S+", str(text or "")))


def _clean_text(value: Any) -> str:
    """Collapse repeated whitespace and coerce to string."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    """Remove duplicate strings while preserving original order."""
    seen = set()
    out = []
    for value in values or []:
        cleaned = _clean_text(value)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def _join_unique_texts(values: list[str], separator: str = "; ") -> str:
    """Join deduplicated text fragments with a stable separator."""
    return separator.join(_dedupe_preserve_order(values))


def _chunk_text_by_word_budget(text: str, max_words: int) -> list[str]:
    """Split text into deterministic chunks that stay within a word budget."""
    cleaned = str(text or "").strip()
    if not cleaned:
        return [""]

    paragraphs = [
        _clean_text(chunk)
        for chunk in re.split(r"\n\s*\n+", cleaned)
        if _clean_text(chunk)
    ]
    if not paragraphs:
        paragraphs = [_clean_text(cleaned)]

    max_words = max(int(max_words or 0), MIN_CHUNK_WORDS)
    chunks: list[str] = []
    current_parts: list[str] = []
    current_words = 0

    for paragraph in paragraphs:
        paragraph_words = _count_words(paragraph)
        if paragraph_words > max_words:
            if current_parts:
                chunks.append("\n\n".join(current_parts))
                current_parts = []
                current_words = 0

            words = paragraph.split()
            for idx in range(0, len(words), max_words):
                chunk_words = words[idx: idx + max_words]
                if chunk_words:
                    chunks.append(" ".join(chunk_words))
            continue

        if current_parts and current_words + paragraph_words > max_words:
            chunks.append("\n\n".join(current_parts))
            current_parts = [paragraph]
            current_words = paragraph_words
            continue

        current_parts.append(paragraph)
        current_words += paragraph_words

    if current_parts:
        chunks.append("\n\n".join(current_parts))

    return chunks or [cleaned]


def _build_shared_prompt_chunks(
    full_text: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    title: str = "",
    abstract: str = "",
) -> list[str]:
    """Create deterministic chunks sized to fit within the shared prompt budget."""
    cleaned_full_text = str(full_text or "").strip()
    if not cleaned_full_text:
        return [""]

    metadata_blob = " ".join([
        str(gene_symbol or ""),
        str(fbgn_id or ""),
        " ".join(sorted(set(synonyms or []))),
        str(title or ""),
        str(abstract or ""),
    ])
    overhead_words = _count_words(metadata_blob) + SHARED_PROMPT_OVERHEAD_WORDS
    chunk_word_budget = max(MIN_CHUNK_WORDS, MAX_PROMPT_WORDS - overhead_words)

    if _count_words(cleaned_full_text) <= chunk_word_budget:
        return [cleaned_full_text]
    return _chunk_text_by_word_budget(cleaned_full_text, chunk_word_budget)


def _model_dump(value: Any) -> dict[str, Any]:
    """Return a plain dict for a parsed Pydantic model."""
    if value is None:
        return {}
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "dict"):
        return value.dict()
    if isinstance(value, dict):
        return dict(value)
    return {}


def _build_responses_input(messages: list[dict[str, str]]) -> tuple[Optional[str], list[dict[str, str]]]:
    """Split chat-style messages into Responses API instructions and input items."""
    instructions_parts: list[str] = []
    input_items: list[dict[str, str]] = []
    for message in messages or []:
        role = str(message.get("role", "user") or "user").strip()
        content = str(message.get("content", "") or "").strip()
        if not content:
            continue
        if role == "system":
            instructions_parts.append(content)
        else:
            input_items.append({"role": role, "content": content})
    instructions = "\n\n".join(instructions_parts).strip()
    return (instructions or None), input_items


def _get_reasoning_config(model_name: Optional[str] = None) -> Optional[dict[str, str]]:
    """Use high reasoning for GPT-5 family models when supported."""
    model = _get_openai_model(model_name).lower()
    if model.startswith("gpt-5"):
        return {"effort": "high"}
    return None


def _parse_structured_completion(
    messages: list[dict[str, str]],
    response_format,
    model_name: Optional[str] = None,
    max_output_tokens: int = 1500,
):
    """Call the Responses API with structured parsing and light retries."""
    last_err = None
    model = _get_openai_model(model_name)
    instructions, input_items = _build_responses_input(messages)
    for attempt in range(3):
        try:
            request_kwargs: dict[str, Any] = {
                "model": model,
                "input": input_items,
                "text_format": response_format,
                "max_output_tokens": max_output_tokens,
            }
            if instructions:
                request_kwargs["instructions"] = instructions
            reasoning = _get_reasoning_config(model)
            if reasoning is not None:
                request_kwargs["reasoning"] = reasoning

            response = _get_openai_client().responses.parse(**request_kwargs)
            parsed = getattr(response, "output_parsed", None)
            if parsed is not None:
                return parsed
            raw_output = str(getattr(response, "output_text", "") or "").strip()
            raise ValueError(f"Empty parsed response from API: {raw_output or 'no output text returned'}")
        except Exception as e:
            last_err = e
            if "Unsupported parameter: 'reasoning.effort'" in str(e):
                try:
                    retry_kwargs: dict[str, Any] = {
                        "model": model,
                        "input": input_items,
                        "text_format": response_format,
                        "max_output_tokens": max_output_tokens,
                    }
                    if instructions:
                        retry_kwargs["instructions"] = instructions
                    response = _get_openai_client().responses.parse(**retry_kwargs)
                    parsed = getattr(response, "output_parsed", None)
                    if parsed is not None:
                        return parsed
                    raw_output = str(getattr(response, "output_text", "") or "").strip()
                    raise ValueError(f"Empty parsed response from API: {raw_output or 'no output text returned'}")
                except Exception as retry_error:
                    last_err = retry_error
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(str(last_err))


def summarize_reference_function(
    text_chunk: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    model_name: Optional[str] = None,
    title: str = "",
    abstract: str = "",
    chunk_index: int = 1,
    total_chunks: int = 1,
) -> dict[str, Any]:
    """Extract function and phenotype evidence for one chunk of a paper."""
    synonyms_str = ", ".join(sorted(set(synonyms))) if synonyms else ""
    sys = (
        "You are an expert biomedical assistant. Use ONLY the provided title, abstract, "
        "and text chunk. Focus only on the target gene. If the chunk contains no direct "
        "evidence about the target gene, set skip_reference to true."
    )
    user = f"""Assess evidence for gene {gene_symbol} (FBgn: {fbgn_id}; Synonyms: {synonyms_str}).

Return JSON with fields:
- function: concise summary of gene {gene_symbol} function evidence from this chunk only
- phenotypes: concise summary of phenotype evidence of gene {gene_symbol} from this chunk only
- skip_reference: true if this chunk does not contain direct evidence for {gene_symbol}
- skip_reason: brief reason when skip_reference is true

Rules:
- Always name {gene_symbol} explicitly
- Be factual and use only the provided text
- If evidence is generic or unrelated to {gene_symbol}, set skip_reference to true
- If one field is absent but the chunk still contains evidence, return an empty string for that field

Title: {title or ''}
Abstract: {abstract or ''}

Text chunk:
{text_chunk or ''}"""
    try:
        out = _parse_structured_completion(
            [{"role": "system", "content": sys}, {"role": "user", "content": user}],
            response_format=FunctionPhenotypeSummary,
            model_name=model_name,
            max_output_tokens=1200,
        )
        parsed = _model_dump(out)
        function_text = _clean_text(parsed.get("function", ""))
        phenotype_text = _clean_text(parsed.get("phenotypes", ""))
        skip_reference = bool(parsed.get("skip_reference", False))
        skip_reason = _clean_text(parsed.get("skip_reason", ""))
        if not skip_reference and not function_text and not phenotype_text:
            skip_reference = True
            skip_reason = skip_reason or "No function or phenotype evidence returned"
        return {
            "function": function_text,
            "phenotypes": phenotype_text,
            "skip_reference": skip_reference,
            "skip_reason": skip_reason,
        }
    except Exception as e:
        return {
            "function": "",
            "phenotypes": "",
            "skip_reference": True,
            "skip_reason": f"API error: {e}",
        }


def rewrite_function_summary(
    gene_symbol: str,
    fbgn_id: str,
    chunk_summaries: list[dict[str, Any]],
    model_name: Optional[str] = None,
) -> dict[str, str]:
    """Rewrite multiple chunk-level summaries into one cohesive paper summary."""
    chunk_lines = []
    for idx, chunk_summary in enumerate(chunk_summaries or [], start=1):
        function_text = _clean_text(chunk_summary.get("function", ""))
        phenotype_text = _clean_text(chunk_summary.get("phenotypes", ""))
        chunk_lines.append(
            f"Chunk {idx}\n"
            f"Function: {function_text or 'None'}\n"
            f"Phenotypes: {phenotype_text or 'None'}"
        )
    merged_input = "\n\n".join(chunk_lines)

    fallback = {
        "function": _join_unique_texts([summary.get("function", "") for summary in chunk_summaries]),
        "phenotypes": _join_unique_texts([summary.get("phenotypes", "") for summary in chunk_summaries]),
    }

    sys = (
        "You are an expert biomedical assistant. Merge chunk-level evidence summaries into a "
        "single cohesive paper-level summary. Use ONLY the provided chunk summaries."
    )
    user = f"""Rewrite these chunk-level summaries for gene {gene_symbol} (FBgn: {fbgn_id}) into one cohesive paper summary.

Return JSON with fields:
- function
- phenotypes

Rules:
- Always name {gene_symbol} explicitly
- Preserve factual content only
- Remove redundancy across chunks
- Do not introduce new claims

Chunk summaries:
{merged_input}"""
    try:
        out = _parse_structured_completion(
            [{"role": "system", "content": sys}, {"role": "user", "content": user}],
            response_format=FinalFunctionPhenotypeSummary,
            model_name=model_name,
            max_output_tokens=1200,
        )
        parsed = _model_dump(out)
        return {
            "function": _clean_text(parsed.get("function", "")) or fallback["function"],
            "phenotypes": _clean_text(parsed.get("phenotypes", "")) or fallback["phenotypes"],
        }
    except Exception:
        return fallback


def _normalize_collection_name(collection: str) -> str:
    """Normalize collection names into stable export values."""
    cleaned = _clean_text(collection)
    lowered = cleaned.lower()
    if not cleaned:
        return ""
    if "bloomington" in lowered or "bdsc" in lowered:
        return "BDSC"
    if "vienna" in lowered or "vdrc" in lowered:
        return "VDRC"
    if "national institute of genetics" in lowered or re.search(r"\bnig\b", lowered):
        return "NIG"
    if "kyoto" in lowered or "dgrc" in lowered or "drosophila genetic resource center" in lowered:
        return "Kyoto"
    return cleaned


def _normalize_stock_id(stock_id: str, collection: str) -> str:
    """Normalize stock IDs to stable, dedupe-friendly forms."""
    cleaned = _clean_text(stock_id)
    cleaned = cleaned.replace("#", "")
    cleaned = re.sub(r"\s+", "", cleaned)
    if not cleaned:
        return ""

    if collection == "BDSC":
        match = re.fullmatch(r"(?:BL|BDSC)?(\d+)", cleaned, flags=re.IGNORECASE)
        if match:
            return f"{match.group(1)}"
    if collection == "VDRC":
        match = re.fullmatch(r"(?:VDRC)?v?(\d+)", cleaned, flags=re.IGNORECASE)
        if match:
            return f"v{match.group(1)}"
    return cleaned


def _normalize_reagent_record(raw_reagent: Any) -> Optional[dict[str, str]]:
    """Normalize one reagent record and enforce canonical (stock_id, collection) pairs."""
    parsed = _model_dump(raw_reagent)
    collection = _normalize_collection_name(parsed.get("collection", ""))
    stock_id = _normalize_stock_id(parsed.get("stock_id", ""), collection)
    if not stock_id or not collection:
        return None
    return {
        "stock_id": stock_id,
        "collection": collection,
        "reagent_type": _clean_text(parsed.get("reagent_type", "")),
        "evidence_snippet": _clean_text(parsed.get("evidence_snippet", "")),
        "functional_validity": _clean_text(parsed.get("functional_validity", "")),
        "reagent_name": _clean_text(parsed.get("reagent_name", "")),
    }


def extract_reference_reagents(
    text_chunk: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    model_name: Optional[str] = None,
    title: str = "",
    abstract: str = "",
    chunk_index: int = 1,
    total_chunks: int = 1,
) -> list[dict[str, str]]:
    """Extract canonical reagent pairs plus metadata from one paper chunk."""
    synonyms_str = ", ".join(sorted(set(synonyms))) if synonyms else ""
    sys = (
        "You are an expert biomedical assistant. Use ONLY the provided title, abstract, "
        "and text chunk. Extract reagent records only when they pertain to the target gene."
    )
    user = f"""Extract reagents for gene {gene_symbol} (FBgn: {fbgn_id}; Synonyms: {synonyms_str}).

Return JSON with a single field:
- reagents: list of objects with stock_id, collection, reagent_type, evidence_snippet, functional_validity, reagent_name

Rules:
- Only include a reagent if both stock_id and collection can be identified
- Focus only on reagents that pertain to {gene_symbol}
- Deduplicate exact repeats within this chunk
- If no qualifying reagents are present, return an empty list
- Use concise evidence_snippet text quoted or paraphrased from the chunk
- Normalize collection names when clear, including BDSC, VDRC, NIG, and Kyoto

Chunk {chunk_index} of {total_chunks}
Title: {title or ''}
Abstract: {abstract or ''}

Text chunk:
{text_chunk or ''}"""
    try:
        out = _parse_structured_completion(
            [{"role": "system", "content": sys}, {"role": "user", "content": user}],
            response_format=ReagentExtraction,
            model_name=model_name,
            max_output_tokens=1800,
        )
        parsed = _model_dump(out)
        normalized = []
        for reagent in parsed.get("reagents", []) or []:
            record = _normalize_reagent_record(reagent)
            if record:
                normalized.append(record)
        return normalized
    except Exception:
        return []


def _merge_reagent_records(base: dict[str, str], candidate: dict[str, str]) -> dict[str, str]:
    """Merge duplicate reagent records while keeping the richest metadata."""
    merged = dict(base or {})
    for field in ("reagent_name", "reagent_type", "functional_validity"):
        if not merged.get(field) and candidate.get(field):
            merged[field] = candidate[field]
    if len(candidate.get("evidence_snippet", "")) > len(merged.get("evidence_snippet", "")):
        merged["evidence_snippet"] = candidate.get("evidence_snippet", "")
    return merged


def _merge_deduplicated_reagents(reagent_records: list[dict[str, str]]) -> list[dict[str, str]]:
    """Merge reagent records across chunks using the canonical pair as the key."""
    deduped: dict[tuple[str, str], dict[str, str]] = {}
    for record in reagent_records or []:
        stock_id = _clean_text(record.get("stock_id", ""))
        collection = _normalize_collection_name(record.get("collection", ""))
        if not stock_id or not collection:
            continue
        key = (stock_id.lower(), collection.lower())
        normalized_record = {
            "stock_id": stock_id,
            "collection": collection,
            "reagent_type": _clean_text(record.get("reagent_type", "")),
            "evidence_snippet": _clean_text(record.get("evidence_snippet", "")),
            "functional_validity": _clean_text(record.get("functional_validity", "")),
            "reagent_name": _clean_text(record.get("reagent_name", "")),
        }
        if key in deduped:
            deduped[key] = _merge_reagent_records(deduped[key], normalized_record)
        else:
            deduped[key] = normalized_record
    return sorted(
        deduped.values(),
        key=lambda item: (item.get("collection", ""), item.get("stock_id", "")),
    )


def _format_reagent_pairs(reagent_records: list[dict[str, str]]) -> str:
    """Format canonical reagent pairs for sheet export."""
    return "; ".join(
        f"({record.get('stock_id', '')}, {record.get('collection', '')})"
        for record in reagent_records or []
        if record.get("stock_id") and record.get("collection")
    )


def _format_reference_summary(function_text: str, phenotype_text: str) -> str:
    """Keep a compact human-readable summary for the existing summary sheet."""
    return (
        f"- Function: {function_text or 'None reported'}\n"
        f"- Phenotypes: {phenotype_text or 'None reported'}"
    )


def gene_mentioned_in_title_abstract(title: str, abstract: str, gene_symbol: str, synonyms: list[str]) -> bool:
    """Check if gene is mentioned in title/abstract."""
    if not title and not abstract:
        return False
    combined_text = f"{title or ''} {abstract or ''}".lower()
    
    if gene_symbol:
        gene_lower = gene_symbol.lower().strip()
        if gene_lower and len(gene_lower) >= 2:
            if re.search(rf"\b{re.escape(gene_lower)}\b", combined_text):
                return True
    
    if synonyms:
        for synonym in synonyms:
            if synonym:
                syn_lower = synonym.lower().strip()
                if syn_lower and len(syn_lower) >= 2:
                    if re.search(rf"\b{re.escape(syn_lower)}\b", combined_text):
                        return True
    return False


###############################################################################
# Classification
###############################################################################

class GeneClassification(BaseModel):
    gene: str
    category: Optional[List[str]]
    confidence: int
    rationale: str


def _normalize_category_output(raw_category, valid_cats: set[str]) -> list[str]:
    """Normalize model category output into a validated list."""
    if raw_category is None:
        return []
    if isinstance(raw_category, list):
        return [c for c in raw_category if c in valid_cats]
    if isinstance(raw_category, str) and raw_category in valid_cats:
        return [raw_category]
    return []


def _safe_confidence(confidence_value) -> int:
    """Clamp model confidence to a 0-100 integer."""
    try:
        return max(0, min(100, int(confidence_value)))
    except Exception:
        return 0


def classify_gene_from_text(gene_symbol: str, keywords_list: list[str], full_text: str):
    """Classify a gene using GPT."""
    if not full_text or not full_text.strip():
        return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": "No text provided for classification"}
    
    cats = [k.strip() for k in keywords_list if k and k.strip() and k.strip().lower() != "none"]
    valid_cats = set(cats) if cats else set()
    cats_line = ", ".join(cats) if cats else "None"

    sys = (
        "You are an expert biomedical assistant. Use ONLY the provided text. "
        "If evidence is weak or ambiguous, return category as empty list [] or None."
    )
    user = (
        f"Classify gene '{gene_symbol}'. Allowed categories: [{cats_line}]. "
        "You may classify into 1 category, N categories, or None (return empty list [] or None). "
        "Return JSON with fields: gene, category (list of strings, can be empty or None), confidence (0-100), rationale (1-2 lines).\n\n"
        "TEXT START\n" + full_text[:200000] + "\nTEXT END"
    )

    try:
        out = _parse_structured_completion(
            [{"role": "system", "content": sys}, {"role": "user", "content": user}],
            response_format=GeneClassification,
            max_output_tokens=4000,
        )
        if not out:
            return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": "Empty response from API"}
        
        category_list = _normalize_category_output(out.category, valid_cats)
        confidence = _safe_confidence(out.confidence)
        
        rationale = str(out.rationale) if hasattr(out, 'rationale') and out.rationale else ""
        
        return {
            "gene": gene_symbol,
            "category": category_list,
            "confidence": confidence,
            "rationale": rationale
        }
    except Exception as e:
        error_msg = str(e)
        if "length limit was reached" in error_msg or "maximum context length" in error_msg:
            try:
                shorter_text = full_text[:50000]
                out = _parse_structured_completion(
                    [{"role": "system", "content": sys}, {"role": "user", "content": user.replace(full_text[:200000], shorter_text)}],
                    response_format=GeneClassification,
                    max_output_tokens=2000,
                )
                if out:
                    category_list = _normalize_category_output(out.category, valid_cats)
                    return {
                        "gene": gene_symbol,
                        "category": category_list,
                        "confidence": _safe_confidence(out.confidence or 0),
                        "rationale": str(out.rationale or "")
                    }
            except Exception as e2:
                return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": f"Error: {e2}"}
        return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": f"Error: {e}"}


###############################################################################
# Main Processing Pipeline
###############################################################################

def process_gene_set(fbgn_ids: list[str], keywords_list: list[str], reference_limit: int = 500):
    """
    Process a set of FBgn IDs through the classification and reagent-finder pipeline.
    Returns (classification_results, all_summaries, validated).
    """
    print(f"\n{'='*60}")
    print(f"Processing {len(fbgn_ids)} genes...")
    print(f"Keywords: {keywords_list}")
    print(f"Reference limit per gene: {reference_limit}")
    print(f"{'='*60}\n")
    
    # Build mappings
    fbgn_to_symbol = build_fbgn_to_symbol_map()
    fbgn_to_primary_names = build_fbgn_to_primary_names_map()
    fbgn_all_names_map = build_fbgn_to_all_names_map()
    pmcid_to_year = build_pmcid_to_year()
    pmcid_to_pmid = build_pmcid_to_pmid()
    pmcid_to_doi = build_pmcid_to_doi()
    
    # Create validated mapping (gene_symbol -> FBgn).
    validated = {
        fbgn_to_symbol.get(fbgn_clean, fbgn_clean): fbgn_clean
        for fbgn_clean in (str(fbgn).strip() for fbgn in fbgn_ids)
        if fbgn_clean.startswith("FBgn")
    }
    
    if not validated:
        print("No valid FBgn IDs found.")
        return {}, [], {}
    
    print(f"Validated {len(validated)} genes")
    
    # Track PMCID sources
    pmcid_source = {}
    fbgn_to_pmcids_raw = defaultdict(set)

    def _add_pmcids_for_gene(fbgn_id: str, pmcs: set[str], source_label: str):
        """Update raw PMCID set and source metadata in one place."""
        if not pmcs:
            return
        fbgn_to_pmcids_raw[fbgn_id].update(pmcs)
        for pmc in pmcs:
            pmcid_source.setdefault(pmc, set()).add(source_label)

    def _source_priority(sources: set[str]) -> int:
        """Higher is better: prefer consensus across providers."""
        has_flybase = 'flybase' in sources
        has_pubmed = 'pubmed' in sources
        has_europe_pmc = 'europe_pmc' in sources
        if has_flybase and has_pubmed and has_europe_pmc:
            return 7
        if has_flybase and has_pubmed:
            return 6
        if has_pubmed and has_europe_pmc:
            return 5
        if has_europe_pmc and has_flybase:
            return 4
        if has_pubmed:
            return 3
        if has_flybase:
            return 2
        if has_europe_pmc:
            return 1
        return 0

    def _format_ref_source(pmc: str) -> str:
        source_labels = []
        pmc_sources = pmcid_source.get(pmc, set())
        if 'flybase' in pmc_sources:
            source_labels.append("FlyBase")
        if 'pubmed' in pmc_sources:
            source_labels.append("PubMed")
        if 'europe_pmc' in pmc_sources:
            source_labels.append("Europe PMC")
        return ", ".join(source_labels) if source_labels else "Unknown"
    
    # 1) Get PMCIDs from FlyBase precomputed files
    print("\n[Step 1] Fetching references from FlyBase...")
    flybase_pmcids = get_pmcids_for_fbgn_list(list(validated.values()))
    for fbgn_id, pmcs in flybase_pmcids.items():
        _add_pmcids_for_gene(fbgn_id, pmcs, 'flybase')
    
    # 2) Search PubMed for additional references
    print("\n[Step 2] Querying PubMed for additional references...")
    for gene_symbol, fbgn_id in validated.items():
        fly_symbol = fbgn_to_symbol.get(fbgn_id, gene_symbol)
        primary_names = fbgn_to_primary_names.get(fbgn_id, set())
        
        pubmed_pmcids = get_pubmed_pmcids_for_fly_gene(fbgn_id, fly_symbol, primary_names, keywords_list)
        _add_pmcids_for_gene(fbgn_id, pubmed_pmcids, 'pubmed')
        print(f"  {fly_symbol}: {len(pubmed_pmcids)} PMCIDs from PubMed")
    
    # 3) Search Europe PMC for additional references
    print("\n[Step 3] Querying Europe PMC for additional references...")
    for gene_symbol, fbgn_id in validated.items():
        fly_symbol = fbgn_to_symbol.get(fbgn_id, gene_symbol)
        primary_names = fbgn_to_primary_names.get(fbgn_id, set())
        
        europe_pmc_pmcids = get_europe_pmc_pmcids_for_fly_gene(fbgn_id, fly_symbol, primary_names, keywords_list)
        _add_pmcids_for_gene(fbgn_id, europe_pmc_pmcids, 'europe_pmc')
        print(f"  {fly_symbol}: {len(europe_pmc_pmcids)} PMCIDs from Europe PMC")
    
    # 4) Sort and limit references
    def _pmc_numeric(p):
        try:
            return int(str(p).upper().lstrip("PMC"))
        except:
            return 0
    
    def _sort_key_sources_then_recency(p):
        p_norm = str(p).upper()
        sources = pmcid_source.get(p, set())
        year = pmcid_to_year.get(p_norm, 0)
        pmc_num = _pmc_numeric(p_norm)
        priority = _source_priority(sources)
        return (priority, year, pmc_num)
    
    fbgn_to_pmcids = {}
    for fbgn, pmcs in fbgn_to_pmcids_raw.items():
        sorted_pmcs = sorted(list(pmcs), key=_sort_key_sources_then_recency, reverse=True)
        fbgn_to_pmcids[fbgn] = set(sorted_pmcs[:reference_limit])
    
    # 5) Filter by keywords and fetch metadata
    print("\n[Step 4] Filtering references by keywords and fetching metadata...")
    all_pmcids = set()
    for fbgn, pmcs in fbgn_to_pmcids.items():
        all_pmcids.update(pmcs)
    
    meta_cache = {}
    high_quality_pmcids_per_gene = defaultdict(set)
    
    total = len(all_pmcids)
    for i, pmc in enumerate(sorted(all_pmcids, key=_sort_key_sources_then_recency, reverse=True)):
        if (i + 1) % 20 == 0:
            print(f"  Processing metadata {i+1}/{total}...")
        try:
            pmc_norm = str(pmc).upper()
            pmid_for_pmc = pmcid_to_pmid.get(pmc_norm, "")
            resolved = resolve_reference_metadata(
                pmc_norm,
                _get_pubmed_fetcher(),
                api_key=NCBI_API_KEY or "",
                pmid_hint=pmid_for_pmc,
                cache_getter=_get_cached_pmid_title_abstract,
                cache_setter=_set_cached_pmid_title_abstract,
            )

            title = str(resolved.get("title", "") or "")
            abstract = str(resolved.get("abstract", "") or "")
            year = str(resolved.get("year", "") or "")
            journal = str(resolved.get("journal", "") or "")
            authors = normalize_authors(resolved.get("authors", []))
            doi = str(resolved.get("doi", "") or "")

            if not year:
                year_val = pmcid_to_year.get(pmc_norm, 0)
                year = str(year_val) if year_val else "n.d."
            if not doi:
                doi = pmcid_to_doi.get(pmc_norm, "")

            meta_cache[pmc] = (title, abstract, year, journal, authors, doi)

            if matches_keywords_list(title, abstract, keywords_list):
                for gene_symbol, fbgn_id in validated.items():
                    if pmc in fbgn_to_pmcids.get(fbgn_id, set()):
                        high_quality_pmcids_per_gene[fbgn_id].add(pmc)
        except Exception:
            meta_cache[pmc] = ("", "", "n.d.", "", [], "")
    
    # 6) Process each gene: summarize and classify
    print("\n[Step 5] Summarizing and classifying genes...")
    
    fbgn_to_summary_chunks = defaultdict(list)
    fbgn_to_pmcids_all = defaultdict(set)
    fbgn_high_quality_ref_summaries = defaultdict(list)
    fbgn_to_full_text_pmcids = defaultdict(set)
    all_summaries = []
    fly_gene_hits = {}
    
    for gene_symbol, fbgn_id in validated.items():
        fly_symbol = fbgn_to_symbol.get(fbgn_id, gene_symbol)
        print(f"\n  Processing: {fly_symbol} ({fbgn_id})")
        
        # Get synonyms
        synonyms = list(fbgn_all_names_map.get(fbgn_id, set()))
        
        # Get high quality PMCIDs for this gene
        pmcs_for_gene = sorted(list(high_quality_pmcids_per_gene.get(fbgn_id, set())), 
                               key=_sort_key_sources_then_recency, reverse=True)
        
        high_q_count = 0
        total_attempts = 0
        
        for pmc in pmcs_for_gene:
            if total_attempts >= 50 or high_q_count >= 6:
                break
            
            meta_data = meta_cache.get(pmc)
            if not meta_data or len(meta_data) != 6:
                continue
            title, abstract, year, journal, authors, doi = meta_data
            
            # Try fetching full text
            pmid_for_pmc = pmcid_to_pmid.get(str(pmc).upper(), "")
            full_text, source_label = fetch_full_text_by_id(
                pmc,
                "pmcid",
                doi=doi,
                pmid=pmid_for_pmc
            )
            
            if full_text and source_label != "Title+Abstract only":
                fbgn_to_full_text_pmcids[fbgn_id].add(pmc)
            
            has_title_abstract_match = gene_mentioned_in_title_abstract(
                title or "", abstract or "", fly_symbol, synonyms
            )
            if not (full_text and source_label != "Title+Abstract only") and not has_title_abstract_match:
                continue

            shared_chunks = _build_shared_prompt_chunks(
                full_text if source_label != "Title+Abstract only" else "",
                fly_symbol,
                fbgn_id,
                synonyms,
                title=title,
                abstract=abstract,
            )
            total_chunks = len(shared_chunks)

            function_chunk_summaries = []
            skip_reasons = []
            for chunk_index, text_chunk in enumerate(shared_chunks, start=1):
                chunk_summary = summarize_reference_function(
                    text_chunk,
                    fly_symbol,
                    fbgn_id,
                    synonyms,
                    title=title,
                    abstract=abstract,
                    chunk_index=chunk_index,
                    total_chunks=total_chunks,
                )
                if chunk_summary.get("skip_reference", True):
                    if chunk_summary.get("skip_reason"):
                        skip_reasons.append(chunk_summary["skip_reason"])
                else:
                    function_chunk_summaries.append(chunk_summary)
            
            total_attempts += 1

            if function_chunk_summaries:
                if len(function_chunk_summaries) == 1:
                    final_function_summary = {
                        "function": _clean_text(function_chunk_summaries[0].get("function", "")),
                        "phenotypes": _clean_text(function_chunk_summaries[0].get("phenotypes", "")),
                    }
                else:
                    final_function_summary = rewrite_function_summary(
                        fly_symbol,
                        fbgn_id,
                        function_chunk_summaries,
                    )

                reagent_records = []
                for chunk_index, text_chunk in enumerate(shared_chunks, start=1):
                    reagent_records.extend(
                        extract_reference_reagents(
                            text_chunk,
                            fly_symbol,
                            fbgn_id,
                            synonyms,
                            title=title,
                            abstract=abstract,
                            chunk_index=chunk_index,
                            total_chunks=total_chunks,
                        )
                    )
                merged_reagents = _merge_deduplicated_reagents(reagent_records)
                reagent_pairs_text = _format_reagent_pairs(merged_reagents)
                summary = _format_reference_summary(
                    final_function_summary.get("function", ""),
                    final_function_summary.get("phenotypes", ""),
                )
                is_high_quality = True
                qc_justification = "Passed QC"
            else:
                final_function_summary = {"function": "", "phenotypes": ""}
                merged_reagents = []
                reagent_pairs_text = ""
                summary = "No evidence found, skip reference"
                is_high_quality = False
                qc_justification = (
                    _join_unique_texts(skip_reasons)
                    or f"No evidence found across {total_chunks} chunk(s)"
                )
            
            ref_source = _format_ref_source(pmc)
            
            # Store summary
            all_summaries.append({
                "paper_id": pmc,
                "title": title,
                "year": year,
                "journal": journal,
                "authors": authors,
                "gene_symbol": fly_symbol,
                "flybase_id": fbgn_id,
                "summary": summary,
                "function_text": final_function_summary.get("function", ""),
                "phenotypes_text": final_function_summary.get("phenotypes", ""),
                "reagent_pairs": reagent_pairs_text,
                "reagents": merged_reagents,
                "abstract_text": abstract or "",
                "is_high_quality": is_high_quality,
                "qc_justification": qc_justification,
                "source": ref_source
            })
            
            fbgn_to_pmcids_all[fbgn_id].add(pmc)
            
            if is_high_quality:
                high_q_count += 1
                fbgn_to_summary_chunks[fbgn_id].append(summary)
                fbgn_high_quality_ref_summaries[fbgn_id].append(pmc)
                print(f"    [{high_q_count}/6] {pmc}: {qc_justification}")
        
        # Classify gene
        chunks = fbgn_to_summary_chunks.get(fbgn_id, [])
        all_high_quality_pmcids = fbgn_to_pmcids_all.get(fbgn_id, set())
        supporting_pmcids = set(fbgn_high_quality_ref_summaries.get(fbgn_id, []))
        full_text_pmcids = fbgn_to_full_text_pmcids.get(fbgn_id, set())
        
        if chunks:
            agg_text = "\n\n".join(chunks)
            
            cls = classify_gene_from_text(fly_symbol, keywords_list, agg_text)
            if not cls or not isinstance(cls, dict):
                cls = {"gene": fly_symbol, "category": [], "confidence": 0, "rationale": "Classification failed"}
            
            category_list = cls.get("category", [])
            if isinstance(category_list, list):
                category_str = "; ".join(category_list) if category_list else "None"
            else:
                category_str = str(category_list) if category_list else "None"
            
            fly_gene_hits[gene_symbol] = {
                "pmcids": all_high_quality_pmcids,
                "category": category_str,
                "confidence": int(cls.get("confidence", 0)),
                "rationale": cls.get("rationale", ""),
                "classified_by": "GPT",
                "supporting_refs": supporting_pmcids,
                "full_text_refs": full_text_pmcids
            }
            print(f"    Classification: {category_str} (confidence: {cls.get('confidence', 0)})")
        else:
            fly_gene_hits[gene_symbol] = {
                "pmcids": all_high_quality_pmcids,
                "category": "None",
                "confidence": 0,
                "rationale": "No relevant references found.",
                "classified_by": "Self",
                "supporting_refs": set(),
                "full_text_refs": full_text_pmcids
            }
            print(f"    Classification: None (no relevant references)")
    
    return fly_gene_hits, all_summaries, validated


def generate_excel_output(fly_gene_hits, all_summaries, validated, output_path, gene_set_df=None):
    """Generate Excel output with Gene Set, Classification, Reference Summaries and Reagents sheets."""
    import pandas as pd
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    
    fbgn_to_symbol = build_fbgn_to_symbol_map()
    pmcid_to_year = build_pmcid_to_year()
    
    def _sort_pmcids_by_recency(pmcids_set):
        return sorted(pmcids_set, key=lambda p: (
            -pmcid_to_year.get(p.upper(), 0),
            p
        ))
    
    # Classification sheet
    class_rows = []
    for fly_sym, fbgn in validated.items():
        info = fly_gene_hits.get(fly_sym, {})
        supporting_refs = info.get("supporting_refs", set())
        full_text_refs = info.get("full_text_refs", set())
        queried_pmcids = info.get("pmcids", set())
        class_rows.append({
            "Fly_Gene_symbol": fly_sym,
            "FlyBase_ID": fbgn,
            "Category": info.get("category", "None"),
            "Confidence": info.get("confidence", 0),
            "Rationale": info.get("rationale", ""),
            "Supporting_Refs": ";".join(_sort_pmcids_by_recency(supporting_refs)),
            "Papers_w_full_text": ";".join(_sort_pmcids_by_recency(full_text_refs)),
            "Queried_PMCIDs": ";".join(_sort_pmcids_by_recency(queried_pmcids)),
        })
    
    class_cols = [
        "Fly_Gene_symbol", "FlyBase_ID", "Category", "Confidence", "Rationale",
        "Supporting_Refs", "Papers_w_full_text", "Queried_PMCIDs"
    ]
    class_df = pd.DataFrame(class_rows, columns=class_cols)
    
    # Sort by category (None last), then by confidence descending
    category_counts = class_df["Category"].value_counts().to_dict()
    sorted_categories = sorted(
        category_counts.keys(),
        key=lambda cat: (
            1 if cat == "None" else 0,
            -category_counts.get(cat, 0),
            cat
        )
    )
    category_order = {cat: idx for idx, cat in enumerate(sorted_categories)}
    class_df["_cat_order"] = class_df["Category"].map(
        lambda x: category_order.get(x, len(sorted_categories))
    )
    class_df = class_df.sort_values(
        by=["_cat_order", "Confidence"],
        ascending=[True, False]
    ).reset_index(drop=True)
    class_df = class_df.drop(columns=["_cat_order"])
    
    # Reference Summaries sheet
    ref_rows = []
    reagent_rows = []
    for sd in all_summaries or []:
        fly_sym = sd.get("gene_symbol", "")
        fbgn = sd.get("flybase_id", "")
        
        if not sd.get("is_high_quality", False):
            continue
        
        cls_info = fly_gene_hits.get(fly_sym, {})
        cls = cls_info.get("category", "None")
        
        if cls == "None" or cls == "Potential Hits":
            continue
        
        year_str = sd.get("year", "")
        try:
            year_int = int(year_str) if year_str and year_str != "n.d." else 0
        except:
            year_int = 0
        
        ref_rows.append({
            "Gene": fly_sym,
            "Gene_ID": fbgn,
            "Paper_ID": sd.get("paper_id", ""),
            "Title": sd.get("title", ""),
            "Year": year_str,
            "_year_sort": year_int,
            "Journal": sd.get("journal", ""),
            "Author(s)": "; ".join(sd.get("authors", []) or []),
            "Reference Summary": sd.get("summary", ""),
            "Abstract": sd.get("abstract_text", ""),
            "Classification": cls,
            "Source": sd.get("source", "Unknown"),
            "Reagent_Pairs": sd.get("reagent_pairs", ""),
        })

        for reagent in sd.get("reagents", []) or []:
            reagent_rows.append({
                "Gene": fly_sym,
                "Paper_ID": sd.get("paper_id", ""),
                "Title": sd.get("title", ""),
                "Stock_ID": reagent.get("stock_id", ""),
                "Collection": reagent.get("collection", ""),
                "Reagent_Type": reagent.get("reagent_type", ""),
                "Functional_Validity": reagent.get("functional_validity", ""),
                "Evidence_Snippet": reagent.get("evidence_snippet", ""),
                "Reagent_Name": reagent.get("reagent_name", ""),
                "Source": sd.get("source", "Unknown"),
            })
    
    ref_df = pd.DataFrame(ref_rows, columns=[
        "Gene", "Gene_ID", "Paper_ID", "Title", "Year", "_year_sort",
        "Journal", "Author(s)", "Reference Summary", "Abstract", "Classification", "Source",
        "Reagent_Pairs"
    ])
    reagent_df = pd.DataFrame(reagent_rows, columns=[
        "Gene", "Paper_ID", "Title", "Stock_ID", "Collection", "Reagent_Type", "Functional_Validity", "Evidence_Snippet",
        "Reagent_Name", "Source"
    ])
    
    # Sort references to mirror classification order
    if not ref_df.empty and not class_df.empty:
        gene_order = {gene: idx for idx, gene in enumerate(class_df["Fly_Gene_symbol"].tolist())}
        ref_df["_gene_order"] = ref_df["Gene"].map(lambda x: gene_order.get(x, 999999))
        ref_df = ref_df.sort_values(by=["_gene_order", "_year_sort"], ascending=[True, False])
        ref_df = ref_df.drop(columns=["_gene_order"])
        if not reagent_df.empty:
            reagent_df["_gene_order"] = reagent_df["Gene"].map(lambda x: gene_order.get(x, 999999))
            reagent_df = reagent_df.sort_values(
                by=["_gene_order", "Collection", "Stock_ID"],
                ascending=[True, True, True]
            )
            reagent_df = reagent_df.drop(columns=["_gene_order"])
    
    if "_year_sort" in ref_df.columns:
        ref_df = ref_df.drop(columns=["_year_sort"])
    # Write Excel
    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        gene_df = gene_set_df.copy() if isinstance(gene_set_df, pd.DataFrame) else pd.DataFrame()
        gene_df.to_excel(writer, index=False, sheet_name="Gene Set")
        class_df.to_excel(writer, index=False, sheet_name="Classification")
        ref_df.to_excel(writer, index=False, sheet_name="Reference Summaries")
        reagent_df.to_excel(writer, index=False, sheet_name="Reagents")
        
        # Format Gene Set sheet
        gene_ws = writer.sheets["Gene Set"]
        for row in gene_ws.iter_rows(min_row=1, max_row=gene_ws.max_row, min_col=1, max_col=gene_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Format Classification sheet
        class_ws = writer.sheets["Classification"]
        class_ws.column_dimensions[get_column_letter(1)].width = 18
        class_ws.column_dimensions[get_column_letter(2)].width = 15
        class_ws.column_dimensions[get_column_letter(3)].width = 15
        class_ws.column_dimensions[get_column_letter(4)].width = 12
        class_ws.column_dimensions[get_column_letter(5)].width = 50
        class_ws.column_dimensions[get_column_letter(6)].width = 30
        class_ws.column_dimensions[get_column_letter(7)].width = 30
        class_ws.column_dimensions[get_column_letter(8)].width = 30
        class_ws.column_dimensions[get_column_letter(9)].width = 50
        
        # Color palette for categories
        color_palette = [
            ("5B9BD5", True), ("FFC000", False), ("70AD47", False),
            ("ED7D31", False), ("9E480E", True), ("00B0F0", False),
            ("FF6B6B", False), ("4ECDC4", False), ("95E1D3", False),
            ("F38181", False), ("7030A0", True), ("FCBAD3", False),
        ]
        
        unique_categories = [cat for cat in class_df["Category"].unique() if cat != "None"]
        category_colors = {}
        category_needs_white_text = {}
        for idx, category in enumerate(unique_categories):
            color_hex, needs_white = color_palette[idx % len(color_palette)]
            category_colors[category] = color_hex
            category_needs_white_text[category] = needs_white
        
        # Apply formatting
        for row in class_ws.iter_rows(min_row=1, max_row=class_ws.max_row, min_col=1, max_col=class_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Apply category colors
        for row_idx in range(2, class_ws.max_row + 1):
            category_cell = class_ws.cell(row=row_idx, column=3)
            category_value = category_cell.value
            if category_value and category_value != "None" and category_value in category_colors:
                color_hex = category_colors[category_value]
                category_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                if category_needs_white_text.get(category_value, False):
                    category_cell.font = Font(color="FFFFFF", bold=True)
        
        # Format Reference Summaries sheet
        ref_ws = writer.sheets["Reference Summaries"]
        ref_ws.column_dimensions[get_column_letter(1)].width = 15
        ref_ws.column_dimensions[get_column_letter(2)].width = 15
        ref_ws.column_dimensions[get_column_letter(3)].width = 12
        ref_ws.column_dimensions[get_column_letter(4)].width = 50
        ref_ws.column_dimensions[get_column_letter(5)].width = 8
        ref_ws.column_dimensions[get_column_letter(6)].width = 25
        ref_ws.column_dimensions[get_column_letter(7)].width = 30
        ref_ws.column_dimensions[get_column_letter(8)].width = 50
        ref_ws.column_dimensions[get_column_letter(9)].width = 50
        ref_ws.column_dimensions[get_column_letter(10)].width = 15
        ref_ws.column_dimensions[get_column_letter(11)].width = 12
        ref_ws.column_dimensions[get_column_letter(12)].width = 35
        
        for row in ref_ws.iter_rows(min_row=1, max_row=ref_ws.max_row, min_col=1, max_col=ref_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Format Reagents sheet
        reagent_ws = writer.sheets["Reagents"]
        reagent_widths = {
            1: 15,
            2: 12,
            3: 45,
            4: 15,
            5: 15,
            6: 18,
            7: 22,
            8: 60,
            9: 25,
            10: 15,
        }
        for col_idx, width in reagent_widths.items():
            reagent_ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row in reagent_ws.iter_rows(min_row=1, max_row=reagent_ws.max_row, min_col=1, max_col=reagent_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Write to file
    with open(output_path, 'wb') as f:
        f.write(excel_buf.getvalue())
    
    print(f"\nOutput written to: {output_path}")
    print(f"  - Classification sheet: {len(class_df)} genes")
    print(f"  - Reference Summaries sheet: {len(ref_df)} entries")
    print(f"  - Reagents sheet: {len(reagent_df)} entries")


def _json_dump(path: Path, payload: dict):
    """Atomically write JSON payload to disk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2, sort_keys=True)
    os.replace(tmp_path, path)


def _json_load(path: Path, default: Any):
    """Load JSON payload, returning default on missing/invalid content."""
    if not path.exists():
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def _normalize_keywords_for_key(keywords_list: list[str]) -> str:
    cleaned = sorted({str(k).strip().lower() for k in (keywords_list or []) if str(k).strip()})
    return ",".join(cleaned)


def _make_gene_store_key(fbgn_id: str, keywords_list: list[str], reference_limit: int) -> str:
    kw_norm = _normalize_keywords_for_key(keywords_list)
    kw_hash = hashlib.sha256(kw_norm.encode("utf-8")).hexdigest()[:16]
    return f"{str(fbgn_id).strip()}|{reference_limit}|{kw_hash}"


def _make_input_fingerprint(fbgn_ids: list[str], keywords_list: list[str], reference_limit: int) -> str:
    normalized_ids = sorted({str(x).strip() for x in fbgn_ids if str(x).strip()})
    payload = {
        "fbgn_ids": normalized_ids,
        "keywords": _normalize_keywords_for_key(keywords_list),
        "reference_limit": int(reference_limit),
    }
    src = json.dumps(payload, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(src.encode("utf-8")).hexdigest()


def _get_batch_root(input_directory: str) -> Path:
    return Path(input_directory) / BATCH_STATE_DIRNAME


def _get_csv_state_dir(input_directory: str, csv_path: str) -> Path:
    csv_abs = str(Path(csv_path).resolve())
    csv_hash = hashlib.sha1(csv_abs.encode("utf-8")).hexdigest()[:10]
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(csv_path).stem)
    return _get_batch_root(input_directory) / f"{safe_stem}_{csv_hash}"


def _serialize_hit_info(hit_info: dict) -> dict:
    out = dict(hit_info or {})
    for set_field in ("pmcids", "supporting_refs", "full_text_refs"):
        val = out.get(set_field, set())
        if isinstance(val, set):
            out[set_field] = sorted(list(val))
        elif isinstance(val, list):
            out[set_field] = sorted(list({str(x) for x in val if str(x)}))
        else:
            out[set_field] = []
    out["confidence"] = int(out.get("confidence", 0) or 0)
    out["category"] = str(out.get("category", "None") or "None")
    out["rationale"] = str(out.get("rationale", "") or "")
    out["classified_by"] = str(out.get("classified_by", "Self") or "Self")
    return out


def _deserialize_hit_info(hit_info: dict) -> dict:
    out = dict(hit_info or {})
    for set_field in ("pmcids", "supporting_refs", "full_text_refs"):
        val = out.get(set_field, [])
        if isinstance(val, set):
            out[set_field] = set(val)
        else:
            out[set_field] = set([str(x) for x in (val or []) if str(x)])
    out["confidence"] = int(out.get("confidence", 0) or 0)
    out["category"] = str(out.get("category", "None") or "None")
    out["rationale"] = str(out.get("rationale", "") or "")
    out["classified_by"] = str(out.get("classified_by", "Self") or "Self")
    return out


def _build_gene_records(
    fly_gene_hits: dict,
    all_summaries: list[dict],
    validated: dict[str, str],
) -> list[dict]:
    summaries_by_fbgn: dict[str, list[dict]] = defaultdict(list)
    for summary in all_summaries or []:
        fbgn = str(summary.get("flybase_id", "")).strip()
        if fbgn:
            summaries_by_fbgn[fbgn].append(summary)

    records: list[dict] = []
    for gene_symbol, fbgn_id in (validated or {}).items():
        info = fly_gene_hits.get(gene_symbol, {})
        records.append({
            "fbgn_id": str(fbgn_id),
            "gene_symbol": str(gene_symbol),
            "hit_info": _serialize_hit_info(info),
            "summaries": summaries_by_fbgn.get(str(fbgn_id), []),
        })
    return records


def _merge_gene_records(records: list[dict]) -> tuple[dict, list[dict], dict]:
    fly_gene_hits: dict[str, dict] = {}
    all_summaries: list[dict] = []
    validated: dict[str, str] = {}
    for record in records or []:
        gene_symbol = str(record.get("gene_symbol", "")).strip()
        fbgn_id = str(record.get("fbgn_id", "")).strip()
        if not gene_symbol or not fbgn_id:
            continue
        validated[gene_symbol] = fbgn_id
        fly_gene_hits[gene_symbol] = _deserialize_hit_info(record.get("hit_info", {}))
        all_summaries.extend(record.get("summaries", []) or [])
    return fly_gene_hits, all_summaries, validated


def _load_run_store(input_directory: str) -> dict:
    run_store_path = _get_batch_root(input_directory) / RUN_STORE_FILENAME
    store = _json_load(run_store_path, default={"version": 1, "genes": {}})
    if not isinstance(store, dict):
        return {"version": 1, "genes": {}}
    if "genes" not in store or not isinstance(store["genes"], dict):
        store["genes"] = {}
    store["version"] = 1
    return store


def _save_run_store(input_directory: str, run_store: dict):
    run_store_path = _get_batch_root(input_directory) / RUN_STORE_FILENAME
    _json_dump(run_store_path, run_store)


def process_csv_file(
    csv_path: str,
    keywords_list: list[str],
    reference_limit: int,
    input_directory: str,
    run_store: dict,
    force_all: bool = False,
):
    """Process a single CSV file containing flybase_gene_id column."""
    print(f"\n{'#'*70}")
    print(f"Processing file: {csv_path}")
    print(f"{'#'*70}")
    
    # Read CSV
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return False
    
    # Check for flybase_gene_id column
    if 'flybase_gene_id' not in df.columns:
        print(f"Skipping: No 'flybase_gene_id' column found")
        return False
    
    # Extract FBgn IDs
    fbgn_ids = df['flybase_gene_id'].dropna().unique().tolist()
    fbgn_ids = [str(x).strip() for x in fbgn_ids if str(x).strip().startswith('FBgn')]
    
    if not fbgn_ids:
        print(f"Skipping: No valid FBgn IDs found")
        return False
    
    print(f"Found {len(fbgn_ids)} unique FBgn IDs")

    input_fingerprint = _make_input_fingerprint(fbgn_ids, keywords_list, reference_limit)
    state_dir = _get_csv_state_dir(input_directory, csv_path)
    state_path = state_dir / "state.json"
    state_dir.mkdir(parents=True, exist_ok=True)

    state_default = {
        "version": 1,
        "csv_path": str(Path(csv_path).resolve()),
        "input_fingerprint": input_fingerprint,
        "total_batches": 0,
        "completed_batches": [],
        "batch_files": {},
        "processed_genes": 0,
        "reused_genes": 0,
    }
    state = _json_load(state_path, default=state_default)
    if not isinstance(state, dict):
        state = dict(state_default)

    if force_all:
        print("Force-all mode enabled: reprocessing all batches for this file.")
        state = dict(state_default)
    elif state.get("input_fingerprint") != input_fingerprint:
        print("Input fingerprint changed for this file; restarting file-specific batch state.")
        state = dict(state_default)

    batches = [fbgn_ids[i:i + BATCH_SIZE] for i in range(0, len(fbgn_ids), BATCH_SIZE)]
    state["total_batches"] = len(batches)
    state["csv_path"] = str(Path(csv_path).resolve())
    state["input_fingerprint"] = input_fingerprint
    completed_batches = set(int(x) for x in (state.get("completed_batches", []) or []))

    if completed_batches and not force_all:
        print(f"Resuming: {len(completed_batches)}/{len(batches)} batches already complete.")

    for batch_index, batch_fbgn_ids in enumerate(batches):
        if (batch_index in completed_batches) and not force_all:
            print(f"  Skipping completed batch {batch_index + 1}/{len(batches)}")
            continue

        print(f"\n[Batch {batch_index + 1}/{len(batches)}] size={len(batch_fbgn_ids)}")
        try:
            dedup_hits = 0
            to_process: list[str] = []
            reused_records: list[dict] = []
            for fbgn_id in batch_fbgn_ids:
                store_key = _make_gene_store_key(fbgn_id, keywords_list, reference_limit)
                if (not force_all) and store_key in run_store.get("genes", {}):
                    reused_records.append(run_store["genes"][store_key])
                    dedup_hits += 1
                else:
                    to_process.append(fbgn_id)

            processed_records: list[dict] = []
            if to_process:
                fly_gene_hits, all_summaries, validated = process_gene_set(
                    to_process, keywords_list, reference_limit
                )
                processed_records = _build_gene_records(fly_gene_hits, all_summaries, validated)

                for record in processed_records:
                    store_key = _make_gene_store_key(
                        record.get("fbgn_id", ""),
                        keywords_list,
                        reference_limit,
                    )
                    run_store.setdefault("genes", {})[store_key] = record
                _save_run_store(input_directory, run_store)

            record_by_fbgn: dict[str, dict] = {}
            for rec in reused_records + processed_records:
                fbgn_id = str(rec.get("fbgn_id", "")).strip()
                if fbgn_id:
                    record_by_fbgn[fbgn_id] = rec
            batch_records = [record_by_fbgn[fbgn] for fbgn in batch_fbgn_ids if fbgn in record_by_fbgn]

            batch_filename = f"batch_{batch_index:05d}.json"
            batch_path = state_dir / batch_filename
            batch_payload = {
                "version": 1,
                "batch_index": batch_index,
                "fbgn_ids": batch_fbgn_ids,
                "records": batch_records,
                "processed_count": len(processed_records),
                "reused_count": dedup_hits,
            }
            _json_dump(batch_path, batch_payload)

            completed_batches.add(batch_index)
            state["completed_batches"] = sorted(list(completed_batches))
            state.setdefault("batch_files", {})[str(batch_index)] = batch_filename
            state["processed_genes"] = int(state.get("processed_genes", 0) or 0) + len(processed_records)
            state["reused_genes"] = int(state.get("reused_genes", 0) or 0) + dedup_hits
            _json_dump(state_path, state)

            print(
                f"  Batch complete: processed={len(processed_records)} "
                f"reused={dedup_hits}"
            )
        except Exception as e:
            print(f"Error in batch {batch_index + 1}: {e}")
            return False

    # Aggregate all completed batch artifacts for this file.
    all_records: list[dict] = []
    for batch_index in range(len(batches)):
        batch_filename = state.get("batch_files", {}).get(str(batch_index), f"batch_{batch_index:05d}.json")
        batch_path = state_dir / batch_filename
        if not batch_path.exists():
            print(f"Error: Missing expected batch artifact {batch_path}")
            return False
        batch_payload = _json_load(batch_path, default={})
        if not isinstance(batch_payload, dict):
            print(f"Error: Invalid batch artifact format: {batch_path}")
            return False
        all_records.extend(batch_payload.get("records", []) or [])

    fly_gene_hits, all_summaries, validated = _merge_gene_records(all_records)
    if not validated:
        print("No genes were successfully validated")
        return False

    # Generate one final output Excel per input CSV.
    base_name = os.path.splitext(csv_path)[0]
    output_path = f"{base_name}_classification.xlsx"
    
    generate_excel_output(fly_gene_hits, all_summaries, validated, output_path, gene_set_df=df)
    
    return True


def run_fbgnid_conversion(input_directory: str, input_gene_col: str = "ext_gene") -> bool:
    """
    Run the same FBgn conversion helper used in fly_stocker_v2-style workflows.
    """
    if not GET_FBGN_IDS_SCRIPT.exists():
        print(f"Error: FBgnID conversion script not found at: {GET_FBGN_IDS_SCRIPT}")
        return False

    cmd = [sys.executable, str(GET_FBGN_IDS_SCRIPT), str(input_directory), str(input_gene_col)]
    print(f"\nRunning FBgnID conversion: {' '.join(cmd)}")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=False)
        if result.stdout:
            print(result.stdout)
        if result.stderr:
            print(result.stderr)
        if result.returncode != 0:
            print(f"FBgnID conversion failed with exit code {result.returncode}")
            return False
        return True
    except Exception as e:
        print(f"Error running FBgnID conversion: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Fly gene classification and reagent-finder pipeline - converts input symbols "
            "to FBgn IDs, summarizes literature, classifies genes, and extracts reagent hits"
        )
    )
    parser.add_argument(
        "input_directory",
        help="Directory containing input CSV files for FBgn conversion, classification, and reagent extraction"
    )
    parser.add_argument(
        "--keywords", "-k",
        default="",
        help="Comma-separated list of classification keywords (e.g., 'circadian,sleep,rhythm')"
    )
    parser.add_argument(
        "--reference-limit", "-r",
        type=int,
        default=500,
        help="Maximum number of references to consider per gene (default: 500)"
    )
    parser.add_argument(
        "--input-gene-col",
        default="ext_gene",
        help="Input gene symbol column for FBgnID conversion (default: ext_gene)."
    )
    parser.add_argument(
        "--force-all",
        action="store_true",
        help="Ignore checkpoint/resume state and reprocess all genes from scratch."
    )
    
    args = parser.parse_args()
    
    # Validate input directory
    if not os.path.isdir(args.input_directory):
        print(f"Error: '{args.input_directory}' is not a valid directory")
        sys.exit(1)
    
    # Validate FlyBase data directory exists
    if not FLYBASE_DATA.exists():
        print(f"Error: FlyBase data directory not found at: {FLYBASE_DATA}")
        print(f"Expected structure:")
        print(f"  {FLYBASE_DATA}/Genes/fb_synonym_fb_*.tsv")
        print(f"  {FLYBASE_DATA}/FlyBase_References/entity_publication_fb_*.tsv")
        print(f"  {FLYBASE_DATA}/FlyBase_References/fbrf_pmid_pmcid_doi_fb_*.tsv")
        sys.exit(1)
    
    # Parse keywords
    keywords_list = [k.strip() for k in args.keywords.split(",") if k.strip()]
    
    print(f"\n{'='*70}")
    print("Fly Gene Classification + Reagent Finder Pipeline (Standalone)")
    print(f"{'='*70}")
    print(f"Input directory: {args.input_directory}")
    print(f"Keywords: {keywords_list if keywords_list else '(none)'}")
    print(f"Reference limit: {args.reference_limit}")
    print(f"Batch size: {BATCH_SIZE} (automatic)")
    print(f"Force all: {'yes' if args.force_all else 'no'}")
    print("Run FBgnID conversion: yes (default)")
    print(f"FBgn conversion input gene column: {args.input_gene_col}")

    # Fail fast on API misconfiguration so runs do not silently produce empty outputs.
    try:
        _get_openai_client()
    except Exception as e:
        print(f"Error: OpenAI API setup invalid: {e}")
        sys.exit(1)
    print(f"OpenAI model: {_get_openai_model()}")

    # Step 0: convert input symbols to FBgn IDs.
    ok = run_fbgnid_conversion(args.input_directory, args.input_gene_col)
    if not ok:
        print("Error: FBgnID conversion failed; aborting.")
        sys.exit(1)
    
    # Find CSV files
    csv_files = [
        os.path.join(args.input_directory, f)
        for f in os.listdir(args.input_directory)
        if f.endswith('.csv') and not f.endswith('_classification.csv')
    ]
    
    if not csv_files:
        print(f"\nNo CSV files found in {args.input_directory}")
        sys.exit(1)
    
    print(f"\nFound {len(csv_files)} CSV files to process")
    batch_root = _get_batch_root(args.input_directory)
    batch_root.mkdir(parents=True, exist_ok=True)
    run_store = _load_run_store(args.input_directory)
    
    # Process each file
    success_count = 0
    for csv_path in sorted(csv_files):
        try:
            if process_csv_file(
                csv_path,
                keywords_list,
                args.reference_limit,
                args.input_directory,
                run_store,
                force_all=args.force_all,
            ):
                success_count += 1
            else:
                print(f"\nError processing {csv_path}; exiting early to preserve checkpoints.")
                sys.exit(1)
        except Exception as e:
            print(f"\nError processing {csv_path}: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    # Persist shared caches for future runs.
    _save_pmid_title_abstract_cache_pending()
    _save_fulltext_method_cache_pending()
    
    print(f"\n{'='*70}")
    print(f"Complete! Successfully processed {success_count}/{len(csv_files)} files")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
