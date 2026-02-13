#!/usr/bin/env python3
"""Backfill/repair Reference Summaries metadata in classification workbooks."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from metapub import PubMedFetcher
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from HelperScripts.metadata_resolver import normalize_authors, resolve_reference_metadata


_TURBO_CANDIDATES = [Path("/Volumes/umms-rallada"), Path("/nfs/turbo/umms-rallada")]
TURBO_ROOT = next((p for p in _TURBO_CANDIDATES if p.exists()), _TURBO_CANDIDATES[0])
DEFAULT_PUBMED_CACHE = TURBO_ROOT / "UM Lab Users" / "Aadish" / "Data" / "PubMed Cache" / "pmid_to_title_abstract.csv"
CANONICAL_CACHE_COLS = [
    "pmid", "pmcid", "title", "abstract", "year", "journal",
    "authors", "doi", "source", "updated_at"
]


def _load_cache(cache_path: Path) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    if not cache_path.exists():
        return out
    try:
        df = pd.read_csv(cache_path, dtype=str, keep_default_na=False)
    except Exception:
        return out
    drop_cols = [c for c in df.columns if str(c).startswith("Unnamed")]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    if "pmid" not in df.columns:
        return out
    for col in CANONICAL_CACHE_COLS:
        if col not in df.columns:
            df[col] = ""
    df = df[CANONICAL_CACHE_COLS]
    for _, row in df.iterrows():
        pmid = str(row.get("pmid", "") or "").strip()
        if not pmid.isdigit():
            continue
        out[pmid] = {
            "pmcid": str(row.get("pmcid", "") or ""),
            "title": str(row.get("title", "") or ""),
            "abstract": str(row.get("abstract", "") or ""),
            "year": str(row.get("year", "") or ""),
            "journal": str(row.get("journal", "") or ""),
            "authors": normalize_authors(row.get("authors", "")),
            "doi": str(row.get("doi", "") or ""),
            "source": str(row.get("source", "") or ""),
            "updated_at": str(row.get("updated_at", "") or ""),
        }
    return out


def _save_cache(cache_path: Path, cache_data: Dict[str, dict]) -> None:
    rows = []
    for pmid in sorted(cache_data.keys()):
        meta = cache_data[pmid] or {}
        rows.append({
            "pmid": str(pmid),
            "pmcid": str(meta.get("pmcid", "") or ""),
            "title": str(meta.get("title", "") or ""),
            "abstract": str(meta.get("abstract", "") or ""),
            "year": str(meta.get("year", "") or ""),
            "journal": str(meta.get("journal", "") or ""),
            "authors": "; ".join(normalize_authors(meta.get("authors", []))),
            "doi": str(meta.get("doi", "") or ""),
            "source": str(meta.get("source", "") or ""),
            "updated_at": str(meta.get("updated_at", "") or ""),
        })
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows, columns=CANONICAL_CACHE_COLS).to_csv(cache_path, index=False)


def _iter_workbooks(input_dir: Path):
    for p in sorted(input_dir.rglob("*_classification.xlsx")):
        if p.is_file():
            yield p


def run_backfill(
    input_dir: Path,
    apply: bool,
    cache_path: Path,
    report_path: Optional[Path],
    api_key: str,
) -> dict:
    cache_data = _load_cache(cache_path)
    fetcher = PubMedFetcher(email="aadish98@gmail.com", api_key=api_key or None)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    def cache_getter(pmid: str) -> Optional[dict]:
        entry = cache_data.get(str(pmid).strip())
        if not entry:
            return None
        return {
            "title": entry.get("title", ""),
            "abstract": entry.get("abstract", ""),
            "year": entry.get("year", ""),
            "journal": entry.get("journal", ""),
            "authors": entry.get("authors", []),
            "doi": entry.get("doi", ""),
            "pmcid": entry.get("pmcid", ""),
            "source": entry.get("source", "cache"),
            "updated_at": entry.get("updated_at", ""),
        }

    def cache_setter(
        pmid: str,
        title: str,
        abstract: str,
        *,
        year: str = "",
        journal: str = "",
        authors=None,
        doi: str = "",
        pmcid: str = "",
        source: str = "",
        updated_at: str = "",
    ):
        pmid_clean = str(pmid or "").strip()
        if not pmid_clean.isdigit():
            return
        prior = cache_data.get(pmid_clean, {})
        cache_data[pmid_clean] = {
            "title": title or prior.get("title", ""),
            "abstract": abstract or prior.get("abstract", ""),
            "year": year or prior.get("year", ""),
            "journal": journal or prior.get("journal", ""),
            "authors": normalize_authors(authors or prior.get("authors", [])),
            "doi": doi or prior.get("doi", ""),
            "pmcid": pmcid or prior.get("pmcid", ""),
            "source": source or prior.get("source", "metapub"),
            "updated_at": updated_at or prior.get("updated_at", now),
        }

    report = {
        "generated_at": now,
        "input_dir": str(input_dir),
        "mode": "apply" if apply else "dry-run",
        "files": [],
        "totals": {
            "files_seen": 0,
            "files_updated": 0,
            "rows_seen": 0,
            "rows_changed": 0,
            "rows_unresolved": 0,
        },
    }

    for wb_path in _iter_workbooks(input_dir):
        file_report = {
            "file": str(wb_path),
            "rows_seen": 0,
            "rows_changed": 0,
            "rows_unresolved": 0,
            "samples": [],
        }
        report["totals"]["files_seen"] += 1

        wb = load_workbook(wb_path)
        if "Reference Summaries" not in wb.sheetnames:
            file_report["error"] = "Missing Reference Summaries sheet"
            report["files"].append(file_report)
            continue

        ws = wb["Reference Summaries"]
        headers = {str(ws.cell(row=1, column=c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        required = ["Paper_ID", "Journal", "Author(s)"]
        if any(h not in headers for h in required):
            file_report["error"] = "Missing required columns"
            report["files"].append(file_report)
            continue

        c_paper = headers["Paper_ID"]
        c_journal = headers["Journal"]
        c_authors = headers["Author(s)"]

        for row_idx in range(2, ws.max_row + 1):
            paper_id = str(ws.cell(row=row_idx, column=c_paper).value or "").strip()
            if not paper_id:
                continue
            file_report["rows_seen"] += 1
            report["totals"]["rows_seen"] += 1

            old_journal = str(ws.cell(row=row_idx, column=c_journal).value or "").strip()
            old_authors = str(ws.cell(row=row_idx, column=c_authors).value or "").strip()

            resolved = resolve_reference_metadata(
                paper_id,
                fetcher,
                api_key=api_key or "",
                cache_getter=cache_getter,
                cache_setter=cache_setter,
            )
            new_journal = str(resolved.get("journal", "") or "").strip() or old_journal
            new_authors = str(resolved.get("authors_display", "") or "").strip() or old_authors

            if not new_journal and not new_authors:
                file_report["rows_unresolved"] += 1
                report["totals"]["rows_unresolved"] += 1
                continue

            changed = (new_journal != old_journal) or (new_authors != old_authors)
            if changed:
                file_report["rows_changed"] += 1
                report["totals"]["rows_changed"] += 1
                if len(file_report["samples"]) < 8:
                    file_report["samples"].append({
                        "paper_id": paper_id,
                        "old_journal": old_journal,
                        "new_journal": new_journal,
                        "old_authors": old_authors,
                        "new_authors": new_authors,
                    })
            if apply:
                ws.cell(row=row_idx, column=c_journal).value = new_journal
                ws.cell(row=row_idx, column=c_authors).value = new_authors

        if apply and file_report["rows_changed"] > 0:
            wb.save(wb_path)
            report["totals"]["files_updated"] += 1
        report["files"].append(file_report)

    if apply:
        _save_cache(cache_path, cache_data)

    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=True, indent=2)

    return report


def main():
    parser = argparse.ArgumentParser(description="Backfill Journal/Author(s) metadata in classification workbooks.")
    parser.add_argument("input_dir", help="Directory containing *_classification.xlsx files.")
    parser.add_argument(
        "--cache-path",
        default=str(DEFAULT_PUBMED_CACHE),
        help="Path to PMID metadata cache CSV (default: shared PubMed cache).",
    )
    parser.add_argument(
        "--report-path",
        default="Logs/backfill_reference_metadata_report.json",
        help="Where to write JSON report.",
    )
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument("--apply", action="store_true", help="Apply changes in-place.")
    mode_group.add_argument("--dry-run", action="store_true", help="Preview changes only (default).")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    if not input_dir.exists():
        raise SystemExit(f"Input directory does not exist: {input_dir}")

    api_key = os.getenv("NCBI_API_KEY", "")
    report = run_backfill(
        input_dir=input_dir,
        apply=bool(args.apply),
        cache_path=Path(args.cache_path).expanduser(),
        report_path=Path(args.report_path).expanduser(),
        api_key=api_key,
    )

    t = report["totals"]
    print(f"Mode: {report['mode']}")
    print(
        f"Files seen={t['files_seen']} updated={t['files_updated']} | "
        f"Rows seen={t['rows_seen']} changed={t['rows_changed']} unresolved={t['rows_unresolved']}"
    )
    print(f"Report: {args.report_path}")


if __name__ == "__main__":
    main()
