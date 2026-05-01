#!/usr/bin/env python3
"""
Fly Gene Classification Pipeline (Standalone)

Usage:
    python flai_gene_classification.py <input_directory> [--keywords "kw1,kw2,..."] [--reference-limit N]

This script processes each CSV file in the specified directory, reads gene symbols from the
configured input column, converts them to FlyBase IDs, gathers literature evidence, and classifies
genes against the supplied keywords.

Output:
    For each input file <name>.csv, creates <name>_classification.xlsx in the same directory.
"""

import os
import sys
import argparse
import requests
import time
import random
import re
import io
import subprocess
import json
import hashlib
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import pandas as pd
from collections import defaultdict
from dataclasses import dataclass, field
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from functools import lru_cache
from pathlib import Path
from datetime import datetime, timezone

PROJECT_ROOT = Path(__file__).resolve().parent
# Prefer repo-local .env for batch jobs, then fall back to default dotenv search.
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")
load_dotenv()

# API Keys
NCBI_API_KEY = os.getenv("NCBI_API_KEY")
UNPAYWALL_TOKEN = os.getenv("UNPAYWALL_TOKEN") or "aadish98@gmail.com"
PUBMED_CACHE_COLUMNS = [
    "pmid", "pmcid", "title", "abstract", "year", "journal",
    "authors", "doi", "source", "updated_at"
]

from metapub import PubMedFetcher
from openai import OpenAI
from glob import glob
from urllib.parse import quote, urlparse
from HelperScripts.flybase_data import (
    SHARED_DATA_ROOT,
    ensure_flybase_data_files,
    resolve_flybase_data_dir,
    resolve_pubmed_cache_dir,
    using_local_data_root,
)
from HelperScripts.gene_models import GeneCatalog, GeneRecord, ReferenceCandidate
from HelperScripts import human_gene_data, mouse_gene_data
from HelperScripts.get_orthologs import convert_directory as convert_ortholog_directory
from HelperScripts.metadata_resolver import normalize_authors, resolve_reference_metadata

PUBMED_CACHE_DIR = resolve_pubmed_cache_dir()
PUBMED_CACHE_PATH = PUBMED_CACHE_DIR / "pmid_to_title_abstract.csv"
FULLTEXT_METHOD_CACHE_PATH = PUBMED_CACHE_DIR / "pmid_to_fulltext_method.csv"

DEFAULT_SUMMARY_MODEL = "gpt-5.4-nano"
DEFAULT_CLASSIFICATION_MODEL = "gpt-5.4"
DEFAULT_SUMMARY_REASONING_EFFORT = "medium"
DEFAULT_CLASSIFICATION_REASONING_EFFORT = "high"
OPENAI_PRICING_CSV_PATH = PROJECT_ROOT / "Data" / "openai_pricing.csv"
SOFT_RUN_REFERENCE_PROFILES_PATH = PROJECT_ROOT / ".flai_system" / "soft_run_reference_profiles.json"
LITELLM_PRICING_URL = (
    "https://raw.githubusercontent.com/BerriAI/litellm/main/"
    "model_prices_and_context_window.json"
)
INPUT_TOKENS_PER_REF = 6500
OUTPUT_CAP_PER_REF = 1200
INPUT_TOKENS_PER_CLASSIFICATION = 2000
OUTPUT_CAP_PER_CLASSIFICATION = 4000
_openai_client: Optional[OpenAI] = None


@dataclass
class ModelPricing:
    model: str
    input_per_1m: float
    cached_input_per_1m: float
    output_per_1m: float
    source: str = ""
    fetched_at_utc: str = ""


@dataclass
class SoftRunFileEstimate:
    csv_path: str
    total_genes: int = 0
    reused_genes: int = 0
    planned_genes: int = 0
    candidate_references: float = 0
    limited_references: float = 0
    metadata_keyword_matches: float = 0
    n_refs: float = 0
    n_genes_classified: int = 0
    keyword_pass_rate: float = 1.0
    keyword_pass_rate_source: str = "fallback"
    keyword_pass_rate_limited_references: int = 0
    keyword_pass_rate_metadata_keyword_matches: int = 0


@dataclass
class SoftRunEstimate:
    files: list[SoftRunFileEstimate] = field(default_factory=list)
    summary_model: str = ""
    classification_model: str = ""
    summary_pricing: Optional[ModelPricing] = None
    classification_pricing: Optional[ModelPricing] = None

    @property
    def total_genes(self) -> int:
        return sum(item.total_genes for item in self.files)

    @property
    def reused_genes(self) -> int:
        return sum(item.reused_genes for item in self.files)

    @property
    def planned_genes(self) -> int:
        return sum(item.planned_genes for item in self.files)

    @property
    def candidate_references(self) -> float:
        return sum(item.candidate_references for item in self.files)

    @property
    def limited_references(self) -> float:
        return sum(item.limited_references for item in self.files)

    @property
    def metadata_keyword_matches(self) -> float:
        return sum(item.metadata_keyword_matches for item in self.files)

    @property
    def n_refs(self) -> float:
        return sum(item.n_refs for item in self.files)

    @property
    def n_genes_classified(self) -> int:
        return sum(item.n_genes_classified for item in self.files)


def _resolve_openai_model(
    model_name: Optional[str],
    *,
    env_var: str,
    default_model: str,
) -> str:
    """Resolve model from explicit arg, task-specific env var, global env var, or default."""
    model = str(
        model_name
        or os.getenv(env_var)
        or default_model
    ).strip()
    return model or default_model


def _get_summary_model(model_name: Optional[str] = None) -> str:
    """Resolve the model used for per-reference summarization."""
    return _resolve_openai_model(
        model_name,
        env_var="OPENAI_SUMMARY_MODEL",
        default_model=DEFAULT_SUMMARY_MODEL,
    )


def _get_classification_model(model_name: Optional[str] = None) -> str:
    """Resolve the model used for final per-gene classification."""
    return _resolve_openai_model(
        model_name,
        env_var="OPENAI_CLASSIFICATION_MODEL",
        default_model=DEFAULT_CLASSIFICATION_MODEL,
    )


def _get_reasoning_effort(env_var: str, default_effort: str) -> Optional[str]:
    effort = str(os.getenv(env_var) or default_effort or "").strip().lower()
    return effort or None


def _get_openai_client() -> OpenAI:
    """Create OpenAI client lazily with a clear missing-key error."""
    global _openai_client
    if _openai_client is not None:
        return _openai_client
    api_key = str(os.getenv("OPENAI_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY is not set. Export it in the environment "
            "or add it to a .env file in the project root."
        )
    _openai_client = OpenAI(api_key=api_key)
    return _openai_client

# Shared FlyBase data location for this standalone script, falling back to repo-local Data/FlyBase.
FLYBASE_DATA = resolve_flybase_data_dir()
GET_FBGN_IDS_SCRIPT = Path(__file__).parent / "HelperScripts" / "GetFBgnIDs.py"
BATCH_SIZE = 50
BATCH_STATE_DIRNAME = ".batch_state"
RUN_STORE_FILENAME = "run_store.json"
CACHE_SCHEMA_VERSION = 2
MAX_PROMPT_WORDS = 5000
SHARED_PROMPT_OVERHEAD_WORDS = 800
MIN_CHUNK_WORDS = 750


def log_step(message: str, detail: str = "", indent: int = 0):
    """Print concise progress markers for long-running batch jobs."""
    prefix = "  " * max(indent, 0)
    suffix = f" - {detail}" if detail else ""
    print(f"{prefix}[Pipeline] {message}{suffix}")


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _normalize_pricing_model_name(model: str) -> str:
    """Normalize dated model names for pricing lookup without guessing families."""
    name = str(model or "").strip()
    if name.startswith("openai/"):
        name = name.split("/", 1)[1]
    return re.sub(r"-\d{4}-\d{2}-\d{2}$", "", name)


def _load_soft_run_system_profiles() -> dict[str, Any]:
    try:
        with open(SOFT_RUN_REFERENCE_PROFILES_PATH, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return {"profiles": [], "species_keyword_pass_rates": {}}
    if not isinstance(payload, dict):
        return {"profiles": [], "species_keyword_pass_rates": {}}
    payload.setdefault("profiles", [])
    payload.setdefault("species_keyword_pass_rates", {})
    return payload


def _save_soft_run_system_profiles(payload: dict[str, Any]):
    SOFT_RUN_REFERENCE_PROFILES_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = SOFT_RUN_REFERENCE_PROFILES_PATH.with_suffix(".json.tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2, sort_keys=True)
    os.replace(tmp_path, SOFT_RUN_REFERENCE_PROFILES_PATH)


def _resolve_soft_run_reference_profile(summary_model: str, classification_model: str) -> Optional[dict[str, object]]:
    """Return empirical token averages for a known soft-run model pair, if available."""
    key = (
        _normalize_pricing_model_name(summary_model),
        _normalize_pricing_model_name(classification_model),
    )
    payload = _load_soft_run_system_profiles()
    profiles = payload.get("profiles", []) if isinstance(payload, dict) else []
    for profile in profiles:
        if not isinstance(profile, dict):
            continue
        profile_key = (
            _normalize_pricing_model_name(str(profile.get("summary_model", "") or "")),
            _normalize_pricing_model_name(str(profile.get("classification_model", "") or "")),
        )
        if profile_key == key:
            return profile
    return None


def _resolve_species_keyword_pass_rate(species: str) -> dict[str, object]:
    """Resolve the empirical keyword pass rate used by fast soft-run."""
    species_key = str(species or "fly").strip().lower() or "fly"
    payload = _load_soft_run_system_profiles()
    rates = payload.get("species_keyword_pass_rates", {}) if isinstance(payload, dict) else {}
    rate_info = rates.get(species_key, {}) if isinstance(rates, dict) else {}
    if isinstance(rate_info, dict):
        rate = _to_float(rate_info.get("keyword_pass_rate"), default=1.0)
        return {
            "species": species_key,
            "keyword_pass_rate": max(0.0, min(1.0, rate)),
            "source": str(rate_info.get("source", "system profile") or "system profile"),
            "calibration_limited_references": int(_to_float(rate_info.get("calibration_limited_references"))),
            "calibration_metadata_keyword_matches": int(_to_float(rate_info.get("calibration_metadata_keyword_matches"))),
            "updated_at_utc": str(rate_info.get("updated_at_utc", "") or ""),
        }
    return {
        "species": species_key,
        "keyword_pass_rate": 1.0,
        "source": "fallback upper bound",
        "calibration_limited_references": 0,
        "calibration_metadata_keyword_matches": 0,
        "updated_at_utc": "",
    }


def _maybe_update_species_keyword_pass_rate(
    species: str,
    observed_limited_references: int,
    observed_metadata_keyword_matches: int,
    *,
    source: str,
    reference_limit: int,
) -> bool:
    """Refresh species pass-rate calibration only when the completed run is larger."""
    species_key = str(species or "fly").strip().lower() or "fly"
    observed_limited = int(observed_limited_references or 0)
    observed_matches = int(observed_metadata_keyword_matches or 0)
    if observed_limited <= 0:
        return False
    payload = _load_soft_run_system_profiles()
    rates = payload.setdefault("species_keyword_pass_rates", {})
    if not isinstance(rates, dict):
        rates = {}
        payload["species_keyword_pass_rates"] = rates
    existing = rates.get(species_key, {})
    existing_limited = 0
    if isinstance(existing, dict):
        existing_limited = int(_to_float(existing.get("calibration_limited_references")))
    if observed_limited < existing_limited:
        return False

    rates[species_key] = {
        "keyword_pass_rate": max(0.0, min(1.0, observed_matches / observed_limited)),
        "calibration_limited_references": observed_limited,
        "calibration_metadata_keyword_matches": observed_matches,
        "reference_limit": int(reference_limit or 0),
        "source": str(source or "real run"),
        "updated_at_utc": _utc_now_iso(),
    }
    _save_soft_run_system_profiles(payload)
    return True


def _refresh_openai_pricing_csv() -> bool:
    """Refresh model pricing from LiteLLM's public JSON without calling OpenAI."""
    try:
        response = requests.get(LITELLM_PRICING_URL, timeout=20)
        if response.status_code != 200:
            raise RuntimeError(f"HTTP {response.status_code}")
        payload = response.json()
        if not isinstance(payload, dict):
            raise RuntimeError("pricing payload was not a JSON object")

        fetched_at = _utc_now_iso()
        excluded_modes = {
            "embedding",
            "image_generation",
            "audio_transcription",
            "audio_speech",
            "moderation",
        }
        rows: list[dict[str, object]] = []
        for model_name, info in payload.items():
            if not isinstance(info, dict):
                continue
            clean_model = str(model_name or "").strip()
            provider = str(info.get("litellm_provider", "") or "").strip().lower()
            if provider != "openai" and not clean_model.startswith("openai/"):
                continue
            mode = str(info.get("mode", "") or "").strip().lower()
            if mode in excluded_modes:
                continue
            lower_model = clean_model.lower()
            if any(token in lower_model for token in ("embedding", "image", "audio", "tts", "whisper", "moderation")):
                continue

            input_per_1m = _to_float(info.get("input_cost_per_token")) * 1_000_000
            cached_input_per_1m = _to_float(
                info.get("cache_read_input_token_cost")
                or info.get("input_cost_per_token_batches")
            ) * 1_000_000
            output_per_1m = _to_float(info.get("output_cost_per_token")) * 1_000_000
            if input_per_1m <= 0 and output_per_1m <= 0:
                continue

            model_key = _normalize_pricing_model_name(clean_model)
            rows.append({
                "model": model_key,
                "input_per_1m": input_per_1m,
                "cached_input_per_1m": cached_input_per_1m,
                "output_per_1m": output_per_1m,
                "source": LITELLM_PRICING_URL,
                "fetched_at_utc": fetched_at,
            })

        if not rows:
            raise RuntimeError("no OpenAI text model pricing rows found")

        df = pd.DataFrame(rows)
        df = df.drop_duplicates(subset=["model"], keep="first").sort_values("model")
        OPENAI_PRICING_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = OPENAI_PRICING_CSV_PATH.with_suffix(".csv.tmp")
        df.to_csv(tmp_path, index=False)
        tmp_path.replace(OPENAI_PRICING_CSV_PATH)
        print(f"  Refreshed OpenAI pricing table: {OPENAI_PRICING_CSV_PATH}")
        return True
    except Exception as e:
        print(f"  [Warning] Could not refresh OpenAI pricing table: {e}")
        return False


def _load_openai_pricing_table(refresh: bool = False) -> dict[str, ModelPricing]:
    """Load pricing from the external CSV, refreshing first for soft-run when requested."""
    if refresh:
        refreshed = _refresh_openai_pricing_csv()
        if not refreshed and OPENAI_PRICING_CSV_PATH.exists():
            try:
                existing = pd.read_csv(OPENAI_PRICING_CSV_PATH, dtype=str, keep_default_na=False)
                fetched_values = sorted({
                    str(x).strip() for x in existing.get("fetched_at_utc", []) if str(x).strip()
                })
                fetched_at = fetched_values[-1] if fetched_values else "unknown"
                print(f"  Using cached OpenAI pricing table from {fetched_at}")
            except Exception:
                print("  Using cached OpenAI pricing table")

    if not OPENAI_PRICING_CSV_PATH.exists():
        raise RuntimeError(
            f"OpenAI pricing table is missing: {OPENAI_PRICING_CSV_PATH}. "
            "Run --soft-run with network access or populate the CSV."
        )

    df = pd.read_csv(OPENAI_PRICING_CSV_PATH, dtype=str, keep_default_na=False)
    required = {"model", "input_per_1m", "cached_input_per_1m", "output_per_1m", "source", "fetched_at_utc"}
    missing = required.difference(df.columns)
    if missing:
        raise RuntimeError(f"OpenAI pricing CSV is missing columns: {', '.join(sorted(missing))}")

    table: dict[str, ModelPricing] = {}
    for _, row in df.iterrows():
        model = _normalize_pricing_model_name(row.get("model", ""))
        if not model:
            continue
        table[model] = ModelPricing(
            model=model,
            input_per_1m=_to_float(row.get("input_per_1m")),
            cached_input_per_1m=_to_float(row.get("cached_input_per_1m")),
            output_per_1m=_to_float(row.get("output_per_1m")),
            source=str(row.get("source", "") or ""),
            fetched_at_utc=str(row.get("fetched_at_utc", "") or ""),
        )
    if not table:
        raise RuntimeError(f"OpenAI pricing CSV has no usable pricing rows: {OPENAI_PRICING_CSV_PATH}")
    return table


def _resolve_model_pricing(model: str, table: dict[str, ModelPricing]) -> ModelPricing:
    model_name = str(model or "").strip()
    candidates = [model_name, _normalize_pricing_model_name(model_name)]
    for candidate in candidates:
        if candidate in table:
            return table[candidate]
    available = ", ".join(sorted(table)[:40])
    if len(table) > 40:
        available += ", ..."
    raise RuntimeError(
        f"No OpenAI pricing entry found for model '{model_name}'. "
        f"Supported pricing models include: {available}"
    )


###############################################################################
# Full-Text Method Cache (PMID -> successful retrieval method)
###############################################################################

_fulltext_method_cache: Dict[str, str] = {}
_fulltext_method_pending: Dict[str, str] = {}
_fulltext_cache_loaded = False

_pmid_title_abstract_cache: Dict[str, dict] = {}
_pmid_title_abstract_pending: Dict[str, dict] = {}
_pmid_title_abstract_loaded = False


def _load_fulltext_method_cache() -> Dict[str, str]:
    """Load persistent PMID->method mapping from turbo-server cache."""
    global _fulltext_cache_loaded, _fulltext_method_cache
    if _fulltext_cache_loaded:
        return _fulltext_method_cache
    try:
        if FULLTEXT_METHOD_CACHE_PATH.exists():
            cache_df = pd.read_csv(
                FULLTEXT_METHOD_CACHE_PATH,
                dtype={"pmid": str, "method": str}
            )
            cache_df["pmid"] = cache_df["pmid"].apply(
                lambda x: str(x).strip() if pd.notna(x) else ""
            )
            cache_df = cache_df[cache_df["pmid"] != ""].copy()
            if len(cache_df) > 0:
                _fulltext_method_cache = {
                    row["pmid"]: str(row["method"]) if pd.notna(row["method"]) else ""
                    for _, row in cache_df.iterrows()
                }
            print(f"  Loaded {len(_fulltext_method_cache)} full-text cache entries")
    except Exception as e:
        print(f"  [Warning] Could not load full-text method cache: {e}")
    _fulltext_cache_loaded = True
    return _fulltext_method_cache


def _get_cached_fulltext_method(pmid: str) -> Optional[str]:
    """Get cached successful retrieval method for a PMID."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return None
    cache = _load_fulltext_method_cache()
    method = cache.get(pmid_clean)
    return method if method else None


def _set_cached_fulltext_method(pmid: str, method: str):
    """Store PMID->method in memory and pending-write buffer."""
    pmid_clean = str(pmid or "").strip()
    method_clean = str(method or "").strip()
    if not pmid_clean or not method_clean:
        return
    _load_fulltext_method_cache()
    _fulltext_method_cache[pmid_clean] = method_clean
    _fulltext_method_pending[pmid_clean] = method_clean


def _save_fulltext_method_cache_pending():
    """Append pending PMID->method mappings to shared cache CSV."""
    if not _fulltext_method_pending:
        return
    try:
        rows = [{"pmid": p, "method": m} for p, m in _fulltext_method_pending.items()]
        new_df = pd.DataFrame(rows)
        FULLTEXT_METHOD_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        if FULLTEXT_METHOD_CACHE_PATH.exists():
            new_df.to_csv(FULLTEXT_METHOD_CACHE_PATH, mode="a", header=False, index=False)
        else:
            new_df.to_csv(FULLTEXT_METHOD_CACHE_PATH, index=False)
        print(f"  Saved {len(_fulltext_method_pending)} new full-text cache entries")
        _fulltext_method_pending.clear()
    except Exception as e:
        print(f"  [Warning] Could not save full-text method cache: {e}")


def _load_pmid_title_abstract_cache() -> Dict[str, dict]:
    """Load persistent PMID metadata cache from turbo-server CSV."""
    global _pmid_title_abstract_loaded, _pmid_title_abstract_cache
    if _pmid_title_abstract_loaded:
        return _pmid_title_abstract_cache
    try:
        if PUBMED_CACHE_PATH.exists():
            cache_df = pd.read_csv(PUBMED_CACHE_PATH, dtype=str, keep_default_na=False)
            drop_cols = [c for c in cache_df.columns if str(c).startswith("Unnamed")]
            if drop_cols:
                cache_df = cache_df.drop(columns=drop_cols)
            if "pmid" not in cache_df.columns:
                cache_df = pd.DataFrame(columns=PUBMED_CACHE_COLUMNS)
            for col in PUBMED_CACHE_COLUMNS:
                if col not in cache_df.columns:
                    cache_df[col] = ""
            cache_df["pmid"] = cache_df["pmid"].apply(
                lambda x: str(x).strip() if pd.notna(x) else ""
            )
            cache_df = cache_df[cache_df["pmid"].str.isdigit()].copy()
            cache_df = cache_df[PUBMED_CACHE_COLUMNS]
            if len(cache_df) > 0:
                for _, row in cache_df.iterrows():
                    pmid = str(row.get("pmid", "") or "").strip()
                    if not pmid:
                        continue
                    authors_list = normalize_authors(row.get("authors", ""))
                    _pmid_title_abstract_cache[pmid] = {
                        "title": str(row.get("title", "") or ""),
                        "abstract": str(row.get("abstract", "") or ""),
                        "year": str(row.get("year", "") or ""),
                        "journal": str(row.get("journal", "") or ""),
                        "authors": authors_list,
                        "doi": str(row.get("doi", "") or ""),
                        "pmcid": str(row.get("pmcid", "") or ""),
                        "source": str(row.get("source", "") or ""),
                        "updated_at": str(row.get("updated_at", "") or ""),
                    }
            print(f"  Loaded {len(_pmid_title_abstract_cache)} PubMed metadata cache entries")
    except Exception as e:
        print(f"  [Warning] Could not load PubMed metadata cache: {e}")
    _pmid_title_abstract_loaded = True
    return _pmid_title_abstract_cache


def _get_cached_pmid_title_abstract(pmid: str) -> Optional[dict]:
    """Get cached metadata for PMID."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return None
    cache = _load_pmid_title_abstract_cache()
    out = cache.get(pmid_clean)
    if not out:
        return None
    return {
        "title": str(out.get("title", "") or ""),
        "abstract": str(out.get("abstract", "") or ""),
        "year": str(out.get("year", "") or ""),
        "journal": str(out.get("journal", "") or ""),
        "authors": normalize_authors(out.get("authors", [])),
        "doi": str(out.get("doi", "") or ""),
        "pmcid": str(out.get("pmcid", "") or ""),
        "source": str(out.get("source", "") or ""),
        "updated_at": str(out.get("updated_at", "") or ""),
    }


def _set_cached_pmid_title_abstract(
    pmid: str,
    title: str,
    abstract: str,
    *,
    year: str = "",
    journal: str = "",
    authors: Optional[list[str]] = None,
    doi: str = "",
    pmcid: str = "",
    source: str = "",
    updated_at: str = "",
):
    """Store PMID metadata in memory and pending-write buffer."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean:
        return
    entry = {
        "title": str(title or ""),
        "abstract": str(abstract or ""),
        "year": str(year or ""),
        "journal": str(journal or ""),
        "authors": normalize_authors(authors or []),
        "doi": str(doi or ""),
        "pmcid": str(pmcid or ""),
        "source": str(source or ""),
        "updated_at": str(updated_at or ""),
    }
    _load_pmid_title_abstract_cache()
    prev = _pmid_title_abstract_cache.get(pmid_clean, {})
    merged = {
        "title": entry["title"] or str(prev.get("title", "") or ""),
        "abstract": entry["abstract"] or str(prev.get("abstract", "") or ""),
        "year": entry["year"] or str(prev.get("year", "") or ""),
        "journal": entry["journal"] or str(prev.get("journal", "") or ""),
        "authors": entry["authors"] or normalize_authors(prev.get("authors", [])),
        "doi": entry["doi"] or str(prev.get("doi", "") or ""),
        "pmcid": entry["pmcid"] or str(prev.get("pmcid", "") or ""),
        "source": entry["source"] or str(prev.get("source", "") or ""),
        "updated_at": entry["updated_at"] or str(prev.get("updated_at", "") or ""),
    }
    _pmid_title_abstract_cache[pmid_clean] = merged
    _pmid_title_abstract_pending[pmid_clean] = entry


def _save_pmid_title_abstract_cache_pending():
    """Persist pending PMID metadata updates to shared cache CSV."""
    if not _pmid_title_abstract_pending:
        return
    try:
        _load_pmid_title_abstract_cache()
        merged_cache: Dict[str, dict] = {}

        if PUBMED_CACHE_PATH.exists():
            try:
                disk_df = pd.read_csv(PUBMED_CACHE_PATH, dtype=str, keep_default_na=False)
                drop_cols = [c for c in disk_df.columns if str(c).startswith("Unnamed")]
                if drop_cols:
                    disk_df = disk_df.drop(columns=drop_cols)
                if "pmid" in disk_df.columns:
                    for col in PUBMED_CACHE_COLUMNS:
                        if col not in disk_df.columns:
                            disk_df[col] = ""
                    disk_df = disk_df[PUBMED_CACHE_COLUMNS]
                    for _, row in disk_df.iterrows():
                        pmid = str(row.get("pmid", "") or "").strip()
                        if not pmid.isdigit():
                            continue
                        merged_cache[pmid] = {
                            "pmcid": str(row.get("pmcid", "") or ""),
                            "title": str(row.get("title", "") or ""),
                            "abstract": str(row.get("abstract", "") or ""),
                            "year": str(row.get("year", "") or ""),
                            "journal": str(row.get("journal", "") or ""),
                            "authors": normalize_authors(row.get("authors", "")),
                            "doi": str(row.get("doi", "") or ""),
                            "source": str(row.get("source", "") or ""),
                            "updated_at": str(row.get("updated_at", "") or ""),
                        }
            except Exception:
                # Keep best-effort behavior even if disk cache is malformed.
                pass

        for pmid, meta in _pmid_title_abstract_cache.items():
            prior = merged_cache.get(pmid, {})
            merged_cache[pmid] = {
                "pmcid": str(meta.get("pmcid", "") or prior.get("pmcid", "")),
                "title": str(meta.get("title", "") or prior.get("title", "")),
                "abstract": str(meta.get("abstract", "") or prior.get("abstract", "")),
                "year": str(meta.get("year", "") or prior.get("year", "")),
                "journal": str(meta.get("journal", "") or prior.get("journal", "")),
                "authors": normalize_authors(meta.get("authors", []) or prior.get("authors", [])),
                "doi": str(meta.get("doi", "") or prior.get("doi", "")),
                "source": str(meta.get("source", "") or prior.get("source", "")),
                "updated_at": str(meta.get("updated_at", "") or prior.get("updated_at", "")),
            }

        rows = []
        for pmid, meta in merged_cache.items():
            rows.append({
                "pmid": str(pmid),
                "pmcid": str(meta.get("pmcid", "") or ""),
                "title": str(meta.get("title", "") or ""),
                "abstract": str(meta.get("abstract", "") or ""),
                "year": str(meta.get("year", "") or ""),
                "journal": str(meta.get("journal", "") or ""),
                "authors": "; ".join(normalize_authors(meta.get("authors", []))),
                "doi": str(meta.get("doi", "") or ""),
                "source": str(meta.get("source", "") or ""),
                "updated_at": str(meta.get("updated_at", "") or ""),
            })
        out_df = pd.DataFrame(rows, columns=PUBMED_CACHE_COLUMNS)
        out_df = out_df.sort_values(by=["pmid"], ascending=True).reset_index(drop=True)
        PUBMED_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        out_df.to_csv(PUBMED_CACHE_PATH, index=False)
        print(
            f"  Saved {len(_pmid_title_abstract_pending)} pending updates "
            f"({len(out_df)} total PubMed metadata rows)"
        )
        _pmid_title_abstract_pending.clear()
    except Exception as e:
        print(f"  [Warning] Could not save PubMed metadata cache: {e}")


def find_latest_tsv(directory: Path, pattern: str) -> Path:
    """
    Find the latest TSV file matching a pattern in a directory.
    Prefers .gz files, falls back to .tsv if no .gz exists.
    """
    # Try gzipped first
    gz_files = sorted(glob(str(directory / f"{pattern}*.tsv.gz")), reverse=True)
    if gz_files:
        return Path(gz_files[0])
    
    # Fall back to uncompressed
    tsv_files = sorted(glob(str(directory / f"{pattern}*.tsv")), reverse=True)
    if tsv_files:
        return Path(tsv_files[0])
    
    raise FileNotFoundError(f"No TSV file matching '{pattern}' found in {directory}")


def load_flybase_tsv(filepath, **kwargs) -> pd.DataFrame:
    """
    Load a FlyBase TSV file, handling various comment/header formats.
    
    FlyBase TSV files have varying formats:
      - Some have ## metadata comments, then #header, then data
      - Some have multiple # comment lines before #header with tabs
      - Some have NO # prefix on headers (plain column names on line 1)
      - Some have #---- separator lines after the header
    
    This function detects and handles all these cases.
    Supports both .tsv and .tsv.gz files.
    """
    import gzip
    
    filepath = Path(filepath)
    
    # Determine opener based on file extension
    if filepath.suffix == '.gz' or str(filepath).endswith('.tsv.gz'):
        opener = lambda f: gzip.open(f, 'rt', encoding='utf-8')
    else:
        opener = lambda f: open(f, 'r', encoding='utf-8')
    
    # First pass: find the header line and count rows to skip
    skip_rows = 0
    header_line = None
    use_custom_header = False
    header_found = False
    
    with opener(filepath) as f:
        for i, line in enumerate(f):
            stripped = line.strip()
            
            # Skip empty lines
            if not stripped:
                skip_rows = i + 1
                continue
            
            # If we already found the header, we're now at data - break
            if header_found:
                if stripped.startswith('#'):
                    skip_rows = i + 1
                    continue
                skip_rows = i
                break
            
            # This line might be the header
            # Check if it starts with # (then strip it) or is a plain header
            if stripped.startswith('#'):
                # Lines with no tabs are comments/metadata.
                if '\t' not in stripped:
                    skip_rows = i + 1
                    continue

                # Header prefixed with # or ## (e.g. "#FBrf\tPMID..." or "##primary_FBid\t...")
                header_content = stripped.lstrip('#')
                cols = header_content.split('\t')
                # Clean column names
                col_names = [c.strip() for c in cols if c.strip()]
                
                if len(col_names) >= 2:
                    # This looks like a valid header
                    header_line = col_names
                    use_custom_header = True
                    header_found = True
                    skip_rows = i + 1
                    # Don't break - continue to skip any separator lines after header
                    continue
                else:
                    # Single column or empty - treat as comment
                    skip_rows = i + 1
                    continue
            else:
                # No # prefix - could be header or data
                cols = stripped.split('\t')
                
                # Heuristic: if first field looks like a FlyBase ID (starts with FB), 
                # this is likely data, not header - let pandas auto-detect
                first_field = cols[0].strip() if cols else ""
                if first_field.startswith('FB') and len(first_field) > 4:
                    # Looks like data (FBst, FBrf, FBal, FBgn, etc.)
                    # This file has no # on header, just let pandas handle it
                    break
                else:
                    # This is likely a plain header (no # prefix)
                    # Let pandas auto-detect it
                    break
    
    # Read the file
    if use_custom_header and header_line:
        df = pd.read_csv(
            filepath,
            sep='\t',
            skiprows=skip_rows,
            names=header_line,
            low_memory=False,
            on_bad_lines='warn',
            **kwargs
        )
    else:
        df = pd.read_csv(
            filepath,
            sep='\t',
            skiprows=skip_rows,
            low_memory=False,
            on_bad_lines='warn',
            **kwargs
        )
    
    return df


###############################################################################
# FlyBase Data Loading (Cached)
###############################################################################

@lru_cache(maxsize=1)
def load_flybase_synonym_table():
    """Load synonyms table from local TSV (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "Genes", "fb_synonym")
    print(f"  Loading FlyBase synonym table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    df = df[df.organism_abbreviation == "Dmel"]
    return df


@lru_cache(maxsize=1)
def build_fbgn_to_symbol_map():
    """Build mapping from FBgn -> current_symbol."""
    df = load_flybase_synonym_table()
    df['current_symbol'] = df['current_symbol'].astype(str).str.strip()
    df['primary_FBid'] = df['primary_FBid'].astype(str).str.strip()
    return dict(zip(df['primary_FBid'], df['current_symbol']))


@lru_cache(maxsize=1)
def build_fbgn_to_all_names_map():
    """Build mapping from FBgn -> set of names including current_symbol and all synonyms."""
    df = load_flybase_synonym_table()
    df = df.rename(columns={"symbol_synonym(s)": "symbol_synonyms"})
    
    fbgn_to_names = defaultdict(set)
    for row in df.itertuples():
        fbgn = str(getattr(row, 'primary_FBid', '')).strip()
        curr = str(getattr(row, 'current_symbol', '')).strip()
        if fbgn:
            if curr:
                fbgn_to_names[fbgn].add(curr)
            syn_field = getattr(row, "symbol_synonyms", "")
            if isinstance(syn_field, str) and syn_field.strip():
                for s in syn_field.split("|"):
                    s = s.strip()
                    if s:
                        fbgn_to_names[fbgn].add(s)
    return {k: set(v) for k, v in fbgn_to_names.items()}


@lru_cache(maxsize=1)
def build_fbgn_to_primary_names_map():
    """Build mapping from FBgn -> set of primary names (current_symbol and current_fullname only)."""
    df = load_flybase_synonym_table()
    
    fbgn_to_names = defaultdict(set)
    for row in df.itertuples():
        fbgn = str(getattr(row, 'primary_FBid', '')).strip()
        if fbgn:
            curr_symbol = str(getattr(row, 'current_symbol', '')).strip()
            if curr_symbol:
                fbgn_to_names[fbgn].add(curr_symbol)
            curr_fullname = str(getattr(row, 'current_fullname', '')).strip()
            if curr_fullname:
                fbgn_to_names[fbgn].add(curr_fullname)
    return {k: set(v) for k, v in fbgn_to_names.items()}


###############################################################################
# FBgn -> PMCID Retrieval from FlyBase Files
###############################################################################

@lru_cache(maxsize=1)
def load_entity_publication():
    """Load entity_publication TSV with FBgn -> PMID mapping (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "FlyBase_References", "entity_publication")
    print(f"  Loading entity publication table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    df = df[df["entity_id"].str.startswith("FBgn", na=False)]
    df["PubMed_id"] = pd.to_numeric(df["PubMed_id"], errors="coerce")
    df.dropna(subset=["PubMed_id"], inplace=True)
    df["PubMed_id"] = df["PubMed_id"].astype(int)
    return df


@lru_cache(maxsize=1)
def load_fbrf_pmid_pmcid():
    """Load fbrf_pmid_pmcid TSV with PMID -> PMCID mapping (finds latest version automatically)."""
    path = find_latest_tsv(FLYBASE_DATA / "FlyBase_References", "fbrf_pmid_pmcid_doi")
    print(f"  Loading PMID/PMCID mapping table: {path.name}")
    df = load_flybase_tsv(path, keep_default_na=False)
    if "pub_type" in df.columns:
        df = df[df["pub_type"].str.strip().str.lower() == "paper"].copy()
    df["PMID"] = pd.to_numeric(df["PMID"], errors="coerce")
    df.dropna(subset=["PMID"], inplace=True)
    df["PMID"] = df["PMID"].astype(int)
    
    # Extract year from miniref
    if "miniref" in df.columns:
        try:
            def extract_year_after_comma(miniref_str):
                miniref_str = str(miniref_str) if miniref_str else ""
                comma_idx = miniref_str.find(",")
                if comma_idx == -1:
                    return None
                after_comma = miniref_str[comma_idx + 1:]
                match = re.search(r"\b(\d{4})\b", after_comma)
                if match:
                    year = int(match.group(1))
                    if 1900 <= year <= 2100:
                        return year
                return None
            _yr = df["miniref"].apply(extract_year_after_comma)
            df["miniref_year"] = pd.to_numeric(_yr, errors="coerce").fillna(0).astype(int)
        except Exception:
            df["miniref_year"] = 0
    else:
        df["miniref_year"] = 0
    return df


def merge_on_pmid():
    """Merge entity_publication & fbrf_pmid_pmcid on PMID => PMCID."""
    entity_df = load_entity_publication().rename(columns={"PubMed_id": "PMID"})
    fbrf_df = load_fbrf_pmid_pmcid()
    merged = pd.merge(entity_df, fbrf_df, on="PMID", how="inner")
    return merged


@lru_cache(maxsize=1)
def build_pmcid_to_year():
    """Build mapping PMCID -> miniref_year."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_year = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip()
        if not pmc or not pmc.lower().startswith("pmc"):
            continue
        try:
            yr = int(getattr(row, "miniref_year", 0) or 0)
        except Exception:
            yr = 0
        pmc_norm = pmc.upper()
        if pmc_norm not in pmcid_to_year or yr > pmcid_to_year[pmc_norm]:
            pmcid_to_year[pmc_norm] = yr
    return pmcid_to_year


@lru_cache(maxsize=1)
def build_pmcid_to_pmid():
    """Build mapping PMCID -> PMID from FlyBase reference metadata."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_pmid = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip().upper()
        pmid = str(getattr(row, "PMID", "")).strip()
        if pmc.startswith("PMC") and pmid.isdigit():
            pmcid_to_pmid[pmc] = pmid
    return pmcid_to_pmid


@lru_cache(maxsize=1)
def build_pmcid_to_doi():
    """Build mapping PMCID -> DOI from FlyBase reference metadata."""
    df = load_fbrf_pmid_pmcid()
    pmcid_to_doi = {}
    for row in df.itertuples():
        pmc = str(getattr(row, "PMCID", "")).strip().upper()
        doi = str(getattr(row, "DOI", "")).strip()
        if pmc.startswith("PMC") and doi:
            pmcid_to_doi[pmc] = doi
    return pmcid_to_doi


def get_pmcids_for_fbgn_list(fbgn_ids):
    """Get PMCIDs for a list of FBgn IDs from FlyBase precomputed files."""
    merged = merge_on_pmid()
    sub = merged[merged["entity_id"].isin(fbgn_ids)]
    
    fbgn_to_pmcids = defaultdict(set)
    for row in sub.itertuples():
        pmc = str(row.PMCID).strip() if pd.notna(row.PMCID) else None
        if pmc and pmc.lower().startswith("pmc"):
            fbgn_to_pmcids[row.entity_id].add(pmc)
    return fbgn_to_pmcids


###############################################################################
# HTTP Helper
###############################################################################

def _http_get(url, headers=None, params=None, timeout=15, max_retries=3, backoff=0.75):
    """HTTP GET with retries for transient failures."""
    merged_headers = {
        "User-Agent": "fly-gene-classifier/1.0 (+https://github.com/Allada-Lab)",
        "Accept": "*/*",
    }
    if headers:
        merged_headers.update(headers)
    for attempt in range(max_retries):
        try:
            r = requests.get(
                url,
                headers=merged_headers,
                params=params or {},
                timeout=timeout,
                allow_redirects=True
            )
            if r.status_code == 200:
                return r
            if r.status_code in (429, 500, 502, 503, 504):
                wait_s = backoff * (2 ** attempt) + random.uniform(0, backoff)
                time.sleep(wait_s)
                continue
            host = urlparse(url).netloc or url
            print(f"  [Warning] HTTP {r.status_code} from {host}")
            return None
        except Exception:
            wait_s = backoff * (2 ** attempt) + random.uniform(0, backoff)
            time.sleep(wait_s)
    return None


###############################################################################
# PubMed / Europe PMC Search
###############################################################################

def search_pubmed_pmids_for_fly_gene(symbol: str, primary_names: list[str], keywords: list[str]) -> list[int]:
    """Search PubMed for fly gene references."""
    return search_pubmed_pmids_for_gene(
        symbol,
        primary_names,
        keywords,
        organism_filter='("drosophila" OR "fruit fly")',
        label="PubMed",
    )


def search_pubmed_pmids_for_gene(
    symbol: str,
    primary_names: list[str],
    keywords: list[str],
    organism_filter: str,
    label: str = "PubMed",
) -> list[int]:
    """Search PubMed for species-scoped gene references."""
    safe_gene_terms = []
    if symbol and symbol.strip():
        safe_gene_terms.append(symbol.strip())
    
    if primary_names:
        for name in primary_names:
            name = (name or "").strip()
            if not name or name == symbol:
                continue
            if len(name) >= 4 or " " in name or "-" in name:
                safe_gene_terms.append(name)
    
    if not safe_gene_terms:
        return []
    
    gene_or = " OR ".join([f'"{t}"' for t in sorted(safe_gene_terms)])
    kw_terms = [k.strip() for k in (keywords or []) if k and k.strip()]
    kw_or = " OR ".join([f'"{k}"' for k in kw_terms]) if kw_terms else ""
    species_filter = str(organism_filter or "").strip()
    
    if kw_or:
        term = f"({gene_or}) AND ({kw_or}) AND {species_filter}" if species_filter else f"({gene_or}) AND ({kw_or})"
    else:
        term = f"({gene_or}) AND {species_filter}" if species_filter else f"({gene_or})"
    
    params = {
        "db": "pubmed",
        "term": term,
        "retmax": 200,
        "retmode": "json",
        "sort": "pub+date",
    }
    if NCBI_API_KEY:
        params["api_key"] = NCBI_API_KEY
    
    print(f"  [{label}] Query: {term[:100]}...")
    r = _http_get("https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi", params=params, timeout=20)
    if r is None:
        return []
    try:
        data = r.json()
        ids = data.get("esearchresult", {}).get("idlist", [])
        return [int(x) for x in ids if str(x).isdigit()]
    except Exception:
        return []


def search_europe_pmc_pmids_for_fly_gene(symbol: str, primary_names: list[str], keywords: list[str]) -> list[int]:
    """Search Europe PMC for fly gene references."""
    return search_europe_pmc_pmids_for_gene(
        symbol,
        primary_names,
        keywords,
        organism_filter='("drosophila" OR "fruit fly") AND "gene"',
        label="Europe PMC",
    )


def search_europe_pmc_pmids_for_gene(
    symbol: str,
    primary_names: list[str],
    keywords: list[str],
    organism_filter: str,
    label: str = "Europe PMC",
    require_open_access: bool = True,
) -> list[int]:
    """Search Europe PMC for species-scoped gene references."""
    safe_gene_terms = []
    if symbol and symbol.strip():
        safe_gene_terms.append(symbol.strip())
    
    if primary_names:
        for name in primary_names:
            name = (name or "").strip()
            if not name or name == symbol:
                continue
            if len(name) >= 4 or " " in name or "-" in name:
                safe_gene_terms.append(name)
    
    if not safe_gene_terms:
        return []
    
    gene_or = " OR ".join([f'"{t}"' for t in sorted(safe_gene_terms)])
    kw_terms = [k.strip() for k in (keywords or []) if k and k.strip()]
    kw_or = " OR ".join([f'"{k}"' for k in kw_terms]) if kw_terms else ""
    species_filter = str(organism_filter or "").strip()
    
    if kw_or:
        query = f"({gene_or}) AND ({kw_or})"
    else:
        query = f"({gene_or})"
    if species_filter:
        query = f"{query} AND {species_filter}"
    if require_open_access:
        query = f"{query} AND OPEN_ACCESS:Y"
    
    params = {
        "query": query,
        "resultType": "core",
        "pageSize": 200,
        "format": "json",
    }
    
    print(f"  [{label}] Query: {query[:100]}...")
    r = _http_get("https://www.ebi.ac.uk/europepmc/webservices/rest/search", params=params, timeout=20)
    if r is None:
        return []
    try:
        data = r.json() or {}
        results = (data.get("resultList") or {}).get("result", []) or []
        pmids = []
        for rec in results:
            if rec.get("source") == "MED":
                pid = rec.get("pmid") or rec.get("id")
                if pid and str(pid).isdigit():
                    pmids.append(int(pid))
        return pmids
    except Exception:
        return []


def batch_pmids_to_pmcids(pmids: list[int]) -> dict[int, str]:
    """Convert batch of PMIDs to PMCIDs."""
    if not pmids:
        return {}
    
    result = {}
    batch_size = 200
    for i in range(0, len(pmids), batch_size):
        batch = pmids[i:i + batch_size]
        ids_str = ",".join(str(p) for p in batch)
        url = "https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/"
        params = {
            "tool": "fly_gene_classifier",
            "email": "aadish98@gmail.com",
            "ids": ids_str,
            "format": "json"
        }
        if NCBI_API_KEY:
            params["api_key"] = NCBI_API_KEY
        
        try:
            r = _http_get(url, params=params, timeout=30)
            if r and r.status_code == 200:
                data = r.json()
                recs = data.get("records", [])
                for rec in recs:
                    pmid_val = rec.get("pmid")
                    pmcid_val = rec.get("pmcid")
                    if pmid_val and pmcid_val:
                        try:
                            result[int(pmid_val)] = pmcid_val
                        except ValueError:
                            pass
        except Exception as e:
            print(f"  [Warning] Batch PMID->PMCID error: {e}")
        time.sleep(0.1)
    
    return result


def get_pubmed_pmcids_for_fly_gene(fbgn_id: str, symbol: str, primary_names: set[str], keywords: list[str]) -> set[str]:
    """Search PubMed and return PMCIDs."""
    pmids = search_pubmed_pmids_for_fly_gene(symbol, list(primary_names), keywords)
    if not pmids:
        return set()
    pmid_to_pmc = batch_pmids_to_pmcids(pmids)
    return set(pmid_to_pmc.values())


def get_europe_pmc_pmcids_for_fly_gene(fbgn_id: str, symbol: str, primary_names: set[str], keywords: list[str]) -> set[str]:
    """Search Europe PMC and return PMCIDs."""
    pmids = search_europe_pmc_pmids_for_fly_gene(symbol, list(primary_names), keywords)
    if not pmids:
        return set()
    pmid_to_pmc = batch_pmids_to_pmcids(pmids)
    return set(pmid_to_pmc.values())


###############################################################################
# Species adapters and reference providers
###############################################################################


def _pmids_to_reference_candidates(
    gene_id: str,
    pmids: list[int] | set[str],
    source_key: str,
    source_label: str,
    snippets: dict[str, str] | None = None,
) -> list[ReferenceCandidate]:
    cleaned_pmids = []
    for pmid in pmids or []:
        text = str(pmid).strip()
        if text.isdigit():
            cleaned_pmids.append(int(text))
    if not cleaned_pmids:
        if pmids:
            print(
                f"  [Warning] {source_label}: no usable PMID values for "
                f"gene {gene_id} from {len(pmids)} input value(s)"
            )
        return []
    pmid_to_pmcid = batch_pmids_to_pmcids(cleaned_pmids)
    candidates = []
    for pmid_int in cleaned_pmids:
        pmid = str(pmid_int)
        pmcid = pmid_to_pmcid.get(pmid_int, "")
        paper_id = pmcid or pmid
        candidates.append(
            ReferenceCandidate(
                gene_id=gene_id,
                paper_id=paper_id,
                source_key=source_key,
                source_label=source_label,
                pmid=pmid,
                pmcid=pmcid,
                snippet=(snippets or {}).get(pmid, ""),
            )
        )
    return candidates


class ReferenceProvider:
    source_key = ""
    source_label = ""

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        raise NotImplementedError


class FlyBaseReferenceProvider(ReferenceProvider):
    source_key = "flybase"
    source_label = "FlyBase"

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        fbgn_ids = [gene.gene_id for gene in genes]
        pmcids_by_gene = get_pmcids_for_fbgn_list(fbgn_ids)
        return {
            gene_id: [
                ReferenceCandidate(
                    gene_id=gene_id,
                    paper_id=str(pmcid).upper(),
                    pmcid=str(pmcid).upper(),
                    source_key=self.source_key,
                    source_label=self.source_label,
                )
                for pmcid in sorted(pmcids or [])
            ]
            for gene_id, pmcids in pmcids_by_gene.items()
        }


class Gene2PubMedProvider(ReferenceProvider):
    source_key = "gene2pubmed"
    source_label = "gene2pubmed"

    def __init__(self, data_module):
        self.data_module = data_module

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        gene_ids = [gene.gene_id for gene in genes]
        pmids_by_gene = self.data_module.get_gene2pubmed_pmids(gene_ids)
        return {
            gene_id: _pmids_to_reference_candidates(gene_id, pmids, self.source_key, self.source_label)
            for gene_id, pmids in pmids_by_gene.items()
        }


class GeneRIFProvider(ReferenceProvider):
    source_key = "generif"
    source_label = "GeneRIF"

    def __init__(self, data_module):
        self.data_module = data_module

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        gene_ids = [gene.gene_id for gene in genes]
        pmids_by_gene = self.data_module.get_generif_pmids(gene_ids)
        out = {}
        for gene_id, pmids in pmids_by_gene.items():
            snippet_map = {
                str(pmid): snippet
                for pmid, snippet in self.data_module.get_generif_snippets(gene_id)
            }
            out[gene_id] = _pmids_to_reference_candidates(
                gene_id, pmids, self.source_key, self.source_label, snippets=snippet_map
            )
        return out


class UniProtProvider(ReferenceProvider):
    source_key = "uniprot"
    source_label = "UniProt"

    def __init__(self, data_module):
        self.data_module = data_module

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        gene_ids = [gene.gene_id for gene in genes]
        pmids_by_gene = self.data_module.get_uniprot_pmids(gene_ids)
        return {
            gene_id: _pmids_to_reference_candidates(gene_id, pmids, self.source_key, self.source_label)
            for gene_id, pmids in pmids_by_gene.items()
        }


class PubMedSearchProvider(ReferenceProvider):
    source_key = "pubmed"
    source_label = "PubMed"

    def __init__(self, organism_filter: str):
        self.organism_filter = organism_filter

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        out = {}
        for gene in genes:
            pmids = search_pubmed_pmids_for_gene(
                gene.symbol,
                list(gene.synonyms),
                keywords,
                organism_filter=self.organism_filter,
                label=self.source_label,
            )
            out[gene.gene_id] = _pmids_to_reference_candidates(gene.gene_id, pmids, self.source_key, self.source_label)
        return out


class EuropePMCSearchProvider(ReferenceProvider):
    source_key = "europe_pmc"
    source_label = "Europe PMC"

    def __init__(self, organism_filter: str, require_open_access: bool = True):
        self.organism_filter = organism_filter
        self.require_open_access = require_open_access

    def collect_many(self, genes: list[GeneRecord], keywords: list[str]) -> dict[str, list[ReferenceCandidate]]:
        out = {}
        for gene in genes:
            pmids = search_europe_pmc_pmids_for_gene(
                gene.symbol,
                list(gene.synonyms),
                keywords,
                organism_filter=self.organism_filter,
                label=self.source_label,
                require_open_access=self.require_open_access,
            )
            out[gene.gene_id] = _pmids_to_reference_candidates(gene.gene_id, pmids, self.source_key, self.source_label)
        return out


@dataclass
class SpeciesAdapter:
    key: str
    display_name: str
    taxon_id: int
    gene_id_column: str
    symbol_column: str
    authority_id_column: str
    output_suffix: str
    pubmed_filter: str
    europe_pmc_filter: str
    source_priority: dict[str, int]
    data_module: Any = None
    local_provider: ReferenceProvider | None = None
    require_open_access: bool = True

    def load_gene_catalog(self) -> GeneCatalog:
        if self.key == "fly":
            return build_fly_gene_catalog()
        return self.data_module.load_gene_catalog()

    def reference_providers(self) -> list[ReferenceProvider]:
        providers: list[ReferenceProvider] = []
        if self.local_provider is not None:
            providers.append(self.local_provider)
        if self.data_module is not None:
            providers.extend([
                Gene2PubMedProvider(self.data_module),
                GeneRIFProvider(self.data_module),
                UniProtProvider(self.data_module),
            ])
        providers.extend([
            PubMedSearchProvider(self.pubmed_filter),
            EuropePMCSearchProvider(self.europe_pmc_filter, require_open_access=self.require_open_access),
        ])
        return providers

    def pubmed_organism_filter(self) -> str:
        return self.pubmed_filter

    def europe_pmc_organism_filter(self) -> str:
        return self.europe_pmc_filter


def build_fly_gene_catalog() -> GeneCatalog:
    symbol_map = build_fbgn_to_symbol_map()
    all_names_map = build_fbgn_to_all_names_map()
    catalog = GeneCatalog(species="fly")
    for fbgn_id, symbol in symbol_map.items():
        names = set(all_names_map.get(fbgn_id, set()))
        names.add(symbol)
        record = GeneRecord(
            species="fly",
            gene_id=fbgn_id,
            symbol=symbol,
            authority_id=fbgn_id,
            synonyms={name for name in names if name},
        )
        catalog.genes[fbgn_id] = record
        for name in record.synonyms:
            catalog.symbol_to_gene_id.setdefault(name, fbgn_id)
    return catalog


def get_species_adapter(key: str) -> SpeciesAdapter:
    key = str(key or "fly").strip().lower()
    if key in ("none", "fly"):
        return SpeciesAdapter(
            key="fly",
            display_name="Fly",
            taxon_id=7227,
            gene_id_column="FlyBase_ID",
            symbol_column="Fly_Gene_symbol",
            authority_id_column="FlyBase_ID",
            output_suffix="_classification.xlsx",
            pubmed_filter='("drosophila" OR "fruit fly")',
            europe_pmc_filter='("drosophila" OR "fruit fly") AND "gene"',
            source_priority={"pubmed": 3, "flybase": 2, "europe_pmc": 1},
            local_provider=FlyBaseReferenceProvider(),
        )
    if key == "human":
        return SpeciesAdapter(
            key="human",
            display_name="Human",
            taxon_id=9606,
            gene_id_column="Entrez_Gene_ID",
            symbol_column="Human_Gene_symbol",
            authority_id_column="HGNC_ID",
            output_suffix="_human_classification.xlsx",
            pubmed_filter='Homo sapiens[Organism]',
            europe_pmc_filter='',
            source_priority={"generif": 5, "uniprot": 4, "gene2pubmed": 3, "pubmed": 2, "europe_pmc": 1},
            data_module=human_gene_data,
            require_open_access=False,
        )
    if key == "mouse":
        return SpeciesAdapter(
            key="mouse",
            display_name="Mouse",
            taxon_id=10090,
            gene_id_column="Entrez_Gene_ID",
            symbol_column="Mouse_Gene_symbol",
            authority_id_column="MGI_ID",
            output_suffix="_mouse_classification.xlsx",
            pubmed_filter='Mus musculus[Organism]',
            europe_pmc_filter='',
            source_priority={"generif": 5, "uniprot": 4, "gene2pubmed": 3, "pubmed": 2, "europe_pmc": 1},
            data_module=mouse_gene_data,
            require_open_access=False,
        )
    raise ValueError(f"Unsupported species adapter: {key}")


###############################################################################
# Metadata Fetching
###############################################################################

_meta_cache = {}


@lru_cache(maxsize=1)
def _get_pubmed_fetcher():
    return PubMedFetcher(email="aadish98@gmail.com", api_key=NCBI_API_KEY)


def fetch_title_abstract_pmid(pmid: str) -> tuple[str, str]:
    """Return (title, abstract) for PMID, using persistent cache first."""
    pmid_clean = str(pmid or "").strip()
    if not pmid_clean or not pmid_clean.isdigit():
        return "", ""

    cached = _get_cached_pmid_title_abstract(pmid_clean)
    if cached:
        return str(cached.get("title", "") or ""), str(cached.get("abstract", "") or "")

    try:
        fetcher = _get_pubmed_fetcher()
        art = fetcher.article_by_pmid(pmid_clean)
        title = str(getattr(art, "title", "") or "")
        abstract = str(getattr(art, "abstract", "") or "")
        _set_cached_pmid_title_abstract(pmid_clean, title, abstract)
        return title, abstract
    except Exception:
        return "", ""


def fetch_title_abstract_pmcid(pmcid, pmid: str = ""):
    """Return (title, abstract, year, journal, authors, doi) for PMCID."""
    pmcid_norm = str(pmcid or "").strip().upper()
    if pmcid_norm in _meta_cache:
        return _meta_cache[pmcid_norm]
    
    title_from_pmid, abstract_from_pmid = fetch_title_abstract_pmid(pmid)

    fetcher = _get_pubmed_fetcher()
    try:
        art = fetcher.article_by_pmcid(pmcid_norm)
        yr = art.year or "n.d."
        journal = getattr(art, "journal", "") or ""
        doi = getattr(art, "doi", "") or ""
        authors = []
        try:
            authors = list(getattr(art, "authors", []) or [])
        except Exception:
            authors = []

        title = str(getattr(art, "title", "") or "") or title_from_pmid
        abstract = str(getattr(art, "abstract", "") or "") or abstract_from_pmid
        if pmid and (title or abstract or journal or authors or doi):
            _set_cached_pmid_title_abstract(
                pmid,
                title,
                abstract,
                year=str(yr or ""),
                journal=journal,
                authors=authors,
                doi=doi,
                pmcid=pmcid_norm,
                source="metapub",
            )

        result = (title, abstract, str(yr), journal, authors, doi)
        _meta_cache[pmcid_norm] = result
        return result
    except Exception:
        result = ("", "", "n.d.", "", [], "")
        _meta_cache[pmcid_norm] = result
        return result


def matches_keywords_list(title, abstract, keywords_list):
    """True if ANY keyword appears in title+abstract."""
    if not keywords_list:
        return True
    text = (title + " " + abstract).lower()
    for kw in keywords_list:
        if not kw:
            continue
        kw_lower = kw.strip().lower()
        if not kw_lower:
            continue
        words = re.findall(r'\b\w+\b', text)
        for word in words:
            if kw_lower in word:
                return True
        if kw_lower in text:
            return True
    return False


###############################################################################
# Full Text Retrieval
###############################################################################

def fetch_from_pmcoa_pdf(pmcid):
    """Fetch PDF from PMC OA."""
    pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf"
    r = _http_get(pdf_url, timeout=25)
    if r is not None and r.status_code == 200:
        text = _extract_pdf_text(r.content)
        if len(text) > 300:
            return text
    return ""


def fetch_from_pmcoa_xml(pmcid):
    """Fetch XML from PMC OA."""
    xml_url = f"https://www.ncbi.nlm.nih.gov/pmc/oai/oai.cgi?verb=GetRecord&identifier=oai:pubmedcentral.nih.gov:{pmcid.replace('PMC', '')}&metadataPrefix=pmc"
    r = _http_get(xml_url, timeout=25)
    if r is not None and r.status_code == 200 and len(r.content) > 500:
        text = re.sub(r"<[^>]+>", " ", r.text)
        return re.sub(r"\s+", " ", text).strip()
    return ""


def fetch_from_europepmc_api(identifier):
    """Fetch from Europe PMC API."""
    if identifier.lower().startswith("pmc"):
        id_param = "PMC" + identifier.replace("PMC", "")
    else:
        id_param = identifier
    url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/{id_param}/fullTextXML"
    r = _http_get(url, timeout=25)
    if r is not None and r.status_code == 200 and len(r.content) > 500:
        text = re.sub(r"<[^>]+>", " ", r.text)
        return re.sub(r"\s+", " ", text).strip()
    return ""


def fetch_from_pmc_html(pmcid):
    """Scrape HTML from PMC page."""
    page_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/"
    r = _http_get(page_url, timeout=25)
    if r is not None and r.status_code == 200:
        text = _extract_html_text(r.text, min_chars=1000)
        if text:
            return text
    return ""


def _normalize_doi(doi: str) -> str:
    """Normalize DOI string to bare DOI value."""
    d = str(doi or "").strip()
    if not d:
        return ""
    d_low = d.lower()
    if d_low.startswith("https://doi.org/"):
        return d[len("https://doi.org/"):]
    if d_low.startswith("http://doi.org/"):
        return d[len("http://doi.org/"):]
    if d_low.startswith("doi:"):
        return d[4:].strip()
    return d


def _is_pdf_response(response) -> bool:
    """Heuristic check for PDF content."""
    if response is None:
        return False
    content_type = str(response.headers.get("content-type", "")).lower()
    if "application/pdf" in content_type or "pdf" in content_type:
        return True
    return bool((response.content or b"")[:5] == b"%PDF-")


def _extract_pdf_text(content: bytes) -> str:
    """Extract text from raw PDF bytes."""
    if not content or len(content) <= 500:
        return ""
    try:
        reader = PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            text += "\n"
        return text
    except Exception:
        return ""


def _extract_html_text(html_text: str, min_chars: int = 300) -> str:
    """Extract cleaned text from HTML."""
    if not html_text or len(html_text) <= 300:
        return ""
    txt = re.sub(r"<script[\s\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
    txt = re.sub(r"<style[\s\S]*?</style>", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"<[^>]+>", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    if len(txt) >= min_chars:
        return txt
    return ""


def fetch_full_text_via_unpaywall(doi: str) -> tuple[str, str]:
    """Try to get full text via Unpaywall."""
    if not doi:
        return "", ""
    doi_norm = _normalize_doi(doi)
    url = f"https://api.unpaywall.org/v2/{quote(doi_norm, safe='')}"
    r = _http_get(url, params={"email": UNPAYWALL_TOKEN}, timeout=20)
    if r is None:
        return "", ""
    try:
        data = r.json()
        oa_locs = data.get("oa_locations", []) or []
        best = data.get("best_oa_location")
        if best:
            oa_locs = [best] + [loc for loc in oa_locs if loc != best]

        for loc in oa_locs:
            pdf_url = (loc or {}).get("url_for_pdf") or ""
            page_url = (loc or {}).get("url") or ""

            if pdf_url:
                r_pdf = _http_get(pdf_url, timeout=30)
                if r_pdf is not None and r_pdf.status_code == 200 and _is_pdf_response(r_pdf):
                    text = _extract_pdf_text(r_pdf.content)
                    if len(text) > 300:
                        return text, "Unpaywall PDF"

            if page_url:
                r_page = _http_get(page_url, timeout=25)
                if r_page is not None and r_page.status_code == 200:
                    if _is_pdf_response(r_page):
                        text = _extract_pdf_text(r_page.content)
                        if len(text) > 300:
                            return text, "Unpaywall PDF"
                    text = _extract_html_text(r_page.text, min_chars=300)
                    if text:
                        return text, "Unpaywall HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_openalex(doi: str) -> tuple[str, str]:
    """Try OpenAlex OA links for DOI retrieval."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""

    r = _http_get(
        "https://api.openalex.org/works",
        params={"filter": f"doi:{doi_norm.lower()}", "per-page": 1},
        timeout=20
    )
    if r is None:
        return "", ""
    try:
        data = r.json() or {}
        results = data.get("results", []) or []
        if not results:
            return "", ""
        work = results[0] or {}
        oa_url = (
            ((work.get("open_access") or {}).get("oa_url"))
            or ((work.get("primary_location") or {}).get("pdf_url"))
            or ((work.get("primary_location") or {}).get("landing_page_url"))
        )
        if not oa_url:
            return "", ""

        rr = _http_get(oa_url, timeout=25)
        if rr is None:
            return "", ""
        if _is_pdf_response(rr):
            text = _extract_pdf_text(rr.content)
            if len(text) > 300:
                return text, "OpenAlex PDF"
        text = _extract_html_text(rr.text, min_chars=300)
        if text:
            return text, "OpenAlex HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_crossref(doi: str) -> tuple[str, str]:
    """Try DOI links exposed by Crossref metadata."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""
    r = _http_get(f"https://api.crossref.org/works/{quote(doi_norm, safe='')}", timeout=20)
    if r is None:
        return "", ""
    try:
        message = (r.json() or {}).get("message", {}) or {}
        links = message.get("link", []) or []
        for link in links:
            url = str((link or {}).get("URL", "")).strip()
            if not url:
                continue
            rr = _http_get(url, timeout=25)
            if rr is None:
                continue
            if _is_pdf_response(rr):
                text = _extract_pdf_text(rr.content)
                if len(text) > 300:
                    return text, "Crossref PDF"
            text = _extract_html_text(rr.text, min_chars=300)
            if text:
                return text, "Crossref HTML"
    except Exception:
        return "", ""
    return "", ""


def fetch_full_text_via_doi_resolver(doi: str) -> tuple[str, str]:
    """Last-resort retrieval from DOI landing page."""
    doi_norm = _normalize_doi(doi)
    if not doi_norm:
        return "", ""
    rr = _http_get(f"https://doi.org/{doi_norm}", timeout=25, max_retries=2)
    if rr is None:
        return "", ""
    if _is_pdf_response(rr):
        text = _extract_pdf_text(rr.content)
        if len(text) > 300:
            return text, "DOI Resolver PDF"
    text = _extract_html_text(rr.text, min_chars=300)
    if text:
        return text, "DOI Resolver HTML"
    return "", ""


def enrich_missing_ids_from_europepmc(pmcid: str = "", doi: str = "") -> tuple[str, str]:
    """Resolve missing PMCID/DOI via Europe PMC search."""
    pmcid_clean = str(pmcid or "").strip().upper()
    doi_clean = _normalize_doi(doi)
    query = ""
    if pmcid_clean:
        query = f"PMCID:{pmcid_clean}"
    elif doi_clean:
        query = f'DOI:"{doi_clean}"'
    else:
        return "", ""

    r = _http_get(
        "https://www.ebi.ac.uk/europepmc/webservices/rest/search",
        params={"query": query, "format": "json", "pageSize": 1},
        timeout=20
    )
    if r is None:
        return "", ""
    try:
        data = r.json() or {}
        results = (data.get("resultList") or {}).get("result", []) or []
        if not results:
            return "", ""
        rec = results[0] or {}
        out_pmcid = str(rec.get("pmcid", "") or "").strip().upper()
        out_doi = _normalize_doi(rec.get("doi", ""))
        return out_pmcid, out_doi
    except Exception:
        return "", ""


def fetch_full_text_by_doi(doi: str) -> tuple[str, str]:
    """DOI-based full-text retrieval cascade."""
    for method in [
        fetch_full_text_via_unpaywall,
        fetch_full_text_via_openalex,
        fetch_full_text_via_crossref,
        fetch_full_text_via_doi_resolver,
    ]:
        try:
            text, label = method(doi)
            if text and len(text) > 300:
                return text, label
        except Exception:
            continue
    return "", ""


def _coerce_fulltext_from_cached_method(cached_method: str, pmcid: str, doi_norm: str) -> tuple[str, str]:
    """Replay one known-good method label and return a valid full-text match."""
    for label, method in [
        ("PMC OA PDF", lambda: (fetch_from_pmcoa_pdf(pmcid), "PMC OA PDF") if pmcid else ("", "")),
        ("PMC OA XML", lambda: (fetch_from_pmcoa_xml(pmcid), "PMC OA XML") if pmcid else ("", "")),
        ("Europe PMC", lambda: (fetch_from_europepmc_api(pmcid), "Europe PMC") if pmcid else ("", "")),
        ("PMC HTML", lambda: (fetch_from_pmc_html(pmcid), "PMC HTML") if pmcid else ("", "")),
    ]:
        if cached_method != label:
            continue
        text, out_label = method()
        if text and len(text) > 300:
            return text, out_label

    for marker, method in [
        ("Unpaywall", fetch_full_text_via_unpaywall),
        ("OpenAlex", fetch_full_text_via_openalex),
        ("Crossref", fetch_full_text_via_crossref),
        ("DOI Resolver", fetch_full_text_via_doi_resolver),
    ]:
        if marker not in cached_method or not doi_norm:
            continue
        text, out_label = method(doi_norm)
        if text and len(text) > 300:
            return text, out_label
    return "", ""


def fetch_full_text_by_id(identifier, id_type="pmcid", doi=None, pmid=None):
    """Cascade through methods to get full text."""
    pmcid = str(identifier or "").strip().upper()
    doi_norm = _normalize_doi(doi)
    pmid_norm = str(pmid or "").strip()

    # Try cached known-good method for this PMID first.
    cached_method = _get_cached_fulltext_method(pmid_norm) if pmid_norm else None
    if cached_method:
        try:
            text, label = _coerce_fulltext_from_cached_method(cached_method, pmcid, doi_norm)
            if text and len(text) > 300:
                return text, label
        except Exception:
            pass

    if not doi_norm and pmcid:
        _epmc_pmcid, epmc_doi = enrich_missing_ids_from_europepmc(pmcid=pmcid, doi=doi_norm)
        if epmc_doi and not doi_norm:
            doi_norm = epmc_doi

    for method, label in [
        (fetch_from_pmcoa_pdf, "PMC OA PDF"),
        (fetch_from_pmcoa_xml, "PMC OA XML"),
        (fetch_from_europepmc_api, "Europe PMC"),
        (fetch_from_pmc_html, "PMC HTML"),
    ]:
        try:
            text = method(pmcid)
            if len(text) > 300:
                if pmid_norm:
                    _set_cached_fulltext_method(pmid_norm, label)
                return text, label
        except Exception:
            continue
    
    # DOI-based fallback cascade
    if doi_norm:
        try:
            text, label = fetch_full_text_by_doi(doi_norm)
            if text and len(text) > 300:
                if pmid_norm:
                    _set_cached_fulltext_method(pmid_norm, label)
                return text, label
        except Exception:
            pass
    
    # Fallback: title+abstract (prefer persistent PMID cache when available).
    title, abstract = fetch_title_abstract_pmid(pmid_norm)
    if not title and not abstract:
        title, abstract, *_ = fetch_title_abstract_pmcid(pmcid, pmid=pmid_norm)
    fallback_text = f"TITLE: {title}\n\nABSTRACT: {abstract}"
    return fallback_text, "Title+Abstract only"


###############################################################################
# Summarization
###############################################################################

class FunctionPhenotypeSummary(BaseModel):
    function: str = ""
    phenotypes: str = ""
    skip_reference: bool = False
    skip_reason: str = ""


class FinalFunctionPhenotypeSummary(BaseModel):
    function: str = ""
    phenotypes: str = ""


def _count_words(text: str) -> int:
    """Approximate word count using whitespace-delimited tokens."""
    return len(re.findall(r"\S+", str(text or "")))


def _clean_text(value: Any) -> str:
    """Collapse repeated whitespace and coerce to string."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    """Remove duplicate strings while preserving original order."""
    seen = set()
    out = []
    for value in values or []:
        cleaned = _clean_text(value)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def _join_unique_texts(values: list[str], separator: str = "; ") -> str:
    """Join deduplicated text fragments with a stable separator."""
    return separator.join(_dedupe_preserve_order(values))


def _chunk_text_by_word_budget(text: str, max_words: int) -> list[str]:
    """Split text into deterministic chunks that stay within a word budget."""
    cleaned = str(text or "").strip()
    if not cleaned:
        return [""]

    paragraphs = [
        _clean_text(chunk)
        for chunk in re.split(r"\n\s*\n+", cleaned)
        if _clean_text(chunk)
    ]
    if not paragraphs:
        paragraphs = [_clean_text(cleaned)]

    max_words = max(int(max_words or 0), MIN_CHUNK_WORDS)
    chunks: list[str] = []
    current_parts: list[str] = []
    current_words = 0

    for paragraph in paragraphs:
        paragraph_words = _count_words(paragraph)
        if paragraph_words > max_words:
            if current_parts:
                chunks.append("\n\n".join(current_parts))
                current_parts = []
                current_words = 0

            words = paragraph.split()
            for idx in range(0, len(words), max_words):
                chunk_words = words[idx: idx + max_words]
                if chunk_words:
                    chunks.append(" ".join(chunk_words))
            continue

        if current_parts and current_words + paragraph_words > max_words:
            chunks.append("\n\n".join(current_parts))
            current_parts = [paragraph]
            current_words = paragraph_words
            continue

        current_parts.append(paragraph)
        current_words += paragraph_words

    if current_parts:
        chunks.append("\n\n".join(current_parts))

    return chunks or [cleaned]


def _build_shared_prompt_chunks(
    full_text: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    title: str = "",
    abstract: str = "",
) -> list[str]:
    """Create deterministic chunks sized to fit within the shared prompt budget."""
    cleaned_full_text = str(full_text or "").strip()
    if not cleaned_full_text:
        return [""]

    metadata_blob = " ".join([
        str(gene_symbol or ""),
        str(fbgn_id or ""),
        " ".join(sorted(set(synonyms or []))),
        str(title or ""),
        str(abstract or ""),
    ])
    overhead_words = _count_words(metadata_blob) + SHARED_PROMPT_OVERHEAD_WORDS
    chunk_word_budget = max(MIN_CHUNK_WORDS, MAX_PROMPT_WORDS - overhead_words)

    if _count_words(cleaned_full_text) <= chunk_word_budget:
        return [cleaned_full_text]
    return _chunk_text_by_word_budget(cleaned_full_text, chunk_word_budget)


def _model_dump(value: Any) -> dict[str, Any]:
    """Return a plain dict for a parsed Pydantic model."""
    if value is None:
        return {}
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "dict"):
        return value.dict()
    if isinstance(value, dict):
        return dict(value)
    return {}


def _compact_numeric_dict(payload: dict[str, Any]) -> dict[str, Any]:
    """Keep usage payloads JSON-friendly without preserving provider internals verbatim."""
    out: dict[str, Any] = {}
    for key, value in (payload or {}).items():
        if isinstance(value, dict):
            nested = _compact_numeric_dict(value)
            if nested:
                out[key] = nested
            continue
        if isinstance(value, (int, float, str, bool)) or value is None:
            out[key] = value
    return out


def _extract_openai_usage(response: Any) -> dict[str, Any]:
    """Extract token usage from a Responses API object."""
    usage = getattr(response, "usage", None)
    usage_dict = _model_dump(usage)
    if not usage_dict and isinstance(usage, dict):
        usage_dict = dict(usage)
    return _compact_numeric_dict(usage_dict)


def _record_openai_usage_event(
    response: Any,
    usage_events: Optional[list[dict[str, Any]]],
    usage_context: Optional[dict[str, Any]],
    *,
    model: str,
    max_output_tokens: int,
    reasoning_effort: Optional[str],
    attempt: int,
):
    """Append one successful OpenAI request event for later calibration."""
    if usage_events is None:
        return
    usage = _extract_openai_usage(response)
    event = {
        "timestamp_utc": _utc_now_iso(),
        "request_type": str((usage_context or {}).get("request_type", "structured_completion") or "structured_completion"),
        "model": str(model or ""),
        "max_output_tokens": int(max_output_tokens or 0),
        "attempt": int(attempt),
        "usage": usage,
    }
    if reasoning_effort:
        event["reasoning_effort"] = str(reasoning_effort)
    for key, value in (usage_context or {}).items():
        if key == "request_type":
            continue
        if isinstance(value, (str, int, float, bool)) or value is None:
            event[key] = value
    usage_events.append(event)


def _summarize_openai_usage_events(events: list[dict[str, Any]]) -> dict[str, int]:
    """Aggregate token usage for a cached gene record."""
    totals = {
        "request_count": 0,
        "input_tokens": 0,
        "output_tokens": 0,
        "total_tokens": 0,
        "cached_input_tokens": 0,
        "reasoning_output_tokens": 0,
    }
    for event in events or []:
        usage = event.get("usage", {}) if isinstance(event, dict) else {}
        if not isinstance(usage, dict):
            continue
        totals["request_count"] += 1
        input_tokens = int(_to_float(usage.get("input_tokens")))
        output_tokens = int(_to_float(usage.get("output_tokens")))
        total_tokens = int(_to_float(usage.get("total_tokens"))) or input_tokens + output_tokens
        totals["input_tokens"] += input_tokens
        totals["output_tokens"] += output_tokens
        totals["total_tokens"] += total_tokens
        input_details = usage.get("input_tokens_details", {})
        if isinstance(input_details, dict):
            totals["cached_input_tokens"] += int(_to_float(input_details.get("cached_tokens")))
        output_details = usage.get("output_tokens_details", {})
        if isinstance(output_details, dict):
            totals["reasoning_output_tokens"] += int(_to_float(output_details.get("reasoning_tokens")))
    return totals


def _build_responses_input(messages: list[dict[str, str]]) -> tuple[Optional[str], list[dict[str, str]]]:
    """Split chat-style messages into Responses API instructions and input items."""
    instructions_parts: list[str] = []
    input_items: list[dict[str, str]] = []
    for message in messages or []:
        role = str(message.get("role", "user") or "user").strip()
        content = str(message.get("content", "") or "").strip()
        if not content:
            continue
        if role == "system":
            instructions_parts.append(content)
        else:
            input_items.append({"role": role, "content": content})
    instructions = "\n\n".join(instructions_parts).strip()
    return (instructions or None), input_items


def _get_reasoning_config(
    model_name: str,
    reasoning_effort: Optional[str] = None,
) -> Optional[dict[str, str]]:
    """Use Responses API reasoning controls for GPT-5 family models when configured."""
    model = str(model_name or "").strip().lower()
    effort = str(reasoning_effort or "").strip().lower()
    if model.startswith("gpt-5") and effort:
        return {"effort": effort}
    return None


def _parse_structured_completion(
    messages: list[dict[str, str]],
    response_format,
    model_name: Optional[str] = None,
    max_output_tokens: int = 1500,
    reasoning_effort: Optional[str] = None,
    usage_events: Optional[list[dict[str, Any]]] = None,
    usage_context: Optional[dict[str, Any]] = None,
):
    """Call the Responses API with structured parsing and light retries."""
    last_err = None
    model = str(model_name or "").strip()
    if not model:
        raise ValueError("model_name is required for structured OpenAI calls")
    instructions, input_items = _build_responses_input(messages)
    for attempt in range(3):
        try:
            request_kwargs: dict[str, Any] = {
                "model": model,
                "input": input_items,
                "text_format": response_format,
                "max_output_tokens": max_output_tokens,
            }
            if instructions:
                request_kwargs["instructions"] = instructions
            reasoning = _get_reasoning_config(model, reasoning_effort=reasoning_effort)
            if reasoning is not None:
                request_kwargs["reasoning"] = reasoning

            response = _get_openai_client().responses.parse(**request_kwargs)
            parsed = getattr(response, "output_parsed", None)
            if parsed is not None:
                _record_openai_usage_event(
                    response,
                    usage_events,
                    usage_context,
                    model=model,
                    max_output_tokens=max_output_tokens,
                    reasoning_effort=reasoning_effort,
                    attempt=attempt + 1,
                )
                return parsed
            raw_output = str(getattr(response, "output_text", "") or "").strip()
            raise ValueError(f"Empty parsed response from API: {raw_output or 'no output text returned'}")
        except Exception as e:
            last_err = e
            error_text = str(e)
            if "401" in error_text or "Unauthorized" in error_text or "Incorrect API key" in error_text:
                raise RuntimeError(f"OpenAI authentication failed: {error_text}") from e
            if "Unsupported parameter: 'reasoning.effort'" in str(e):
                try:
                    retry_kwargs: dict[str, Any] = {
                        "model": model,
                        "input": input_items,
                        "text_format": response_format,
                        "max_output_tokens": max_output_tokens,
                    }
                    if instructions:
                        retry_kwargs["instructions"] = instructions
                    response = _get_openai_client().responses.parse(**retry_kwargs)
                    parsed = getattr(response, "output_parsed", None)
                    if parsed is not None:
                        _record_openai_usage_event(
                            response,
                            usage_events,
                            usage_context,
                            model=model,
                            max_output_tokens=max_output_tokens,
                            reasoning_effort=None,
                            attempt=attempt + 1,
                        )
                        return parsed
                    raw_output = str(getattr(response, "output_text", "") or "").strip()
                    raise ValueError(f"Empty parsed response from API: {raw_output or 'no output text returned'}")
                except Exception as retry_error:
                    last_err = retry_error
                    retry_error_text = str(retry_error)
                    if "401" in retry_error_text or "Unauthorized" in retry_error_text or "Incorrect API key" in retry_error_text:
                        raise RuntimeError(f"OpenAI authentication failed: {retry_error_text}") from retry_error
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(str(last_err))


def _build_summary_messages(
    text_chunk: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    title: str = "",
    abstract: str = "",
) -> list[dict[str, str]]:
    synonyms_str = ", ".join(sorted(set(synonyms))) if synonyms else ""
    sys = (
        "You are an expert biomedical assistant. Use ONLY the provided title, abstract, "
        "and text chunk. Focus only on the target gene. If the chunk contains no direct "
        "evidence about the target gene, set skip_reference to true."
    )
    user = f"""Assess evidence for gene {gene_symbol} (Gene ID: {fbgn_id}; Synonyms: {synonyms_str}).

Return JSON with fields:
- function: concise summary of gene {gene_symbol} function evidence from this chunk only
- phenotypes: concise summary of phenotype evidence of gene {gene_symbol} from this chunk only
- skip_reference: true if this chunk does not contain direct evidence for {gene_symbol}
- skip_reason: brief reason when skip_reference is true

Rules:
- Always name {gene_symbol} explicitly
- Be factual and use only the provided text
- If evidence is generic or unrelated to {gene_symbol}, set skip_reference to true
- If one field is absent but the chunk still contains evidence, return an empty string for that field

Title: {title or ''}
Abstract: {abstract or ''}

Text chunk:
{text_chunk or ''}"""
    return [{"role": "system", "content": sys}, {"role": "user", "content": user}]


def _build_rewrite_messages(
    gene_symbol: str,
    fbgn_id: str,
    chunk_summaries: list[dict[str, Any]],
) -> list[dict[str, str]]:
    chunk_lines = []
    for idx, chunk_summary in enumerate(chunk_summaries or [], start=1):
        function_text = _clean_text(chunk_summary.get("function", ""))
        phenotype_text = _clean_text(chunk_summary.get("phenotypes", ""))
        chunk_lines.append(
            f"Chunk {idx}\n"
            f"Function: {function_text or 'None'}\n"
            f"Phenotypes: {phenotype_text or 'None'}"
        )
    merged_input = "\n\n".join(chunk_lines)
    sys = (
        "You are an expert biomedical assistant. Merge chunk-level evidence summaries into a "
        "single cohesive paper-level summary. Use ONLY the provided chunk summaries."
    )
    user = f"""Rewrite these chunk-level summaries for gene {gene_symbol} (Gene ID: {fbgn_id}) into one cohesive paper summary.

Return JSON with fields:
- function
- phenotypes

Rules:
- Always name {gene_symbol} explicitly
- Preserve factual content only
- Remove redundancy across chunks
- Do not introduce new claims

Chunk summaries:
{merged_input}"""
    return [{"role": "system", "content": sys}, {"role": "user", "content": user}]


def _build_classification_messages(
    gene_symbol: str,
    keywords_list: list[str],
    full_text: str,
) -> list[dict[str, str]]:
    cats = [k.strip() for k in keywords_list if k and k.strip() and k.strip().lower() != "none"]
    cats_line = ", ".join(cats) if cats else "None"
    sys = (
        "You are an expert biomedical assistant. Use ONLY the provided text. "
        "If evidence is weak or ambiguous, return category as empty list [] or None."
    )
    user = (
        f"Classify gene '{gene_symbol}'. Allowed categories: [{cats_line}]. "
        "You may classify into 1 category, N categories, or None (return empty list [] or None). "
        "Return JSON with fields: gene, category (list of strings, can be empty or None), confidence (0-100), rationale (1-2 lines).\n\n"
        "TEXT START\n" + full_text[:200000] + "\nTEXT END"
    )
    return [{"role": "system", "content": sys}, {"role": "user", "content": user}]


def summarize_reference_function(
    text_chunk: str,
    gene_symbol: str,
    fbgn_id: str,
    synonyms: list[str],
    model_name: Optional[str] = None,
    title: str = "",
    abstract: str = "",
    chunk_index: int = 1,
    total_chunks: int = 1,
    paper_id: str = "",
    usage_events: Optional[list[dict[str, Any]]] = None,
) -> dict[str, Any]:
    """Extract function and phenotype evidence for one chunk of a paper."""
    try:
        summary_model = _get_summary_model(model_name)
        out = _parse_structured_completion(
            _build_summary_messages(text_chunk, gene_symbol, fbgn_id, synonyms, title=title, abstract=abstract),
            response_format=FunctionPhenotypeSummary,
            model_name=summary_model,
            max_output_tokens=OUTPUT_CAP_PER_REF,
            reasoning_effort=_get_reasoning_effort(
                "OPENAI_SUMMARY_REASONING_EFFORT",
                DEFAULT_SUMMARY_REASONING_EFFORT,
            ),
            usage_events=usage_events,
            usage_context={
                "request_type": "summary_chunk",
                "gene_symbol": gene_symbol,
                "gene_id": fbgn_id,
                "paper_id": paper_id,
                "chunk_index": int(chunk_index),
                "total_chunks": int(total_chunks),
                "chunk_words": _count_words(text_chunk),
            },
        )
        parsed = _model_dump(out)
        function_text = _clean_text(parsed.get("function", ""))
        phenotype_text = _clean_text(parsed.get("phenotypes", ""))
        skip_reference = bool(parsed.get("skip_reference", False))
        skip_reason = _clean_text(parsed.get("skip_reason", ""))
        if not skip_reference and not function_text and not phenotype_text:
            skip_reference = True
            skip_reason = skip_reason or "No function or phenotype evidence returned"
        return {
            "function": function_text,
            "phenotypes": phenotype_text,
            "skip_reference": skip_reference,
            "skip_reason": skip_reason,
        }
    except Exception as e:
        if "OpenAI authentication failed" in str(e):
            raise
        return {
            "function": "",
            "phenotypes": "",
            "skip_reference": True,
            "skip_reason": f"API error: {e}",
        }


def rewrite_function_summary(
    gene_symbol: str,
    fbgn_id: str,
    chunk_summaries: list[dict[str, Any]],
    model_name: Optional[str] = None,
    paper_id: str = "",
    usage_events: Optional[list[dict[str, Any]]] = None,
) -> dict[str, str]:
    """Rewrite multiple chunk-level summaries into one cohesive paper summary."""
    fallback = {
        "function": _join_unique_texts([summary.get("function", "") for summary in chunk_summaries]),
        "phenotypes": _join_unique_texts([summary.get("phenotypes", "") for summary in chunk_summaries]),
    }
    try:
        summary_model = _get_summary_model(model_name)
        out = _parse_structured_completion(
            _build_rewrite_messages(gene_symbol, fbgn_id, chunk_summaries),
            response_format=FinalFunctionPhenotypeSummary,
            model_name=summary_model,
            max_output_tokens=OUTPUT_CAP_PER_REF,
            reasoning_effort=_get_reasoning_effort(
                "OPENAI_SUMMARY_REASONING_EFFORT",
                DEFAULT_SUMMARY_REASONING_EFFORT,
            ),
            usage_events=usage_events,
            usage_context={
                "request_type": "summary_rewrite",
                "gene_symbol": gene_symbol,
                "gene_id": fbgn_id,
                "paper_id": paper_id,
                "chunk_count": len(chunk_summaries or []),
            },
        )
        parsed = _model_dump(out)
        return {
            "function": _clean_text(parsed.get("function", "")) or fallback["function"],
            "phenotypes": _clean_text(parsed.get("phenotypes", "")) or fallback["phenotypes"],
        }
    except Exception as e:
        if "OpenAI authentication failed" in str(e):
            raise
        return fallback


def _format_reference_summary(function_text: str, phenotype_text: str) -> str:
    """Keep a compact human-readable summary for the existing summary sheet."""
    return (
        f"- Function: {function_text or 'None reported'}\n"
        f"- Phenotypes: {phenotype_text or 'None reported'}"
    )


def gene_mentioned_in_title_abstract(title: str, abstract: str, gene_symbol: str, synonyms: list[str]) -> bool:
    """Check if gene is mentioned in title/abstract."""
    if not title and not abstract:
        return False
    combined_text = f"{title or ''} {abstract or ''}".lower()
    
    if gene_symbol:
        gene_lower = gene_symbol.lower().strip()
        if gene_lower and len(gene_lower) >= 2:
            if re.search(rf"\b{re.escape(gene_lower)}\b", combined_text):
                return True
    
    if synonyms:
        for synonym in synonyms:
            if synonym:
                syn_lower = synonym.lower().strip()
                if syn_lower and len(syn_lower) >= 2:
                    if re.search(rf"\b{re.escape(syn_lower)}\b", combined_text):
                        return True
    return False


###############################################################################
# Classification
###############################################################################

class GeneClassification(BaseModel):
    gene: str
    category: Optional[List[str]]
    confidence: int
    rationale: str


def _normalize_category_output(raw_category, valid_cats: set[str]) -> list[str]:
    """Normalize model category output into a validated list."""
    if raw_category is None:
        return []
    if isinstance(raw_category, list):
        return [c for c in raw_category if c in valid_cats]
    if isinstance(raw_category, str) and raw_category in valid_cats:
        return [raw_category]
    return []


def _safe_confidence(confidence_value) -> int:
    """Clamp model confidence to a 0-100 integer."""
    try:
        return max(0, min(100, int(confidence_value)))
    except Exception:
        return 0


def classify_gene_from_text(
    gene_symbol: str,
    keywords_list: list[str],
    full_text: str,
    model_name: Optional[str] = None,
    gene_id: str = "",
    usage_events: Optional[list[dict[str, Any]]] = None,
    high_quality_reference_count: int = 0,
):
    """Classify a gene using GPT."""
    if not full_text or not full_text.strip():
        return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": "No text provided for classification"}
    
    cats = [k.strip() for k in keywords_list if k and k.strip() and k.strip().lower() != "none"]
    valid_cats = set(cats) if cats else set()
    messages = _build_classification_messages(gene_symbol, keywords_list, full_text)

    try:
        classification_model = _get_classification_model(model_name)
        out = _parse_structured_completion(
            messages,
            response_format=GeneClassification,
            model_name=classification_model,
            max_output_tokens=OUTPUT_CAP_PER_CLASSIFICATION,
            reasoning_effort=_get_reasoning_effort(
                "OPENAI_CLASSIFICATION_REASONING_EFFORT",
                DEFAULT_CLASSIFICATION_REASONING_EFFORT,
            ),
            usage_events=usage_events,
            usage_context={
                "request_type": "classification",
                "gene_symbol": gene_symbol,
                "gene_id": gene_id,
                "high_quality_reference_count": int(high_quality_reference_count),
                "classification_text_words": _count_words(full_text),
            },
        )
        if not out:
            return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": "Empty response from API"}
        
        category_list = _normalize_category_output(out.category, valid_cats)
        confidence = _safe_confidence(out.confidence)
        
        rationale = str(out.rationale) if hasattr(out, 'rationale') and out.rationale else ""
        
        return {
            "gene": gene_symbol,
            "category": category_list,
            "confidence": confidence,
            "rationale": rationale
        }
    except Exception as e:
        error_msg = str(e)
        if "length limit was reached" in error_msg or "maximum context length" in error_msg:
            try:
                shorter_text = full_text[:50000]
                out = _parse_structured_completion(
                    _build_classification_messages(gene_symbol, keywords_list, shorter_text),
                    response_format=GeneClassification,
                    model_name=_get_classification_model(model_name),
                    max_output_tokens=2000,
                    reasoning_effort=_get_reasoning_effort(
                        "OPENAI_CLASSIFICATION_REASONING_EFFORT",
                        DEFAULT_CLASSIFICATION_REASONING_EFFORT,
                    ),
                    usage_events=usage_events,
                    usage_context={
                        "request_type": "classification_retry_short_context",
                        "gene_symbol": gene_symbol,
                        "gene_id": gene_id,
                        "high_quality_reference_count": int(high_quality_reference_count),
                        "classification_text_words": _count_words(shorter_text),
                    },
                )
                if out:
                    category_list = _normalize_category_output(out.category, valid_cats)
                    return {
                        "gene": gene_symbol,
                        "category": category_list,
                        "confidence": _safe_confidence(out.confidence or 0),
                        "rationale": str(out.rationale or "")
                    }
            except Exception as e2:
                return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": f"Error: {e2}"}
        return {"gene": gene_symbol, "category": [], "confidence": 0, "rationale": f"Error: {e}"}


###############################################################################
# Main Processing Pipeline
###############################################################################

def process_gene_set(
    genes: list[GeneRecord] | list[str],
    keywords_list: list[str],
    reference_limit: int = 500,
    adapter: SpeciesAdapter | None = None,
):
    """Process a normalized gene set through the shared classification pipeline."""
    adapter = adapter or get_species_adapter("fly")
    if genes and isinstance(genes[0], str):
        catalog = adapter.load_gene_catalog()
        genes = [catalog.get(str(g).strip()) for g in genes]  # type: ignore[index]
        genes = [gene for gene in genes if gene is not None]

    genes = [gene for gene in (genes or []) if isinstance(gene, GeneRecord) and gene.gene_id]
    print(f"\n{'='*60}")
    print(f"Processing {len(genes)} {adapter.display_name} genes...")
    print(f"Keywords: {keywords_list}")
    print(f"Reference limit per gene: {reference_limit}")
    print(f"{'='*60}\n")
    log_step("Validating gene records", f"{len(genes)} candidate genes", indent=1)

    validated: dict[str, GeneRecord] = {gene.symbol: gene for gene in genes}
    if not validated:
        print("No valid genes found.")
        return {}, [], {}, {}, {}

    if adapter.key == "fly":
        log_step("Loading FlyBase reference maps", indent=1)
    pmcid_to_year = build_pmcid_to_year() if adapter.key == "fly" else {}
    pmcid_to_pmid = build_pmcid_to_pmid() if adapter.key == "fly" else {}
    pmcid_to_doi = build_pmcid_to_doi() if adapter.key == "fly" else {}

    paper_sources: dict[str, set[str]] = defaultdict(set)
    paper_labels: dict[str, dict[str, str]] = defaultdict(dict)
    paper_id_to_pmid: dict[str, str] = {}
    paper_id_to_doi: dict[str, str] = {}
    gene_to_candidates: dict[str, dict[str, ReferenceCandidate]] = defaultdict(dict)

    def _paper_key(paper_id: str) -> str:
        text = str(paper_id or "").strip()
        return text.upper() if text.upper().startswith("PMC") else text

    def _add_candidate(candidate: ReferenceCandidate):
        paper_key = _paper_key(candidate.paper_id)
        if not paper_key:
            return
        candidate.paper_id = paper_key
        existing = gene_to_candidates[candidate.gene_id].get(paper_key)
        if existing:
            existing.source_labels.add(candidate.source_label)
            if candidate.pmid and not existing.pmid:
                existing.pmid = candidate.pmid
            if candidate.pmcid and not existing.pmcid:
                existing.pmcid = candidate.pmcid
            if candidate.snippet and not existing.snippet:
                existing.snippet = candidate.snippet
        else:
            gene_to_candidates[candidate.gene_id][paper_key] = candidate
        paper_sources[paper_key].add(candidate.source_key)
        paper_labels[paper_key][candidate.source_key] = candidate.source_label
        if candidate.pmid:
            paper_id_to_pmid[paper_key] = candidate.pmid
        if candidate.doi:
            paper_id_to_doi[paper_key] = candidate.doi

    providers = adapter.reference_providers()
    log_step("Collecting candidate references", f"{len(providers)} source(s)", indent=1)
    for step_index, provider in enumerate(providers, start=1):
        print(f"\n[Step {step_index}] Fetching references from {provider.source_label}...")
        try:
            collected = provider.collect_many(genes, keywords_list)
        except Exception as e:
            print(f"  [Warning] {provider.source_label} failed: {e}")
            collected = {}
        source_total = 0
        genes_with_refs = 0
        for gene in genes:
            candidates = collected.get(gene.gene_id, []) or []
            source_total += len(candidates)
            if candidates:
                genes_with_refs += 1
            for candidate in candidates:
                _add_candidate(candidate)
        print(
            f"  {provider.source_label}: {source_total} references "
            f"across {genes_with_refs}/{len(genes)} genes"
        )

    def _source_priority(sources: set[str]) -> int:
        if len(sources) >= 2:
            return 100 + len(sources)
        return max((adapter.source_priority.get(source, 0) for source in sources), default=0)

    def _paper_numeric(paper_id: str) -> int:
        try:
            return int(re.sub(r"\D+", "", str(paper_id)))
        except Exception:
            return 0

    def _sort_key_sources_then_recency(paper_id: str):
        paper_norm = _paper_key(paper_id)
        year = pmcid_to_year.get(paper_norm, 0)
        return (_source_priority(paper_sources.get(paper_norm, set())), year, _paper_numeric(paper_norm))

    def _format_ref_source(paper_id: str) -> str:
        paper_norm = _paper_key(paper_id)
        labels_by_key = paper_labels.get(paper_norm, {})
        ordered_keys = sorted(
            labels_by_key,
            key=lambda key: adapter.source_priority.get(key, 0),
            reverse=True,
        )
        labels = [labels_by_key[key] for key in ordered_keys]
        return ", ".join(labels) if labels else "Unknown"

    gene_to_papers = {}
    for gene_id, candidate_map in gene_to_candidates.items():
        sorted_papers = sorted(candidate_map.keys(), key=_sort_key_sources_then_recency, reverse=True)
        gene_to_papers[gene_id] = set(sorted_papers[:reference_limit])
    calibration_limited_references = sum(len(papers) for papers in gene_to_papers.values())

    log_step("Ranking and limiting references", f"limit={reference_limit} per gene", indent=1)
    print("\n[Metadata] Filtering references by keywords and fetching metadata...")
    all_paper_ids = set()
    for papers in gene_to_papers.values():
        all_paper_ids.update(papers)
    log_step("Fetching metadata and applying keyword filter", f"{len(all_paper_ids)} unique references", indent=1)

    meta_cache = {}
    high_quality_papers_per_gene = defaultdict(set)
    total = len(all_paper_ids)
    for i, paper_id in enumerate(sorted(all_paper_ids, key=_sort_key_sources_then_recency, reverse=True)):
        if (i + 1) % 20 == 0:
            print(f"  Processing metadata {i+1}/{total}...")
        paper_norm = _paper_key(paper_id)
        try:
            pmid_for_paper = (
                paper_id_to_pmid.get(paper_norm, "")
                or (paper_norm if paper_norm.isdigit() else "")
                or pmcid_to_pmid.get(paper_norm, "")
            )
            resolved = resolve_reference_metadata(
                paper_norm,
                _get_pubmed_fetcher(),
                api_key=NCBI_API_KEY or "",
                pmid_hint=pmid_for_paper,
                cache_getter=_get_cached_pmid_title_abstract,
                cache_setter=_set_cached_pmid_title_abstract,
            )
            title = str(resolved.get("title", "") or "")
            abstract = str(resolved.get("abstract", "") or "")
            year = str(resolved.get("year", "") or "")
            journal = str(resolved.get("journal", "") or "")
            authors = normalize_authors(resolved.get("authors", []))
            doi = str(resolved.get("doi", "") or "") or paper_id_to_doi.get(paper_norm, "")
            if not year:
                year_val = pmcid_to_year.get(paper_norm, 0)
                year = str(year_val) if year_val else "n.d."
            if not doi:
                doi = pmcid_to_doi.get(paper_norm, "")
            meta_cache[paper_norm] = (title, abstract, year, journal, authors, doi)
            for gene in genes:
                if paper_norm in gene_to_papers.get(gene.gene_id, set()):
                    candidate = gene_to_candidates.get(gene.gene_id, {}).get(paper_norm)
                    snippet = candidate.snippet if candidate else ""
                    if matches_keywords_list(title, abstract, keywords_list) or matches_keywords_list("", snippet, keywords_list):
                        high_quality_papers_per_gene[gene.gene_id].add(paper_norm)
        except Exception:
            meta_cache[paper_norm] = ("", "", "n.d.", "", [], "")
    calibration_metadata_keyword_matches = sum(len(papers) for papers in high_quality_papers_per_gene.values())

    print("\n[Summaries] Summarizing and classifying genes...")
    log_step("Retrieving full text, summarizing evidence, and classifying genes", indent=1)
    gene_to_summary_chunks = defaultdict(list)
    gene_to_papers_all = defaultdict(set)
    gene_high_quality_ref_summaries = defaultdict(list)
    gene_to_full_text_papers = defaultdict(set)
    gene_to_openai_usage_events: dict[str, list[dict[str, Any]]] = defaultdict(list)
    all_summaries = []
    gene_hits = {}

    for gene in genes:
        print(f"\n  Processing: {gene.symbol} ({gene.display_id})")
        synonyms = list(gene.synonyms or set())
        papers_for_gene = sorted(
            list(high_quality_papers_per_gene.get(gene.gene_id, set())),
            key=_sort_key_sources_then_recency,
            reverse=True,
        )
        print(f"    Candidate references after keyword filter: {len(papers_for_gene)}")
        high_q_count = 0
        total_attempts = 0
        for paper_id in papers_for_gene:
            if total_attempts >= 50 or high_q_count >= 6:
                break
            meta_data = meta_cache.get(paper_id)
            if not meta_data or len(meta_data) != 6:
                continue
            title, abstract, year, journal, authors, doi = meta_data
            candidate = gene_to_candidates.get(gene.gene_id, {}).get(paper_id)
            pmid_for_paper = (
                (candidate.pmid if candidate else "")
                or paper_id_to_pmid.get(paper_id, "")
                or (paper_id if paper_id.isdigit() else "")
            )
            pmcid_for_paper = (candidate.pmcid if candidate else "") or (paper_id if paper_id.startswith("PMC") else "")
            if pmcid_for_paper:
                full_text, source_label = fetch_full_text_by_id(
                    pmcid_for_paper,
                    "pmcid",
                    doi=doi,
                    pmid=pmid_for_paper,
                )
            else:
                full_text, source_label = fetch_full_text_by_id(
                    "",
                    "pmcid",
                    doi=doi,
                    pmid=pmid_for_paper,
                )
            if candidate and candidate.snippet and source_label == "Title+Abstract only":
                full_text = f"{full_text}\n\nGENERIF: {candidate.snippet}"
            if full_text and source_label != "Title+Abstract only":
                gene_to_full_text_papers[gene.gene_id].add(paper_id)
            has_title_abstract_match = gene_mentioned_in_title_abstract(
                title or "", abstract or "", gene.symbol, synonyms
            ) or bool(candidate and candidate.snippet and gene_mentioned_in_title_abstract("", candidate.snippet, gene.symbol, synonyms))
            if not (full_text and source_label != "Title+Abstract only") and not has_title_abstract_match:
                continue

            shared_chunks = _build_shared_prompt_chunks(
                full_text if source_label != "Title+Abstract only" else "",
                gene.symbol,
                gene.gene_id,
                synonyms,
                title=title,
                abstract=abstract,
            )
            total_chunks = len(shared_chunks)
            chunk_word_counts = [_count_words(chunk) for chunk in shared_chunks]
            function_chunk_summaries = []
            skip_reasons = []
            for chunk_index, text_chunk in enumerate(shared_chunks, start=1):
                chunk_summary = summarize_reference_function(
                    text_chunk,
                    gene.symbol,
                    gene.gene_id,
                    synonyms,
                    title=title,
                    abstract=abstract,
                    chunk_index=chunk_index,
                    total_chunks=total_chunks,
                    paper_id=paper_id,
                    usage_events=gene_to_openai_usage_events[gene.gene_id],
                )
                if chunk_summary.get("skip_reference", True):
                    if chunk_summary.get("skip_reason"):
                        skip_reasons.append(chunk_summary["skip_reason"])
                else:
                    function_chunk_summaries.append(chunk_summary)

            total_attempts += 1
            if function_chunk_summaries:
                if len(function_chunk_summaries) == 1:
                    final_function_summary = {
                        "function": _clean_text(function_chunk_summaries[0].get("function", "")),
                        "phenotypes": _clean_text(function_chunk_summaries[0].get("phenotypes", "")),
                    }
                else:
                    final_function_summary = rewrite_function_summary(
                        gene.symbol,
                        gene.gene_id,
                        function_chunk_summaries,
                        paper_id=paper_id,
                        usage_events=gene_to_openai_usage_events[gene.gene_id],
                    )
                summary = _format_reference_summary(
                    final_function_summary.get("function", ""),
                    final_function_summary.get("phenotypes", ""),
                )
                is_high_quality = True
                qc_justification = "Passed QC"
            else:
                final_function_summary = {"function": "", "phenotypes": ""}
                summary = "No evidence found, skip reference"
                is_high_quality = False
                qc_justification = _join_unique_texts(skip_reasons) or f"No evidence found across {total_chunks} chunk(s)"

            all_summaries.append({
                "paper_id": paper_id,
                "title": title,
                "year": year,
                "journal": journal,
                "authors": authors,
                "gene_symbol": gene.symbol,
                "gene_id": gene.gene_id,
                "authority_id": gene.authority_id,
                "species": gene.species,
                "trace": dict(gene.trace or {}),
                "summary": summary,
                "function_text": final_function_summary.get("function", ""),
                "phenotypes_text": final_function_summary.get("phenotypes", ""),
                "abstract_text": abstract or "",
                "full_text_source": source_label,
                "full_text_word_count": _count_words(full_text if source_label != "Title+Abstract only" else ""),
                "chunk_count": total_chunks,
                "chunk_word_counts": chunk_word_counts,
                "is_high_quality": is_high_quality,
                "qc_justification": qc_justification,
                "source": _format_ref_source(paper_id),
            })
            gene_to_papers_all[gene.gene_id].add(paper_id)
            if is_high_quality:
                high_q_count += 1
                gene_to_summary_chunks[gene.gene_id].append(summary)
                gene_high_quality_ref_summaries[gene.gene_id].append(paper_id)
                print(f"    [{high_q_count}/6] {paper_id}: {qc_justification}")

        chunks = gene_to_summary_chunks.get(gene.gene_id, [])
        all_high_quality_papers = gene_to_papers_all.get(gene.gene_id, set())
        supporting_papers = set(gene_high_quality_ref_summaries.get(gene.gene_id, []))
        full_text_papers = gene_to_full_text_papers.get(gene.gene_id, set())
        if chunks:
            print(f"    Classifying from {len(chunks)} high-quality summary chunk(s)")
            agg_text = "\n\n".join(chunks)
            cls = classify_gene_from_text(
                gene.symbol,
                keywords_list,
                agg_text,
                gene_id=gene.gene_id,
                usage_events=gene_to_openai_usage_events[gene.gene_id],
                high_quality_reference_count=len(chunks),
            )
            if not cls or not isinstance(cls, dict):
                cls = {"gene": gene.symbol, "category": [], "confidence": 0, "rationale": "Classification failed"}
            category_list = cls.get("category", [])
            category_str = "; ".join(category_list) if isinstance(category_list, list) and category_list else str(category_list or "None")
            gene_hits[gene.symbol] = {
                "pmcids": all_high_quality_papers,
                "category": category_str,
                "confidence": int(cls.get("confidence", 0)),
                "rationale": cls.get("rationale", ""),
                "classified_by": "GPT",
                "supporting_refs": supporting_papers,
                "full_text_refs": full_text_papers,
            }
            print(f"    Classification: {category_str} (confidence: {cls.get('confidence', 0)})")
        else:
            gene_hits[gene.symbol] = {
                "pmcids": all_high_quality_papers,
                "category": "None",
                "confidence": 0,
                "rationale": "No relevant references found.",
                "classified_by": "Self",
                "supporting_refs": set(),
                "full_text_refs": full_text_papers,
            }
            print("    Classification: None (no relevant references)")

    log_step("Gene set complete", f"{len(gene_hits)} genes classified, {len(all_summaries)} reference summaries", indent=1)
    calibration_stats = {
        "species": adapter.key,
        "limited_references": int(calibration_limited_references),
        "metadata_keyword_matches": int(calibration_metadata_keyword_matches),
    }
    return gene_hits, all_summaries, validated, dict(gene_to_openai_usage_events), calibration_stats


def _plan_soft_run_gene_set(
    genes: list[GeneRecord],
    keywords_list: list[str],
    reference_limit: int = 500,
    adapter: SpeciesAdapter | None = None,
) -> SoftRunFileEstimate:
    """Plan OpenAI-bound references using fast candidate counts; never fetch full text."""
    adapter = adapter or get_species_adapter("fly")
    genes = [gene for gene in (genes or []) if isinstance(gene, GeneRecord) and gene.gene_id]
    estimate = SoftRunFileEstimate(csv_path="", total_genes=len(genes), planned_genes=len(genes))
    if not genes:
        return estimate

    pass_rate_info = _resolve_species_keyword_pass_rate(adapter.key)
    keyword_pass_rate = _to_float(pass_rate_info.get("keyword_pass_rate"), default=1.0)
    estimate.keyword_pass_rate = keyword_pass_rate
    estimate.keyword_pass_rate_source = str(pass_rate_info.get("source", "") or "fallback")
    estimate.keyword_pass_rate_limited_references = int(_to_float(pass_rate_info.get("calibration_limited_references")))
    estimate.keyword_pass_rate_metadata_keyword_matches = int(
        _to_float(pass_rate_info.get("calibration_metadata_keyword_matches"))
    )
    print(
        f"  Fast soft-run keyword-pass proxy: species={adapter.key}, "
        f"pass_rate={keyword_pass_rate:.4f}, source={estimate.keyword_pass_rate_source}"
    )

    if adapter.key == "fly":
        log_step("Loading FlyBase reference maps", indent=1)
    pmcid_to_year = build_pmcid_to_year() if adapter.key == "fly" else {}

    paper_sources: dict[str, set[str]] = defaultdict(set)
    gene_to_candidates: dict[str, dict[str, ReferenceCandidate]] = defaultdict(dict)

    def _paper_key(paper_id: str) -> str:
        text = str(paper_id or "").strip()
        return text.upper() if text.upper().startswith("PMC") else text

    def _add_candidate(candidate: ReferenceCandidate):
        paper_key = _paper_key(candidate.paper_id)
        if not paper_key:
            return
        candidate.paper_id = paper_key
        existing = gene_to_candidates[candidate.gene_id].get(paper_key)
        if existing:
            existing.source_labels.add(candidate.source_label)
            if candidate.pmid and not existing.pmid:
                existing.pmid = candidate.pmid
            if candidate.pmcid and not existing.pmcid:
                existing.pmcid = candidate.pmcid
            if candidate.snippet and not existing.snippet:
                existing.snippet = candidate.snippet
        else:
            gene_to_candidates[candidate.gene_id][paper_key] = candidate
        paper_sources[paper_key].add(candidate.source_key)

    providers = adapter.reference_providers()
    log_step("Collecting candidate references", f"{len(providers)} source(s)", indent=1)
    for step_index, provider in enumerate(providers, start=1):
        print(f"\n[Soft-run Step {step_index}] Fetching references from {provider.source_label}...")
        try:
            collected = provider.collect_many(genes, keywords_list)
        except Exception as e:
            print(f"  [Warning] {provider.source_label} failed: {e}")
            collected = {}
        source_total = 0
        genes_with_refs = 0
        for gene in genes:
            candidates = collected.get(gene.gene_id, []) or []
            source_total += len(candidates)
            if candidates:
                genes_with_refs += 1
            for candidate in candidates:
                _add_candidate(candidate)
        print(
            f"  {provider.source_label}: {source_total} references "
            f"across {genes_with_refs}/{len(genes)} genes"
        )

    def _source_priority(sources: set[str]) -> int:
        if len(sources) >= 2:
            return 100 + len(sources)
        return max((adapter.source_priority.get(source, 0) for source in sources), default=0)

    def _paper_numeric(paper_id: str) -> int:
        try:
            return int(re.sub(r"\D+", "", str(paper_id)))
        except Exception:
            return 0

    def _sort_key_sources_then_recency(paper_id: str):
        paper_norm = _paper_key(paper_id)
        year = pmcid_to_year.get(paper_norm, 0)
        return (_source_priority(paper_sources.get(paper_norm, set())), year, _paper_numeric(paper_norm))

    gene_to_papers: dict[str, set[str]] = {}
    for gene_id, candidate_map in gene_to_candidates.items():
        estimate.candidate_references += len(candidate_map)
        sorted_papers = sorted(candidate_map.keys(), key=_sort_key_sources_then_recency, reverse=True)
        limited = set(sorted_papers[:reference_limit])
        estimate.limited_references += len(limited)
        gene_to_papers[gene_id] = limited

    log_step("Estimating keyword matches from candidate counts", f"pass_rate={keyword_pass_rate:.4f}", indent=1)
    for papers in gene_to_papers.values():
        limited_count = len(papers)
        estimated_refs_for_gene = min(limited_count * keyword_pass_rate, reference_limit)
        estimate.metadata_keyword_matches += estimated_refs_for_gene
        estimate.n_refs += estimated_refs_for_gene
        if estimated_refs_for_gene >= 1:
            estimate.n_genes_classified += 1
    return estimate


def _estimate_costs(
    n_refs: float,
    n_genes_classified: int,
    summary_pricing: ModelPricing,
    classification_pricing: ModelPricing,
    average_profile: Optional[dict[str, object]] = None,
) -> dict[str, float]:
    summary_input_tokens = n_refs * INPUT_TOKENS_PER_REF
    summary_output_tokens = n_refs * OUTPUT_CAP_PER_REF
    classification_input_tokens = n_genes_classified * INPUT_TOKENS_PER_CLASSIFICATION
    classification_output_tokens = n_genes_classified * OUTPUT_CAP_PER_CLASSIFICATION
    input_cost = (
        (summary_input_tokens / 1_000_000) * summary_pricing.input_per_1m
        + (classification_input_tokens / 1_000_000) * classification_pricing.input_per_1m
    )
    output_cost = (
        (summary_output_tokens / 1_000_000) * summary_pricing.output_per_1m
        + (classification_output_tokens / 1_000_000) * classification_pricing.output_per_1m
    )
    result = {
        "summary_input_tokens": float(summary_input_tokens),
        "summary_output_tokens": float(summary_output_tokens),
        "classification_input_tokens": float(classification_input_tokens),
        "classification_output_tokens": float(classification_output_tokens),
        "input_tokens": float(summary_input_tokens + classification_input_tokens),
        "output_tokens": float(summary_output_tokens + classification_output_tokens),
        "input_cost": input_cost,
        "output_cost": output_cost,
        "total_cost": input_cost + output_cost,
    }
    if average_profile:
        summary_expected_input_tokens = (
            n_refs * _to_float(average_profile.get("summary_input_tokens_per_reference"))
        )
        summary_expected_output_tokens = (
            n_refs * _to_float(average_profile.get("summary_output_tokens_per_reference"))
        )
        classification_expected_input_tokens = (
            n_genes_classified * _to_float(average_profile.get("classification_input_tokens_per_gene"))
        )
        classification_expected_output_tokens = (
            n_genes_classified * _to_float(average_profile.get("classification_output_tokens_per_gene"))
        )
        expected_input_cost = (
            (summary_expected_input_tokens / 1_000_000) * summary_pricing.input_per_1m
            + (classification_expected_input_tokens / 1_000_000) * classification_pricing.input_per_1m
        )
        expected_output_cost = (
            (summary_expected_output_tokens / 1_000_000) * summary_pricing.output_per_1m
            + (classification_expected_output_tokens / 1_000_000) * classification_pricing.output_per_1m
        )
        result.update({
            "expected_summary_input_tokens": float(summary_expected_input_tokens),
            "expected_summary_output_tokens": float(summary_expected_output_tokens),
            "expected_classification_input_tokens": float(classification_expected_input_tokens),
            "expected_classification_output_tokens": float(classification_expected_output_tokens),
            "expected_input_tokens": float(summary_expected_input_tokens + classification_expected_input_tokens),
            "expected_output_tokens": float(summary_expected_output_tokens + classification_expected_output_tokens),
            "expected_model_requests": float(
                n_refs * _to_float(average_profile.get("summary_requests_per_reference"))
                + n_genes_classified * _to_float(average_profile.get("classification_requests_per_gene"), 1.0)
            ),
            "expected_input_cost": expected_input_cost,
            "expected_output_cost": expected_output_cost,
            "expected_total_cost": expected_input_cost + expected_output_cost,
        })
    return result


def _print_soft_run_report(estimate: SoftRunEstimate):
    summary_pricing = estimate.summary_pricing
    classification_pricing = estimate.classification_pricing
    if summary_pricing is None or classification_pricing is None:
        raise RuntimeError("Soft-run estimate is missing model pricing")

    def _money(value: float) -> str:
        return f"${value:,.4f}"

    def _refs(value: float) -> str:
        return f"{value:,.1f}" if abs(float(value or 0) - int(float(value or 0))) > 1e-9 else f"{int(value):,}"

    average_profile = _resolve_soft_run_reference_profile(
        estimate.summary_model,
        estimate.classification_model,
    )

    print(f"\n{'='*70}")
    print("Soft-Run OpenAI Cost Estimate")
    print(f"{'='*70}")
    print("No OpenAI calls were made. No full paper text was fetched.")
    print(f"Summary model: {estimate.summary_model}")
    print(
        f"  Rates: input=${summary_pricing.input_per_1m:g}/1M, "
        f"output=${summary_pricing.output_per_1m:g}/1M"
    )
    print(f"Classification model: {estimate.classification_model}")
    print(
        f"  Rates: input=${classification_pricing.input_per_1m:g}/1M, "
        f"output=${classification_pricing.output_per_1m:g}/1M"
    )
    print("\nAssumptions:")
    print(f"  INPUT_TOKENS_PER_REF={INPUT_TOKENS_PER_REF}")
    print(f"  OUTPUT_CAP_PER_REF={OUTPUT_CAP_PER_REF}")
    print(f"  INPUT_TOKENS_PER_CLASSIFICATION={INPUT_TOKENS_PER_CLASSIFICATION}")
    print(f"  OUTPUT_CAP_PER_CLASSIFICATION={OUTPUT_CAP_PER_CLASSIFICATION}")
    if average_profile:
        print("\nEmpirical reference-average profile:")
        print(f"  {average_profile.get('label', 'unnamed profile')}")
        print(f"  Source: {average_profile.get('source', 'unknown')}")
        print(
            "  Per reference: "
            f"summary_requests={_to_float(average_profile.get('summary_requests_per_reference')):.2f}, "
            f"summary_input_tokens={_to_float(average_profile.get('summary_input_tokens_per_reference')):,.0f}, "
            f"summary_output_tokens={_to_float(average_profile.get('summary_output_tokens_per_reference')):,.0f}"
        )
        print(
            "  Per GPT-classified gene: "
            f"classification_input_tokens={_to_float(average_profile.get('classification_input_tokens_per_gene')):,.0f}, "
            f"classification_output_tokens={_to_float(average_profile.get('classification_output_tokens_per_gene')):,.0f}"
        )

    print("\nPer-file estimates:")
    for file_estimate in estimate.files:
        costs = _estimate_costs(
            file_estimate.n_refs,
            file_estimate.n_genes_classified,
            summary_pricing,
            classification_pricing,
            average_profile,
        )
        print(f"\n  {file_estimate.csv_path}")
        print(
            f"    genes: total={file_estimate.total_genes}, reused={file_estimate.reused_genes}, "
            f"planned={file_estimate.planned_genes}"
        )
        print(
            f"    references: candidates={file_estimate.candidate_references}, "
            f"after_reference_limit={file_estimate.limited_references}, "
            f"estimated_keyword_matches={_refs(file_estimate.metadata_keyword_matches)}"
        )
        print(
            f"    fast keyword-pass proxy: species_rate={file_estimate.keyword_pass_rate:.4f}, "
            f"calibration_limited_refs={file_estimate.keyword_pass_rate_limited_references:,}, "
            f"source={file_estimate.keyword_pass_rate_source}"
        )
        print(
            f"    N_refs={_refs(file_estimate.n_refs)}, "
            f"N_genes_classified={file_estimate.n_genes_classified}"
        )
        print(
            f"    tokens: input={int(costs['input_tokens']):,}, "
            f"capped_output={int(costs['output_tokens']):,}"
        )
        if "expected_total_cost" in costs:
            print(
                f"    reference-average tokens: input={int(costs['expected_input_tokens']):,}, "
                f"output={int(costs['expected_output_tokens']):,}, "
                f"requests={costs['expected_model_requests']:,.1f}"
            )
            print(
                f"    Reference-average cost: {_money(costs['expected_total_cost'])} "
                "(empirical planning estimate)"
            )
        print(
            f"    Estimated input cost: {_money(costs['input_cost'])}\n"
            f"    Capped output cost: {_money(costs['output_cost'])} (upper-bound component, not expected spend)\n"
            f"    Estimated total upper-bound cost: {_money(costs['total_cost'])}"
        )

    total_costs = _estimate_costs(
        estimate.n_refs,
        estimate.n_genes_classified,
        summary_pricing,
        classification_pricing,
        average_profile,
    )
    print(f"\n{'-'*70}")
    print("Total")
    print(
        f"Genes: total={estimate.total_genes}, reused={estimate.reused_genes}, "
        f"planned={estimate.planned_genes}"
    )
    print(
        f"References: candidates={estimate.candidate_references}, "
        f"after_reference_limit={estimate.limited_references}, "
        f"estimated_keyword_matches={_refs(estimate.metadata_keyword_matches)}"
    )
    print(f"N_refs={_refs(estimate.n_refs)}, N_genes_classified={estimate.n_genes_classified}")
    print(
        f"Tokens: input={int(total_costs['input_tokens']):,}, "
        f"capped_output={int(total_costs['output_tokens']):,}"
    )
    if "expected_total_cost" in total_costs:
        print(
            f"Reference-average tokens: input={int(total_costs['expected_input_tokens']):,}, "
            f"output={int(total_costs['expected_output_tokens']):,}, "
            f"requests={total_costs['expected_model_requests']:,.1f}"
        )
        print(f"Reference-average cost: {_money(total_costs['expected_total_cost'])} (empirical planning estimate)")
    print(f"Estimated input cost: {_money(total_costs['input_cost'])}")
    print(f"Capped output cost: {_money(total_costs['output_cost'])} (upper-bound component, not expected spend)")
    print(f"Estimated total upper-bound cost: {_money(total_costs['total_cost'])}")
    print(
        "\nNote: this fast soft-run estimate uses candidate-reference counts and an empirical "
        "species keyword-pass rate instead of resolving every title/abstract. It does not fetch "
        "full paper text, does not model full-text chunking, and does not model the "
        "evidence-quality stop condition that can end real processing after six high-quality "
        "references per gene."
    )
    print(f"{'='*70}")


def generate_excel_output(gene_hits, all_summaries, validated, output_path, gene_set_df=None, adapter: SpeciesAdapter | None = None):
    """Generate species-aware Excel output with Gene Set, Classification, and Reference Summaries sheets."""
    import pandas as pd
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    log_step("Building Excel output", str(output_path), indent=1)

    records: list[GeneRecord] = []
    for symbol, value in (validated or {}).items():
        if isinstance(value, GeneRecord):
            records.append(value)
        else:
            records.append(GeneRecord(
                species="fly",
                gene_id=str(value),
                symbol=str(symbol),
                authority_id=str(value),
                synonyms={str(symbol)},
            ))
    if adapter is None:
        species_key = records[0].species if records else "fly"
        adapter = get_species_adapter(species_key)

    pmcid_to_year = build_pmcid_to_year()
    
    def _sort_refs_by_recency(refs_set):
        return sorted(refs_set, key=lambda p: (-pmcid_to_year.get(str(p).upper(), 0), str(p)))

    def _trace(record: GeneRecord, key: str) -> str:
        return str((record.trace or {}).get(key, "") or "")
    
    log_step("Preparing classification sheet", f"{len(records)} genes", indent=2)
    # Classification sheet
    class_rows = []
    for record in records:
        info = gene_hits.get(record.symbol, {})
        supporting_refs = info.get("supporting_refs", set())
        full_text_refs = info.get("full_text_refs", set())
        queried_refs = info.get("pmcids", set())
        if adapter.key == "fly":
            row = {
                "Fly_Gene_symbol": record.symbol,
                "FlyBase_ID": record.gene_id,
            }
        else:
            row = {
                adapter.symbol_column: record.symbol,
                adapter.authority_id_column: record.authority_id,
                adapter.gene_id_column: record.gene_id,
            }
        row.update({
            "Category": info.get("category", "None"),
            "Confidence": info.get("confidence", 0),
            "Rationale": info.get("rationale", ""),
        })
        if adapter.key != "fly":
            row.update({
                "Fly_Gene_symbol_primary": _trace(record, "fly_gene_symbol_primary"),
                "Fly_Gene_symbol_input": _trace(record, "fly_gene_symbol_input"),
                "FlyBase_ID": _trace(record, "flybase_gene_id"),
                "DIOPT_Score": _trace(record, "diopt_score"),
                "DIOPT_Best_Match": _trace(record, "diopt_best_match"),
            })
        row.update({
            "Supporting_Refs": ";".join(_sort_refs_by_recency(supporting_refs)),
            "Papers_w_full_text": ";".join(_sort_refs_by_recency(full_text_refs)),
            "Queried_Refs": ";".join(_sort_refs_by_recency(queried_refs)),
        })
        class_rows.append(row)
    
    if adapter.key == "fly":
        class_cols = [
            "Fly_Gene_symbol", "FlyBase_ID", "Category", "Confidence", "Rationale",
            "Supporting_Refs", "Papers_w_full_text", "Queried_Refs"
        ]
    else:
        class_cols = [
            adapter.symbol_column, adapter.authority_id_column, adapter.gene_id_column,
            "Category", "Confidence", "Rationale",
            "Fly_Gene_symbol_primary", "Fly_Gene_symbol_input", "FlyBase_ID",
            "DIOPT_Score", "DIOPT_Best_Match",
            "Supporting_Refs", "Papers_w_full_text", "Queried_Refs",
        ]
    class_df = pd.DataFrame(class_rows, columns=class_cols)
    
    # Sort by category (None last), then by confidence descending
    if not class_df.empty:
        category_counts = class_df["Category"].value_counts().to_dict()
        sorted_categories = sorted(
            category_counts.keys(),
            key=lambda cat: (
                1 if cat == "None" else 0,
                -category_counts.get(cat, 0),
                cat
            )
        )
        category_order = {cat: idx for idx, cat in enumerate(sorted_categories)}
        class_df["_cat_order"] = class_df["Category"].map(
            lambda x: category_order.get(x, len(sorted_categories))
        )
        class_df = class_df.sort_values(
            by=["_cat_order", "Confidence"],
            ascending=[True, False]
        ).reset_index(drop=True)
        class_df = class_df.drop(columns=["_cat_order"])
    
    log_step("Preparing reference summaries sheet", f"{len(all_summaries or [])} summaries", indent=2)
    # Reference Summaries sheet
    ref_rows = []
    for sd in all_summaries or []:
        gene_sym = sd.get("gene_symbol", "")
        gene_id = sd.get("gene_id", sd.get("flybase_id", ""))
        
        if not sd.get("is_high_quality", False):
            continue
        
        cls_info = gene_hits.get(gene_sym, {})
        cls = cls_info.get("category", "None")
        
        if cls == "None" or cls == "Potential Hits":
            continue
        
        year_str = sd.get("year", "")
        try:
            year_int = int(year_str) if year_str and year_str != "n.d." else 0
        except:
            year_int = 0
        
        if adapter.key == "fly":
            row = {
                "Gene": gene_sym,
                "Gene_ID": gene_id,
            }
        else:
            trace = sd.get("trace", {}) or {}
            row = {
                adapter.symbol_column: gene_sym,
                adapter.authority_id_column: sd.get("authority_id", ""),
                adapter.gene_id_column: gene_id,
                "Fly_Gene_symbol_primary": trace.get("fly_gene_symbol_primary", ""),
            }
        row.update({
            "Paper_ID": sd.get("paper_id", ""),
            "Title": sd.get("title", ""),
            "Year": year_str,
            "_year_sort": year_int,
            "Journal": sd.get("journal", ""),
            "Author(s)": "; ".join(sd.get("authors", []) or []),
            "Reference Summary": sd.get("summary", ""),
            "Abstract": sd.get("abstract_text", ""),
            "Classification": cls,
            "Source": sd.get("source", "Unknown"),
        })
        ref_rows.append(row)
    
    if adapter.key == "fly":
        ref_cols = [
            "Gene", "Gene_ID", "Paper_ID", "Title", "Year", "_year_sort",
            "Journal", "Author(s)", "Reference Summary", "Abstract", "Classification", "Source",
        ]
    else:
        ref_cols = [
            adapter.symbol_column, adapter.authority_id_column, adapter.gene_id_column,
            "Fly_Gene_symbol_primary",
            "Paper_ID", "Title", "Year", "_year_sort",
            "Journal", "Author(s)", "Reference Summary", "Abstract", "Classification", "Source",
        ]
    ref_df = pd.DataFrame(ref_rows, columns=ref_cols)
    
    # Sort references to mirror classification order
    if not ref_df.empty and not class_df.empty:
        class_gene_col = "Fly_Gene_symbol" if adapter.key == "fly" else adapter.symbol_column
        ref_gene_col = "Gene" if adapter.key == "fly" else adapter.symbol_column
        gene_order = {gene: idx for idx, gene in enumerate(class_df[class_gene_col].tolist())}
        ref_df["_gene_order"] = ref_df[ref_gene_col].map(lambda x: gene_order.get(x, 999999))
        ref_df = ref_df.sort_values(by=["_gene_order", "_year_sort"], ascending=[True, False])
        ref_df = ref_df.drop(columns=["_gene_order"])
    
    if "_year_sort" in ref_df.columns:
        ref_df = ref_df.drop(columns=["_year_sort"])
    log_step("Writing Excel workbook", indent=2)
    # Write Excel
    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        gene_df = gene_set_df.copy() if isinstance(gene_set_df, pd.DataFrame) else pd.DataFrame()
        gene_df.to_excel(writer, index=False, sheet_name="Gene Set")
        class_df.to_excel(writer, index=False, sheet_name="Classification")
        ref_df.to_excel(writer, index=False, sheet_name="Reference Summaries")
        
        # Format Gene Set sheet
        gene_ws = writer.sheets["Gene Set"]
        for row in gene_ws.iter_rows(min_row=1, max_row=gene_ws.max_row, min_col=1, max_col=gene_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Format Classification sheet
        class_ws = writer.sheets["Classification"]
        class_ws.column_dimensions[get_column_letter(1)].width = 18
        class_ws.column_dimensions[get_column_letter(2)].width = 15
        class_ws.column_dimensions[get_column_letter(3)].width = 15
        class_ws.column_dimensions[get_column_letter(4)].width = 12
        class_ws.column_dimensions[get_column_letter(5)].width = 50
        class_ws.column_dimensions[get_column_letter(6)].width = 30
        class_ws.column_dimensions[get_column_letter(7)].width = 30
        class_ws.column_dimensions[get_column_letter(8)].width = 30
        class_ws.column_dimensions[get_column_letter(9)].width = 35
        for col_idx in range(10, class_ws.max_column + 1):
            class_ws.column_dimensions[get_column_letter(col_idx)].width = 30
        
        # Color palette for categories
        color_palette = [
            ("5B9BD5", True), ("FFC000", False), ("70AD47", False),
            ("ED7D31", False), ("9E480E", True), ("00B0F0", False),
            ("FF6B6B", False), ("4ECDC4", False), ("95E1D3", False),
            ("F38181", False), ("7030A0", True), ("FCBAD3", False),
        ]
        
        unique_categories = [cat for cat in class_df["Category"].unique() if cat != "None"]
        category_colors = {}
        category_needs_white_text = {}
        for idx, category in enumerate(unique_categories):
            color_hex, needs_white = color_palette[idx % len(color_palette)]
            category_colors[category] = color_hex
            category_needs_white_text[category] = needs_white
        
        # Apply formatting
        for row in class_ws.iter_rows(min_row=1, max_row=class_ws.max_row, min_col=1, max_col=class_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Apply category colors
        for row_idx in range(2, class_ws.max_row + 1):
            category_col = class_cols.index("Category") + 1
            category_cell = class_ws.cell(row=row_idx, column=category_col)
            category_value = category_cell.value
            if category_value and category_value != "None" and category_value in category_colors:
                color_hex = category_colors[category_value]
                category_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                if category_needs_white_text.get(category_value, False):
                    category_cell.font = Font(color="FFFFFF", bold=True)
        
        # Format Reference Summaries sheet
        ref_ws = writer.sheets["Reference Summaries"]
        ref_ws.column_dimensions[get_column_letter(1)].width = 15
        ref_ws.column_dimensions[get_column_letter(2)].width = 15
        ref_ws.column_dimensions[get_column_letter(3)].width = 12
        ref_ws.column_dimensions[get_column_letter(4)].width = 50
        ref_ws.column_dimensions[get_column_letter(5)].width = 8
        ref_ws.column_dimensions[get_column_letter(6)].width = 25
        ref_ws.column_dimensions[get_column_letter(7)].width = 30
        ref_ws.column_dimensions[get_column_letter(8)].width = 50
        ref_ws.column_dimensions[get_column_letter(9)].width = 50
        ref_ws.column_dimensions[get_column_letter(10)].width = 15
        ref_ws.column_dimensions[get_column_letter(11)].width = 12
        ref_ws.column_dimensions[get_column_letter(12)].width = 35
        
        for row in ref_ws.iter_rows(min_row=1, max_row=ref_ws.max_row, min_col=1, max_col=ref_ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Write to file
    with open(output_path, 'wb') as f:
        f.write(excel_buf.getvalue())
    
    print(f"\nOutput written to: {output_path}")
    print(f"  - Classification sheet: {len(class_df)} genes")
    print(f"  - Reference Summaries sheet: {len(ref_df)} entries")


def _json_dump(path: Path, payload: dict):
    """Atomically write JSON payload to disk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2, sort_keys=True)
    os.replace(tmp_path, path)


def _json_load(path: Path, default: Any):
    """Load JSON payload, returning default on missing/invalid content."""
    if not path.exists():
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def _normalize_keywords_for_key(keywords_list: list[str]) -> str:
    cleaned = sorted({str(k).strip().lower() for k in (keywords_list or []) if str(k).strip()})
    return ",".join(cleaned)


def _make_gene_store_key(gene_id: str, keywords_list: list[str], reference_limit: int, species: str = "fly") -> str:
    kw_norm = _normalize_keywords_for_key(keywords_list)
    kw_hash = hashlib.sha256(kw_norm.encode("utf-8")).hexdigest()[:16]
    return f"v{CACHE_SCHEMA_VERSION}|{str(species).strip()}|{str(gene_id).strip()}|{reference_limit}|{kw_hash}"


def _make_input_fingerprint(gene_ids: list[str], keywords_list: list[str], reference_limit: int, species: str = "fly", diopt_filter: str = "") -> str:
    normalized_ids = sorted({str(x).strip() for x in gene_ids if str(x).strip()})
    payload = {
        "cache_schema_version": CACHE_SCHEMA_VERSION,
        "species": str(species or "fly"),
        "gene_ids": normalized_ids,
        "keywords": _normalize_keywords_for_key(keywords_list),
        "reference_limit": int(reference_limit),
        "diopt_filter": str(diopt_filter or ""),
    }
    src = json.dumps(payload, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(src.encode("utf-8")).hexdigest()


def _get_batch_root(input_directory: str) -> Path:
    return Path(input_directory) / BATCH_STATE_DIRNAME


def _get_csv_state_dir(input_directory: str, csv_path: str, species: str = "fly") -> Path:
    csv_abs = str(Path(csv_path).resolve())
    csv_hash = hashlib.sha1(f"{species}|{csv_abs}".encode("utf-8")).hexdigest()[:10]
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(csv_path).stem)
    return _get_batch_root(input_directory) / f"{species}_{safe_stem}_{csv_hash}"


def _serialize_hit_info(hit_info: dict) -> dict:
    out = dict(hit_info or {})
    for set_field in ("pmcids", "supporting_refs", "full_text_refs"):
        val = out.get(set_field, set())
        if isinstance(val, set):
            out[set_field] = sorted(list(val))
        elif isinstance(val, list):
            out[set_field] = sorted(list({str(x) for x in val if str(x)}))
        else:
            out[set_field] = []
    out["confidence"] = int(out.get("confidence", 0) or 0)
    out["category"] = str(out.get("category", "None") or "None")
    out["rationale"] = str(out.get("rationale", "") or "")
    out["classified_by"] = str(out.get("classified_by", "Self") or "Self")
    return out


def _deserialize_hit_info(hit_info: dict) -> dict:
    out = dict(hit_info or {})
    for set_field in ("pmcids", "supporting_refs", "full_text_refs"):
        val = out.get(set_field, [])
        if isinstance(val, set):
            out[set_field] = set(val)
        else:
            out[set_field] = set([str(x) for x in (val or []) if str(x)])
    out["confidence"] = int(out.get("confidence", 0) or 0)
    out["category"] = str(out.get("category", "None") or "None")
    out["rationale"] = str(out.get("rationale", "") or "")
    out["classified_by"] = str(out.get("classified_by", "Self") or "Self")
    return out


def _serialize_gene_record(gene: GeneRecord) -> dict:
    return {
        "species": gene.species,
        "gene_id": gene.gene_id,
        "symbol": gene.symbol,
        "authority_id": gene.authority_id,
        "synonyms": sorted(list(gene.synonyms or [])),
        "trace": dict(gene.trace or {}),
    }


def _deserialize_gene_record(payload: dict) -> GeneRecord:
    if not isinstance(payload, dict):
        return GeneRecord(species="fly", gene_id="", symbol="")
    return GeneRecord(
        species=str(payload.get("species", "fly") or "fly"),
        gene_id=str(payload.get("gene_id", payload.get("fbgn_id", "")) or ""),
        symbol=str(payload.get("symbol", payload.get("gene_symbol", "")) or ""),
        authority_id=str(payload.get("authority_id", payload.get("fbgn_id", "")) or ""),
        synonyms=set(str(x) for x in (payload.get("synonyms", []) or []) if str(x)),
        trace=dict(payload.get("trace", {}) or {}),
    )


def _build_gene_records(
    gene_hits: dict,
    all_summaries: list[dict],
    validated: dict[str, GeneRecord],
    openai_usage_by_gene_id: Optional[dict[str, list[dict[str, Any]]]] = None,
) -> list[dict]:
    summaries_by_gene_id: dict[str, list[dict]] = defaultdict(list)
    for summary in all_summaries or []:
        gene_id = str(summary.get("gene_id", summary.get("flybase_id", ""))).strip()
        if gene_id:
            summaries_by_gene_id[gene_id].append(summary)

    records: list[dict] = []
    for gene_symbol, gene in (validated or {}).items():
        if not isinstance(gene, GeneRecord):
            gene = GeneRecord(species="fly", gene_id=str(gene), symbol=str(gene_symbol), authority_id=str(gene))
        info = gene_hits.get(gene_symbol, {})
        openai_usage_events = list((openai_usage_by_gene_id or {}).get(str(gene.gene_id), []) or [])
        records.append({
            "gene_record": _serialize_gene_record(gene),
            # Legacy fields for older run stores and easier inspection.
            "fbgn_id": str(gene.gene_id) if gene.species == "fly" else str(gene.trace.get("flybase_gene_id", "")),
            "gene_symbol": str(gene_symbol),
            "hit_info": _serialize_hit_info(info),
            "summaries": summaries_by_gene_id.get(str(gene.gene_id), []),
            "openai_usage": {
                "totals": _summarize_openai_usage_events(openai_usage_events),
                "requests": openai_usage_events,
            },
        })
    return records


def _merge_gene_records(records: list[dict]) -> tuple[dict, list[dict], dict[str, GeneRecord]]:
    gene_hits: dict[str, dict] = {}
    all_summaries: list[dict] = []
    validated: dict[str, GeneRecord] = {}
    for record in records or []:
        gene_record_payload = record.get("gene_record")
        if isinstance(gene_record_payload, dict):
            gene = _deserialize_gene_record(gene_record_payload)
        else:
            gene = GeneRecord(
                species="fly",
                gene_id=str(record.get("fbgn_id", "")).strip(),
                symbol=str(record.get("gene_symbol", "")).strip(),
                authority_id=str(record.get("fbgn_id", "")).strip(),
            )
        if not gene.symbol or not gene.gene_id:
            continue
        validated[gene.symbol] = gene
        gene_hits[gene.symbol] = _deserialize_hit_info(record.get("hit_info", {}))
        all_summaries.extend(record.get("summaries", []) or [])
    return gene_hits, all_summaries, validated


def _load_run_store(input_directory: str) -> dict:
    run_store_path = _get_batch_root(input_directory) / RUN_STORE_FILENAME
    store = _json_load(run_store_path, default={"version": CACHE_SCHEMA_VERSION, "genes": {}})
    if not isinstance(store, dict):
        return {"version": CACHE_SCHEMA_VERSION, "genes": {}}
    if int(store.get("version", 0) or 0) != CACHE_SCHEMA_VERSION:
        return {"version": CACHE_SCHEMA_VERSION, "genes": {}}
    if "genes" not in store or not isinstance(store["genes"], dict):
        store["genes"] = {}
    store["version"] = CACHE_SCHEMA_VERSION
    return store


def _save_run_store(input_directory: str, run_store: dict):
    run_store_path = _get_batch_root(input_directory) / RUN_STORE_FILENAME
    _json_dump(run_store_path, run_store)


def _input_id_column_for_adapter(adapter: SpeciesAdapter) -> str:
    if adapter.key == "fly":
        return "flybase_gene_id"
    return f"{adapter.key}_entrez_gene_id"


def _input_symbol_column_for_adapter(adapter: SpeciesAdapter) -> str:
    if adapter.key == "fly":
        return "ext_gene"
    return f"{adapter.key}_gene_symbol"


def _normalize_gene_id_value(value: object, adapter: SpeciesAdapter) -> str:
    """Normalize IDs read from CSVs where blank rows can coerce Entrez IDs to floats."""
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    if adapter.key != "fly":
        match = re.fullmatch(r"(\d+)\.0+", text)
        if match:
            return match.group(1)
    return text


def _gene_records_from_df(df: pd.DataFrame, adapter: SpeciesAdapter, input_gene_col: str = "ext_gene") -> list[GeneRecord]:
    catalog = adapter.load_gene_catalog()
    id_col = _input_id_column_for_adapter(adapter)
    if id_col not in df.columns:
        return []
    symbol_col = input_gene_col if adapter.key == "fly" else _input_symbol_column_for_adapter(adapter)
    authority_col = "flybase_gene_id" if adapter.key == "fly" else ("hgnc_id" if adapter.key == "human" else "mgi_id")
    def _merge_trace_value(existing: str, incoming: str) -> str:
        values = []
        seen = set()
        for raw in (existing, incoming):
            for part in str(raw or "").split(";"):
                cleaned = part.strip()
                if cleaned and cleaned not in seen:
                    seen.add(cleaned)
                    values.append(cleaned)
        return ";".join(values)

    records_by_id: dict[str, GeneRecord] = {}
    for _, row in df.iterrows():
        gene_id = _normalize_gene_id_value(row.get(id_col, ""), adapter)
        status = str(row.get("status", "mapped") or "mapped").strip().lower()
        if not gene_id or status == "unmapped":
            continue
        catalog_record = catalog.get(gene_id)
        if catalog_record:
            record = GeneRecord(
                species=catalog_record.species,
                gene_id=catalog_record.gene_id,
                symbol=catalog_record.symbol,
                authority_id=catalog_record.authority_id,
                synonyms=set(catalog_record.synonyms),
                trace={},
            )
        else:
            symbol = str(row.get(symbol_col, "") or gene_id).strip()
            authority_id = _normalize_gene_id_value(row.get(authority_col, ""), adapter) or gene_id
            if adapter.key != "fly":
                print(
                    f"  [Warning] {adapter.display_name} gene ID {gene_id} "
                    f"not found in catalog; using CSV symbol '{symbol or gene_id}'"
                )
            record = GeneRecord(
                species=adapter.key,
                gene_id=gene_id,
                symbol=symbol,
                authority_id=authority_id,
                synonyms={symbol} if symbol else set(),
            )
        if adapter.key != "fly":
            incoming_trace = {
                "fly_gene_symbol_input": str(row.get("fly_gene_symbol_input", "") or ""),
                "fly_gene_symbol_primary": str(row.get("fly_gene_symbol_primary", "") or ""),
                "flybase_gene_id": str(row.get("flybase_gene_id", "") or ""),
                "diopt_score": str(row.get("diopt_score", "") or ""),
                "diopt_best_match": str(row.get("diopt_best_match", "") or ""),
                "diopt_supporting_algorithms": str(row.get("diopt_supporting_algorithms", "") or ""),
            }
            record.trace = incoming_trace
        if record.gene_id in records_by_id:
            existing = records_by_id[record.gene_id]
            existing.synonyms.update(record.synonyms)
            for key, value in (record.trace or {}).items():
                existing.trace[key] = _merge_trace_value(existing.trace.get(key, ""), value)
        else:
            records_by_id[record.gene_id] = record
    return list(records_by_id.values())


def process_csv_file(
    csv_path: str,
    keywords_list: list[str],
    reference_limit: int,
    input_directory: str,
    run_store: dict,
    adapter: SpeciesAdapter | None = None,
    diopt_filter: str = "",
    input_gene_col: str = "ext_gene",
    force_all: bool = False,
):
    """Process a single CSV file containing species-specific gene IDs."""
    adapter = adapter or get_species_adapter("fly")
    print(f"\n{'#'*70}")
    print(f"Processing file: {csv_path}")
    print(f"{'#'*70}")
    
    # Read CSV
    log_step("Reading input CSV", str(csv_path), indent=1)
    try:
        df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return False
    
    id_col = _input_id_column_for_adapter(adapter)
    if id_col not in df.columns:
        print(f"Skipping: No '{id_col}' column found")
        return False

    log_step("Resolving input rows to gene records", f"id column={id_col}", indent=1)
    gene_records = _gene_records_from_df(df, adapter, input_gene_col=input_gene_col)
    if not gene_records:
        print(f"Skipping: No valid {adapter.display_name} gene IDs found")
        return False

    gene_ids = [gene.gene_id for gene in gene_records]
    print(f"Found {len(gene_ids)} unique {adapter.display_name} gene IDs")

    input_fingerprint = _make_input_fingerprint(
        gene_ids, keywords_list, reference_limit, species=adapter.key, diopt_filter=diopt_filter
    )
    state_dir = _get_csv_state_dir(input_directory, csv_path, species=adapter.key)
    state_path = state_dir / "state.json"
    state_dir.mkdir(parents=True, exist_ok=True)
    log_step("Preparing checkpoint state", str(state_dir), indent=1)

    state_default = {
        "version": CACHE_SCHEMA_VERSION,
        "csv_path": str(Path(csv_path).resolve()),
        "input_fingerprint": input_fingerprint,
        "total_batches": 0,
        "completed_batches": [],
        "batch_files": {},
        "processed_genes": 0,
        "reused_genes": 0,
    }
    state = _json_load(state_path, default=state_default)
    if not isinstance(state, dict):
        state = dict(state_default)

    if force_all:
        print("Force-all mode enabled: reprocessing all batches for this file.")
        state = dict(state_default)
    elif state.get("input_fingerprint") != input_fingerprint:
        print("Input fingerprint changed for this file; restarting file-specific batch state.")
        state = dict(state_default)

    batches = [gene_records[i:i + BATCH_SIZE] for i in range(0, len(gene_records), BATCH_SIZE)]
    state["total_batches"] = len(batches)
    state["csv_path"] = str(Path(csv_path).resolve())
    state["input_fingerprint"] = input_fingerprint
    completed_batches = set(int(x) for x in (state.get("completed_batches", []) or []))
    log_step("Preparing batches", f"{len(batches)} batch(es), size={BATCH_SIZE}", indent=1)

    if completed_batches and not force_all:
        print(f"Resuming: {len(completed_batches)}/{len(batches)} batches already complete.")

    calibration_totals = {
        "limited_references": 0,
        "metadata_keyword_matches": 0,
    }
    for batch_index, batch_gene_records in enumerate(batches):
        if (batch_index in completed_batches) and not force_all:
            print(f"  Skipping completed batch {batch_index + 1}/{len(batches)}")
            continue

        print(f"\n[Batch {batch_index + 1}/{len(batches)}] size={len(batch_gene_records)}")
        try:
            dedup_hits = 0
            to_process: list[GeneRecord] = []
            reused_records: list[dict] = []
            for gene in batch_gene_records:
                store_key = _make_gene_store_key(
                    gene.gene_id, keywords_list, reference_limit, species=adapter.key
                )
                if (not force_all) and store_key in run_store.get("genes", {}):
                    reused_records.append(run_store["genes"][store_key])
                    dedup_hits += 1
                else:
                    to_process.append(gene)
            print(f"  Cache status: {dedup_hits} reused, {len(to_process)} to process")

            processed_records: list[dict] = []
            if to_process:
                log_step("Running gene classification batch", f"{len(to_process)} uncached gene(s)", indent=1)
                gene_hits, all_summaries, validated, openai_usage_by_gene_id, calibration_stats = process_gene_set(
                    to_process, keywords_list, reference_limit, adapter=adapter
                )
                calibration_totals["limited_references"] += int(calibration_stats.get("limited_references", 0) or 0)
                calibration_totals["metadata_keyword_matches"] += int(
                    calibration_stats.get("metadata_keyword_matches", 0) or 0
                )
                processed_records = _build_gene_records(
                    gene_hits,
                    all_summaries,
                    validated,
                    openai_usage_by_gene_id=openai_usage_by_gene_id,
                )

                for record in processed_records:
                    gene_record = _deserialize_gene_record(record.get("gene_record", {}))
                    store_key = _make_gene_store_key(
                        gene_record.gene_id,
                        keywords_list,
                        reference_limit,
                        species=adapter.key,
                    )
                    run_store.setdefault("genes", {})[store_key] = record
                _save_run_store(input_directory, run_store)
                log_step("Saved run-store cache", f"{len(run_store.get('genes', {}))} cached gene entries", indent=1)

            record_by_gene_id: dict[str, dict] = {}
            for rec in reused_records + processed_records:
                gene_record = _deserialize_gene_record(rec.get("gene_record", {}))
                if gene_record.gene_id:
                    record_by_gene_id[gene_record.gene_id] = rec
            batch_records = [
                record_by_gene_id[gene.gene_id]
                for gene in batch_gene_records
                if gene.gene_id in record_by_gene_id
            ]

            batch_filename = f"batch_{batch_index:05d}.json"
            batch_path = state_dir / batch_filename
            log_step("Writing batch checkpoint", batch_filename, indent=1)
            batch_payload = {
                "version": CACHE_SCHEMA_VERSION,
                "batch_index": batch_index,
                "gene_ids": [gene.gene_id for gene in batch_gene_records],
                "species": adapter.key,
                "records": batch_records,
                "processed_count": len(processed_records),
                "reused_count": dedup_hits,
            }
            _json_dump(batch_path, batch_payload)

            completed_batches.add(batch_index)
            state["completed_batches"] = sorted(list(completed_batches))
            state.setdefault("batch_files", {})[str(batch_index)] = batch_filename
            state["processed_genes"] = int(state.get("processed_genes", 0) or 0) + len(processed_records)
            state["reused_genes"] = int(state.get("reused_genes", 0) or 0) + dedup_hits
            _json_dump(state_path, state)

            print(
                f"  Batch complete: processed={len(processed_records)} "
                f"reused={dedup_hits}"
            )
        except Exception as e:
            print(f"Error in batch {batch_index + 1}: {e}")
            return False

    log_step("Aggregating batch checkpoints", f"{len(batches)} batch artifact(s)", indent=1)
    # Aggregate all completed batch artifacts for this file.
    all_records: list[dict] = []
    for batch_index in range(len(batches)):
        batch_filename = state.get("batch_files", {}).get(str(batch_index), f"batch_{batch_index:05d}.json")
        batch_path = state_dir / batch_filename
        if not batch_path.exists():
            print(f"Error: Missing expected batch artifact {batch_path}")
            return False
        batch_payload = _json_load(batch_path, default={})
        if not isinstance(batch_payload, dict):
            print(f"Error: Invalid batch artifact format: {batch_path}")
            return False
        all_records.extend(batch_payload.get("records", []) or [])

    gene_hits, all_summaries, validated = _merge_gene_records(all_records)
    if not validated:
        print("No genes were successfully validated")
        return False

    # Generate one final output Excel per input CSV.
    base_name = os.path.splitext(csv_path)[0]
    output_path = f"{base_name}_classification.xlsx"
    log_step("Generating final workbook", output_path, indent=1)
    
    generate_excel_output(gene_hits, all_summaries, validated, output_path, gene_set_df=df, adapter=adapter)

    if calibration_totals["limited_references"] > 0:
        updated = _maybe_update_species_keyword_pass_rate(
            adapter.key,
            calibration_totals["limited_references"],
            calibration_totals["metadata_keyword_matches"],
            source=str(Path(csv_path).resolve()),
            reference_limit=reference_limit,
        )
        if updated:
            log_step(
                "Updated fast soft-run keyword-pass calibration",
                (
                    f"species={adapter.key}, "
                    f"limited_refs={calibration_totals['limited_references']}, "
                    f"matches={calibration_totals['metadata_keyword_matches']}"
                ),
                indent=1,
            )
    
    return True


def soft_run_csv_file(
    csv_path: str,
    keywords_list: list[str],
    reference_limit: int,
    input_directory: str,
    run_store: dict,
    adapter: SpeciesAdapter | None = None,
    diopt_filter: str = "",
    input_gene_col: str = "ext_gene",
    force_all: bool = False,
) -> Optional[SoftRunFileEstimate]:
    """Estimate OpenAI-bound work for a CSV without OpenAI or full-text calls."""
    adapter = adapter or get_species_adapter("fly")
    print(f"\n{'#'*70}")
    print(f"Soft-run estimating file: {csv_path}")
    print(f"{'#'*70}")

    log_step("Reading input CSV", str(csv_path), indent=1)
    try:
        df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return None

    id_col = _input_id_column_for_adapter(adapter)
    if id_col not in df.columns:
        print(f"Skipping: No '{id_col}' column found")
        return None

    log_step("Resolving input rows to gene records", f"id column={id_col}", indent=1)
    gene_records = _gene_records_from_df(df, adapter, input_gene_col=input_gene_col)
    if not gene_records:
        print(f"Skipping: No valid {adapter.display_name} gene IDs found")
        return None

    gene_ids = [gene.gene_id for gene in gene_records]
    print(f"Found {len(gene_ids)} unique {adapter.display_name} gene IDs")

    input_fingerprint = _make_input_fingerprint(
        gene_ids, keywords_list, reference_limit, species=adapter.key, diopt_filter=diopt_filter
    )
    state_dir = _get_csv_state_dir(input_directory, csv_path, species=adapter.key)
    state_path = state_dir / "state.json"
    state_default = {
        "version": CACHE_SCHEMA_VERSION,
        "csv_path": str(Path(csv_path).resolve()),
        "input_fingerprint": input_fingerprint,
        "total_batches": 0,
        "completed_batches": [],
        "batch_files": {},
        "processed_genes": 0,
        "reused_genes": 0,
    }
    state = _json_load(state_path, default=state_default)
    if not isinstance(state, dict):
        state = dict(state_default)
    if force_all or state.get("input_fingerprint") != input_fingerprint:
        completed_batches: set[int] = set()
    else:
        completed_batches = set(int(x) for x in (state.get("completed_batches", []) or []))

    batches = [gene_records[i:i + BATCH_SIZE] for i in range(0, len(gene_records), BATCH_SIZE)]
    estimate = SoftRunFileEstimate(csv_path=csv_path, total_genes=len(gene_records))
    pass_rate_info = _resolve_species_keyword_pass_rate(adapter.key)
    estimate.keyword_pass_rate = _to_float(pass_rate_info.get("keyword_pass_rate"), default=1.0)
    estimate.keyword_pass_rate_source = str(pass_rate_info.get("source", "") or "fallback")
    estimate.keyword_pass_rate_limited_references = int(_to_float(pass_rate_info.get("calibration_limited_references")))
    estimate.keyword_pass_rate_metadata_keyword_matches = int(
        _to_float(pass_rate_info.get("calibration_metadata_keyword_matches"))
    )
    if completed_batches and not force_all:
        print(f"Soft-run cache: {len(completed_batches)}/{len(batches)} completed batch(es) excluded from new cost")

    for batch_index, batch_gene_records in enumerate(batches):
        if (batch_index in completed_batches) and not force_all:
            estimate.reused_genes += len(batch_gene_records)
            continue

        to_process: list[GeneRecord] = []
        dedup_hits = 0
        for gene in batch_gene_records:
            store_key = _make_gene_store_key(
                gene.gene_id, keywords_list, reference_limit, species=adapter.key
            )
            if (not force_all) and store_key in run_store.get("genes", {}):
                dedup_hits += 1
            else:
                to_process.append(gene)
        estimate.reused_genes += dedup_hits
        estimate.planned_genes += len(to_process)
        print(
            f"\n[Soft-run Batch {batch_index + 1}/{len(batches)}] "
            f"reused={dedup_hits}, planned={len(to_process)}"
        )
        if not to_process:
            continue
        batch_estimate = _plan_soft_run_gene_set(
            to_process,
            keywords_list,
            reference_limit=reference_limit,
            adapter=adapter,
        )
        estimate.candidate_references += batch_estimate.candidate_references
        estimate.limited_references += batch_estimate.limited_references
        estimate.metadata_keyword_matches += batch_estimate.metadata_keyword_matches
        estimate.n_refs += batch_estimate.n_refs
        estimate.n_genes_classified += batch_estimate.n_genes_classified
        estimate.keyword_pass_rate = batch_estimate.keyword_pass_rate
        estimate.keyword_pass_rate_source = batch_estimate.keyword_pass_rate_source
        estimate.keyword_pass_rate_limited_references = batch_estimate.keyword_pass_rate_limited_references
        estimate.keyword_pass_rate_metadata_keyword_matches = batch_estimate.keyword_pass_rate_metadata_keyword_matches

    return estimate


def run_fbgnid_conversion(
    input_directory: str,
    input_gene_col: str = "ext_gene",
    flybase_data_dir: Path | None = None,
) -> bool:
    """
    Run the same FBgn conversion helper used in fly_stocker_v2-style workflows.
    """
    if not GET_FBGN_IDS_SCRIPT.exists():
        print(f"Error: FBgnID conversion script not found at: {GET_FBGN_IDS_SCRIPT}")
        return False

    cmd = [sys.executable, str(GET_FBGN_IDS_SCRIPT), str(input_directory), str(input_gene_col)]
    if flybase_data_dir is not None:
        cmd.extend(["--flybase-data-dir", str(flybase_data_dir)])
    print(f"\nRunning FBgnID conversion: {' '.join(cmd)}")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=False)
        if result.stdout:
            print(result.stdout)
        if result.stderr:
            print(result.stderr)
        if result.returncode != 0:
            print(f"FBgnID conversion failed with exit code {result.returncode}")
            return False
        return True
    except Exception as e:
        print(f"Error running FBgnID conversion: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Fly gene classification pipeline - converts input symbols "
            "to FBgn IDs, summarizes literature, and classifies genes"
        )
    )
    parser.add_argument(
        "input_directory",
        help="Directory containing input CSV files for FBgn conversion and classification"
    )
    parser.add_argument(
        "--keywords", "-k",
        default="",
        help="Comma-separated list of classification keywords (e.g., 'circadian,sleep,rhythm')"
    )
    parser.add_argument(
        "--reference-limit", "-r",
        type=int,
        default=500,
        help="Maximum number of references to consider per gene (default: 500)"
    )
    parser.add_argument(
        "--input-gene-col",
        default="ext_gene",
        help="Input gene symbol column for FBgnID conversion (default: ext_gene)."
    )
    parser.add_argument(
        "--flybase-data-dir",
        default="",
        help="Optional FlyBase data directory. Defaults to shared lab data, or local Data/FlyBase if unavailable."
    )
    parser.add_argument(
        "--force-all",
        action="store_true",
        help="Ignore checkpoint/resume state and reprocess all genes from scratch."
    )
    parser.add_argument(
        "--soft-run",
        action="store_true",
        help=(
            "Estimate OpenAI API usage and cost from metadata only. "
            "Does not call OpenAI, fetch full paper text, or write classification outputs."
        )
    )
    parser.add_argument(
        "--orthologs",
        choices=["none", "human", "mouse"],
        default="none",
        help="Optionally map fly genes to human or mouse orthologs before classification."
    )
    parser.add_argument(
        "--diopt-filter",
        default="exclude_low_score_2",
        help="DIOPT filter to use for ortholog lookup (default: exclude_low_score_2)."
    )
    parser.add_argument(
        "--human-data-dir",
        default="",
        help="Cache directory for human HGNC/gene2pubmed/GeneRIF/UniProt data."
    )
    parser.add_argument(
        "--mouse-data-dir",
        default="",
        help="Cache directory for mouse MGI/gene2pubmed/GeneRIF/UniProt data."
    )
    parser.add_argument(
        "--refresh-human-data",
        action="store_true",
        help="Re-download cached human reference data before processing."
    )
    parser.add_argument(
        "--refresh-mouse-data",
        action="store_true",
        help="Re-download cached mouse reference data before processing."
    )
    parser.add_argument(
        "--diopt-cache-dir",
        default="",
        help="Optional cache directory for DIOPT API responses."
    )
    parser.add_argument(
        "--diopt-workers",
        type=int,
        default=8,
        help="Concurrent DIOPT lookup workers for ortholog mapping (default: 8)."
    )
    
    args = parser.parse_args()

    global FLYBASE_DATA
    FLYBASE_DATA = resolve_flybase_data_dir(args.flybase_data_dir or None)
    
    # Validate input directory
    log_step("Validating input directory", args.input_directory)
    if not os.path.isdir(args.input_directory):
        print(f"Error: '{args.input_directory}' is not a valid directory")
        sys.exit(1)
    
    if using_local_data_root() and not args.flybase_data_dir and not os.environ.get("FLYBASE_DATA_DIR"):
        print(f"Shared lab data root is not accessible: {SHARED_DATA_ROOT}")
        print(f"Using local data fallback: {FLYBASE_DATA}")

    log_step("Checking FlyBase data files", str(FLYBASE_DATA))
    missing_flybase = ensure_flybase_data_files(FLYBASE_DATA, download_missing=True)
    if missing_flybase:
        print(f"Error: FlyBase data files are still missing under: {FLYBASE_DATA}")
        print("Expected structure:")
        print(f"  {FLYBASE_DATA}/Genes/fb_synonym_fb_*.tsv[.gz]")
        print(f"  {FLYBASE_DATA}/FlyBase_References/entity_publication_fb_*.tsv[.gz]")
        print(f"  {FLYBASE_DATA}/FlyBase_References/fbrf_pmid_pmcid_doi_fb_*.tsv[.gz]")
        print("Could not retrieve:")
        for requirement in missing_flybase:
            print(f"  - {requirement['label']}: {requirement['url']}")
        sys.exit(1)
    
    # Parse keywords
    keywords_list = [k.strip() for k in args.keywords.split(",") if k.strip()]
    
    print(f"\n{'='*70}")
    print("Fly Gene Classification Pipeline (Standalone)")
    print(f"{'='*70}")
    print(f"Input directory: {args.input_directory}")
    print(f"Keywords: {keywords_list if keywords_list else '(none)'}")
    print(f"Reference limit: {args.reference_limit}")
    print(f"Batch size: {BATCH_SIZE} (automatic)")
    print(f"FlyBase data: {FLYBASE_DATA}")
    print(f"PubMed cache: {PUBMED_CACHE_DIR}")
    print(f"Force all: {'yes' if args.force_all else 'no'}")
    print(f"Soft run: {'yes' if args.soft_run else 'no'}")
    print(f"Ortholog mode: {args.orthologs}")
    print(f"DIOPT filter: {args.diopt_filter}")
    if args.orthologs != "none":
        print(f"DIOPT workers: {max(1, args.diopt_workers)}")
    print("Run FBgnID conversion: yes (default)")
    print(f"FBgn conversion input gene column: {args.input_gene_col}")

    # Fail fast on API misconfiguration so runs do not silently produce empty outputs.
    summary_model = _get_summary_model()
    classification_model = _get_classification_model()
    soft_run_estimate: Optional[SoftRunEstimate] = None
    if args.soft_run:
        log_step("Refreshing OpenAI pricing table for soft-run")
        try:
            pricing_table = _load_openai_pricing_table(refresh=True)
            summary_pricing = _resolve_model_pricing(summary_model, pricing_table)
            classification_pricing = _resolve_model_pricing(classification_model, pricing_table)
        except Exception as e:
            print(f"Error: could not prepare OpenAI pricing table: {e}")
            sys.exit(1)
        soft_run_estimate = SoftRunEstimate(
            summary_model=summary_model,
            classification_model=classification_model,
            summary_pricing=summary_pricing,
            classification_pricing=classification_pricing,
        )
        print(f"OpenAI summary model: {summary_model}")
        print(f"OpenAI classification model: {classification_model}")
    else:
        log_step("Validating OpenAI configuration")
        try:
            _get_openai_client()
        except Exception as e:
            print(f"Error: OpenAI API setup invalid: {e}")
            sys.exit(1)
        print(f"OpenAI summary model: {summary_model}")
        print(f"OpenAI classification model: {classification_model}")
        print(
            "OpenAI reasoning effort: "
            f"summary={_get_reasoning_effort('OPENAI_SUMMARY_REASONING_EFFORT', DEFAULT_SUMMARY_REASONING_EFFORT)}, "
            f"classification={_get_reasoning_effort('OPENAI_CLASSIFICATION_REASONING_EFFORT', DEFAULT_CLASSIFICATION_REASONING_EFFORT)}"
        )

    # Step 0: convert input symbols to FBgn IDs.
    log_step("Converting input gene symbols to FBgn IDs")
    ok = run_fbgnid_conversion(args.input_directory, args.input_gene_col, FLYBASE_DATA)
    if not ok:
        print("Error: FBgnID conversion failed; aborting.")
        sys.exit(1)

    log_step("Configuring species data caches")
    human_gene_data.configure(args.human_data_dir or None, refresh=args.refresh_human_data)
    mouse_gene_data.configure(args.mouse_data_dir or None, refresh=args.refresh_mouse_data)

    adapter = get_species_adapter("fly" if args.orthologs == "none" else args.orthologs)

    if args.orthologs != "none":
        log_step("Mapping fly genes to orthologs", args.orthologs)
        try:
            generated = convert_ortholog_directory(
                args.input_directory,
                orthologs=args.orthologs,
                input_gene_col=args.input_gene_col,
                diopt_filter=args.diopt_filter,
                flybase_data_dir=FLYBASE_DATA,
                cache_dir=args.diopt_cache_dir or None,
                reuse_existing=not args.force_all,
                diopt_workers=args.diopt_workers,
            )
        except Exception as e:
            print(f"Error creating {args.orthologs} ortholog CSVs: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
        csv_files = [str(path) for path in generated]
    else:
        log_step("Finding source CSV files")
        # Find source CSV files, excluding generated ortholog and output files.
        csv_files = [
            os.path.join(args.input_directory, f)
            for f in os.listdir(args.input_directory)
            if (
                f.endswith('.csv')
                and not f.endswith('_classification.csv')
                and not f.endswith('_human.csv')
                and not f.endswith('_mouse.csv')
            )
        ]
    
    if not csv_files:
        print(f"\nNo CSV files found in {args.input_directory}")
        sys.exit(1)
    
    print(f"\nFound {len(csv_files)} CSV files to process")
    batch_root = _get_batch_root(args.input_directory)
    if args.soft_run:
        log_step("Inspecting batch state directory", str(batch_root))
    else:
        log_step("Preparing batch state directory", str(batch_root))
        batch_root.mkdir(parents=True, exist_ok=True)
    log_step("Loading run-store cache")
    run_store = _load_run_store(args.input_directory)
    
    # Process each file
    success_count = 0
    for csv_path in sorted(csv_files):
        try:
            if args.soft_run:
                file_estimate = soft_run_csv_file(
                    csv_path,
                    keywords_list,
                    args.reference_limit,
                    args.input_directory,
                    run_store,
                    adapter=adapter,
                    diopt_filter=args.diopt_filter if args.orthologs != "none" else "",
                    input_gene_col=args.input_gene_col,
                    force_all=args.force_all,
                )
                if file_estimate is not None:
                    soft_run_estimate.files.append(file_estimate)  # type: ignore[union-attr]
                    success_count += 1
                else:
                    print(f"\nError estimating {csv_path}; exiting early.")
                    sys.exit(1)
            else:
                if process_csv_file(
                    csv_path,
                    keywords_list,
                    args.reference_limit,
                    args.input_directory,
                    run_store,
                    adapter=adapter,
                    diopt_filter=args.diopt_filter if args.orthologs != "none" else "",
                    input_gene_col=args.input_gene_col,
                    force_all=args.force_all,
                ):
                    success_count += 1
                else:
                    print(f"\nError processing {csv_path}; exiting early to preserve checkpoints.")
                    sys.exit(1)
        except Exception as e:
            print(f"\nError processing {csv_path}: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    # Persist shared caches for future runs.
    log_step("Saving metadata caches")
    _save_pmid_title_abstract_cache_pending()
    if not args.soft_run:
        _save_fulltext_method_cache_pending()

    if args.soft_run:
        _print_soft_run_report(soft_run_estimate)  # type: ignore[arg-type]
    
    print(f"\n{'='*70}")
    action = "estimated" if args.soft_run else "processed"
    print(f"Complete! Successfully {action} {success_count}/{len(csv_files)} files")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
