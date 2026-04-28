#!/usr/bin/env python3
"""
Summarize fl.AI gene-set processing outputs.

Given a directory containing one or more original fly gene-set CSVs, this script
counts:
  - Unique fly genes in each original CSV.
  - Unique human / mouse genes the fly set mapped to (from sibling
    `Human/<stem>_human.csv` and `Mouse/<stem>_mouse.csv`).
  - Unique human / mouse genes assigned to each classification bucket
    (from sibling `*_classification.xlsx` workbooks).

Multi-keyword classifications are treated as their own canonical bucket: the
categories are split, trimmed, sorted case-insensitively, and joined with
", ", so `sleep; circadian` and `circadian, sleep` collapse to the same
`circadian, sleep` bucket but remain distinct from `sleep` or `circadian`
alone.

Output: a single formatted Excel sheet written to `<input_dir>/summary.xlsx`.
The sheet lays out each human / mouse run as a chronological flow:
fly input -> ortholog mapping -> classification -> bucket breakdown.

Usage:
    python summarize_gene_set.py <input_directory> [--input-gene-col ext_gene]
                                                   [--output summary.xlsx]
                                                   [--human-dir PATH]
                                                   [--mouse-dir PATH]
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


GENERATED_SUFFIXES = ("_human.csv", "_mouse.csv", "_classification.csv")
CATEGORY_SPLIT = re.compile(r"[;,]")


SPECIES_CONFIG = {
    "human": {
        "label": "Human",
        "subdir": "Human",
        "csv_suffix": "_human.csv",
        "classification_suffix": "_human_classification.xlsx",
        "entrez_col": "human_entrez_gene_id",
        "symbol_col": "human_gene_symbol",
        "classification_symbol_col": "Human_Gene_symbol",
        "classification_id_col": "Entrez_Gene_ID",
    },
    "mouse": {
        "label": "Mouse",
        "subdir": "Mouse",
        "csv_suffix": "_mouse.csv",
        "classification_suffix": "_mouse_classification.xlsx",
        "entrez_col": "mouse_entrez_gene_id",
        "symbol_col": "mouse_gene_symbol",
        "classification_symbol_col": "Mouse_Gene_symbol",
        "classification_id_col": "Entrez_Gene_ID",
    },
}


@dataclass
class SpeciesSummary:
    """Counts and per-bucket tallies for one species derived from one gene set."""

    species_label: str
    mapped_unique: int = 0
    classified_unique: int = 0
    unclassified_unique: int = 0
    buckets: dict[str, int] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)


@dataclass
class GeneSetSummary:
    """Top-level per-file summary covering fly counts and species derivations."""

    gene_set: str
    original_path: Path
    fly_unique: int = 0
    fly_notes: list[str] = field(default_factory=list)
    species: dict[str, SpeciesSummary] = field(default_factory=dict)


def is_generated_csv(path: Path) -> bool:
    name = path.name.lower()
    return any(name.endswith(suffix) for suffix in GENERATED_SUFFIXES)


def find_original_csvs(input_dir: Path) -> list[Path]:
    return sorted(
        p for p in input_dir.glob("*.csv")
        if p.is_file() and not is_generated_csv(p)
    )


def _clean(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    if text.lower() in {"", "nan", "none", "n/a", "na"}:
        return ""
    return text


def count_unique_fly_genes(
    csv_path: Path,
    input_gene_col: str,
) -> tuple[int, list[str]]:
    notes: list[str] = []
    try:
        df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    except Exception as exc:
        return 0, [f"Failed to read CSV: {exc}"]

    if df.empty:
        return 0, ["CSV is empty"]

    if "flybase_gene_id" in df.columns:
        ids = {
            _clean(v) for v in df["flybase_gene_id"]
            if _clean(v).startswith("FBgn")
        }
        unmapped = sum(1 for v in df["flybase_gene_id"] if not _clean(v).startswith("FBgn"))
        if unmapped:
            notes.append(f"{unmapped} row(s) without an FBgn ID were excluded from the fly count")
        return len(ids), notes

    if input_gene_col in df.columns:
        symbols = {_clean(v) for v in df[input_gene_col] if _clean(v)}
        notes.append(
            f"`flybase_gene_id` column missing; counted unique values in `{input_gene_col}` instead"
        )
        return len(symbols), notes

    return 0, [
        f"Neither `flybase_gene_id` nor `{input_gene_col}` columns are present"
    ]


def count_unique_mapped(
    csv_path: Path,
    species_cfg: dict,
) -> tuple[int, list[str]]:
    notes: list[str] = []
    try:
        df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    except Exception as exc:
        return 0, [f"Failed to read ortholog CSV: {exc}"]

    if df.empty:
        return 0, ["Ortholog CSV is empty"]

    entrez_col = species_cfg["entrez_col"]
    symbol_col = species_cfg["symbol_col"]
    if entrez_col in df.columns:
        ids = {_clean(v) for v in df[entrez_col] if _clean(v)}
        return len(ids), notes
    if symbol_col in df.columns:
        ids = {_clean(v) for v in df[symbol_col] if _clean(v)}
        notes.append(
            f"`{entrez_col}` column missing; counted unique values in `{symbol_col}` instead"
        )
        return len(ids), notes
    return 0, [f"Neither `{entrez_col}` nor `{symbol_col}` columns present in {csv_path.name}"]


_EMPTY_CATEGORY_TOKENS = {"", "nan", "none", "n/a", "na"}


def canonical_bucket(category_value: object) -> str:
    text = "" if category_value is None else str(category_value).strip()
    if text.lower() in _EMPTY_CATEGORY_TOKENS:
        return ""
    parts = [p.strip() for p in CATEGORY_SPLIT.split(text)]
    parts = [p for p in parts if p and p.lower() not in _EMPTY_CATEGORY_TOKENS]
    if not parts:
        return ""
    parts_sorted = sorted(set(parts), key=lambda s: s.lower())
    return ", ".join(parts_sorted)


def summarize_classification(
    workbook_path: Path,
    species_cfg: dict,
) -> tuple[int, int, dict[str, int], list[str]]:
    """Return (classified_unique, unclassified_unique, bucket_counts, notes)."""
    notes: list[str] = []
    try:
        df = pd.read_excel(
            workbook_path,
            sheet_name="Classification",
            dtype=str,
            keep_default_na=False,
        )
    except ValueError as exc:
        return 0, 0, {}, [f"`Classification` sheet missing in {workbook_path.name}: {exc}"]
    except Exception as exc:
        return 0, 0, {}, [f"Failed to read {workbook_path.name}: {exc}"]

    if df.empty:
        return 0, 0, {}, [f"`Classification` sheet in {workbook_path.name} is empty"]

    id_col = species_cfg["classification_id_col"] if species_cfg["classification_id_col"] in df.columns else None
    if id_col is None and species_cfg["classification_symbol_col"] in df.columns:
        id_col = species_cfg["classification_symbol_col"]
        notes.append(
            f"`{species_cfg['classification_id_col']}` column missing; "
            f"using `{id_col}` for unique-gene counts"
        )
    if id_col is None:
        return 0, 0, {}, [
            f"Neither `{species_cfg['classification_id_col']}` nor "
            f"`{species_cfg['classification_symbol_col']}` columns are present"
        ]
    if "Category" not in df.columns:
        return 0, 0, {}, [f"`Category` column missing from {workbook_path.name}"]

    classified: set[str] = set()
    unclassified: set[str] = set()
    bucket_to_genes: dict[str, set[str]] = {}
    for _, row in df.iterrows():
        gene_id = _clean(row.get(id_col, ""))
        if not gene_id:
            continue
        bucket = canonical_bucket(row.get("Category", ""))
        if bucket:
            classified.add(gene_id)
            bucket_to_genes.setdefault(bucket, set()).add(gene_id)
        else:
            unclassified.add(gene_id)

    unclassified -= classified
    bucket_counts = {bucket: len(genes) for bucket, genes in bucket_to_genes.items()}
    return len(classified), len(unclassified), bucket_counts, notes


def resolve_species_dir(
    input_dir: Path,
    species_cfg: dict,
    override: Path | None,
) -> Path | None:
    if override is not None:
        return override
    candidate = input_dir / species_cfg["subdir"]
    if candidate.is_dir():
        return candidate
    return None


def summarize_species_for_file(
    stem: str,
    species_cfg: dict,
    species_dir: Path | None,
) -> SpeciesSummary:
    summary = SpeciesSummary(species_label=species_cfg["label"])
    if species_dir is None:
        summary.notes.append(f"`{species_cfg['subdir']}/` folder not found; skipping {species_cfg['label']}")
        return summary

    csv_path = species_dir / f"{stem}{species_cfg['csv_suffix']}"
    if csv_path.exists():
        mapped, mapped_notes = count_unique_mapped(csv_path, species_cfg)
        summary.mapped_unique = mapped
        summary.notes.extend(mapped_notes)
    else:
        summary.notes.append(f"Ortholog CSV not found: {csv_path}")

    workbook_path = species_dir / f"{stem}{species_cfg['classification_suffix']}"
    if workbook_path.exists():
        classified, unclassified, buckets, class_notes = summarize_classification(
            workbook_path, species_cfg
        )
        summary.classified_unique = classified
        summary.unclassified_unique = unclassified
        summary.buckets = buckets
        summary.notes.extend(class_notes)
    else:
        summary.notes.append(f"Classification workbook not found: {workbook_path}")

    return summary


def build_summary(
    input_dir: Path,
    *,
    input_gene_col: str,
    human_dir: Path | None,
    mouse_dir: Path | None,
) -> list[GeneSetSummary]:
    originals = find_original_csvs(input_dir)
    if not originals:
        return []

    resolved = {
        "human": resolve_species_dir(input_dir, SPECIES_CONFIG["human"], human_dir),
        "mouse": resolve_species_dir(input_dir, SPECIES_CONFIG["mouse"], mouse_dir),
    }

    results: list[GeneSetSummary] = []
    for csv_path in originals:
        stem = csv_path.stem
        gene_set = GeneSetSummary(gene_set=stem, original_path=csv_path)
        gene_set.fly_unique, gene_set.fly_notes = count_unique_fly_genes(csv_path, input_gene_col)
        for species_key, cfg in SPECIES_CONFIG.items():
            gene_set.species[species_key] = summarize_species_for_file(
                stem, cfg, resolved[species_key]
            )
        results.append(gene_set)
    return results


def _format_bucket_breakdown(species: SpeciesSummary) -> str:
    if not species.buckets:
        return "No classified bucket counts available"
    sorted_buckets = sorted(
        species.buckets.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )
    return "\n".join(f"{bucket}: {count}" for bucket, count in sorted_buckets)


def _flow_rows(summaries: list[GeneSetSummary]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for gene_set in summaries:
        for species_key in ("human", "mouse"):
            species = gene_set.species[species_key]
            rows.append({
                "Gene Set": gene_set.gene_set,
                "Pipeline": f"Fly -> {species.species_label}",
                "Step 1: Original Fly Gene Set": f"{gene_set.fly_unique} unique fly genes",
                "Step 2: Ortholog Mapping": (
                    f"{species.mapped_unique} unique {species.species_label.lower()} genes mapped"
                ),
                "Step 3: Keyword / Subset Classification": (
                    f"{species.classified_unique} classified into bucket(s)\n"
                    f"{species.unclassified_unique} unclassified / no bucket"
                ),
                "Step 4: Classified Buckets": _format_bucket_breakdown(species),
            })
    return rows


def write_summary_workbook(rows: list[dict[str, object]], output_path: Path) -> None:
    columns = [
        "Gene Set",
        "Pipeline",
        "Step 1: Original Fly Gene Set",
        "Step 2: Ortholog Mapping",
        "Step 3: Keyword / Subset Classification",
        "Step 4: Classified Buckets",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    title_fill = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    human_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    mouse_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    body_font = Font(size=13)
    count_font = Font(size=14, bold=True)
    bucket_font = Font(size=13, bold=True)
    thin_side = Side(border_style="thin", color="BFBFBF")
    medium_side = Side(border_style="medium", color="7F7F7F")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Gene Set Summary: Fly to Human / Mouse Pipeline"
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append(columns)
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = cell_border

    last_gene_set: str | None = None
    for row_data in rows:
        ws.append([row_data[col] for col in columns])
        excel_row = ws.max_row
        pipeline = str(row_data.get("Pipeline", ""))
        species_fill = human_fill if "Human" in pipeline else mouse_fill
        gene_set_changed = last_gene_set is not None and row_data["Gene Set"] != last_gene_set
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = cell_border
            cell.fill = species_fill
            cell.font = count_font if col_name.startswith("Step") else body_font
            if col_name == "Step 4: Classified Buckets":
                cell.font = bucket_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if gene_set_changed:
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=medium_side,
                    bottom=thin_side,
                )
        ws.row_dimensions[excel_row].height = 92
        last_gene_set = row_data["Gene Set"]

    widths = {
        "Gene Set": 32,
        "Pipeline": 20,
        "Step 1: Original Fly Gene Set": 26,
        "Step 2: Ortholog Mapping": 30,
        "Step 3: Keyword / Subset Classification": 34,
        "Step 4: Classified Buckets": 36,
    }
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]

    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "A3"
    if ws.max_row > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{ws.max_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Summarize fly gene-set processing outputs. Counts original unique "
            "fly genes, mapped human/mouse orthologs, and per-bucket "
            "classification totals."
        )
    )
    parser.add_argument(
        "input_directory",
        help="Directory containing the original fly gene-set CSV(s).",
    )
    parser.add_argument(
        "--input-gene-col",
        default="ext_gene",
        help=(
            "Symbol column to fall back on when an original CSV lacks "
            "`flybase_gene_id` (default: ext_gene)."
        ),
    )
    parser.add_argument(
        "--output",
        default="summary.xlsx",
        help="Output filename, written next to the original gene set (default: summary.xlsx).",
    )
    parser.add_argument(
        "--human-dir",
        default=None,
        help="Override path to the directory containing `*_human.csv` and `*_human_classification.xlsx`.",
    )
    parser.add_argument(
        "--mouse-dir",
        default=None,
        help="Override path to the directory containing `*_mouse.csv` and `*_mouse_classification.xlsx`.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    input_dir = Path(args.input_directory).expanduser().resolve()
    if not input_dir.is_dir():
        print(f"Error: input directory does not exist: {input_dir}", file=sys.stderr)
        return 1

    human_dir = Path(args.human_dir).expanduser().resolve() if args.human_dir else None
    mouse_dir = Path(args.mouse_dir).expanduser().resolve() if args.mouse_dir else None

    summaries = build_summary(
        input_dir,
        input_gene_col=args.input_gene_col,
        human_dir=human_dir,
        mouse_dir=mouse_dir,
    )
    if not summaries:
        print(f"No original gene-set CSVs found in {input_dir}", file=sys.stderr)
        return 1

    rows = _flow_rows(summaries)
    output_path = (input_dir / args.output) if not Path(args.output).is_absolute() else Path(args.output)
    write_summary_workbook(rows, output_path)

    print(f"Summary written: {output_path}")
    for gene_set in summaries:
        human = gene_set.species["human"]
        mouse = gene_set.species["mouse"]
        print(
            f"  {gene_set.gene_set}: fly={gene_set.fly_unique} | "
            f"human mapped={human.mapped_unique}, classified={human.classified_unique}, "
            f"buckets={len(human.buckets)} | "
            f"mouse mapped={mouse.mapped_unique}, classified={mouse.classified_unique}, "
            f"buckets={len(mouse.buckets)}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
